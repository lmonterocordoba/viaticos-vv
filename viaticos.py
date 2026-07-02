#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Viáticos — Villar & Villar Abogados
Roles: trabajador (ve solo su pago) | administradora (ve todo, valida)
"""

import os, re, shutil, json, threading, webbrowser, hashlib, uuid, unicodedata
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template_string, request, jsonify,
                   send_file, session, redirect, url_for)
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# ─── RUTAS ────────────────────────────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_XLS = os.path.join(BASE_DIR, "Solicitud y Comprobación de Gastos Villar Abogados - 2026.xlsx")
CALENDAR     = os.path.join(BASE_DIR, "Viáticos Liogero", "Calendario de viajes.xlsx")
EMAIL_CONFIG_F  = os.path.join(BASE_DIR, "email_config.json")
CASETAS_F       = os.path.join(BASE_DIR, "casetas.json")

# En Railway: DATA_DIR=/data (volumen persistente). En local: usa rutas conocidas.
_DATA_DIR = os.environ.get('DATA_DIR', None)
if _DATA_DIR:
    USUARIOS_F  = os.path.join(_DATA_DIR, "usuarios.json")
    SUBS_DIR    = os.path.join(_DATA_DIR, "submissions")
    UPLOADS_DIR = os.path.join(_DATA_DIR, "uploads")
    os.makedirs(_DATA_DIR, exist_ok=True)
    # Sembrar usuarios.json desde el repo si el volumen está vacío
    _seed = os.path.join(BASE_DIR, "usuarios.json")
    if not os.path.exists(USUARIOS_F) and os.path.exists(_seed):
        import shutil as _sh; _sh.copy2(_seed, USUARIOS_F)
else:
    USUARIOS_F  = os.path.expanduser("~/viaticos_data/usuarios.json")
    SUBS_DIR    = os.path.join(BASE_DIR, "submissions")
    UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")

os.makedirs(SUBS_DIR,   exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)

# ─── REGLAS DE NEGOCIO ────────────────────────────────────────────────────────

EMPLOYEE_KM_RATE    = 2.60
EMPLOYEE_MEAL_RATE  = 290.00
EMPLOYEE_HOTEL_RATE = 2100.00
CLIENT_HOTEL_MAX    = 2500.00
CLIENT_MEAL_RATE    = 370.00

# Grupos de tabulador para cobro al cliente.
# Liogero y Benavides comparten el "Tabulador Villar Abogados 2026".
# Los demás clientes usan el "Tabulador Querétaro 2026" de su propia carpeta.
# El sistema carga el PDF de tabulador que esté en la carpeta del cliente;
# si un cliente del grupo VV no tiene PDF propio, cae al tabulador de Liogero.
TABULADOR_GRUPO_VV       = {"Viáticos Liogero", "Viáticos Benavides"}
TABULADOR_GRUPO_VV_REF   = "Viáticos Liogero"   # carpeta de referencia para el grupo VV

def _get_tabulador(cliente_folder_name):
    """Carga el tabulador correcto según el grupo del cliente.
    cliente_folder_name: ej. 'Viáticos Liogero', 'Viáticos INSECOM'."""
    from_dirs = []
    if nfc(cliente_folder_name) in {nfc(x) for x in TABULADOR_GRUPO_VV}:
        # Grupo Villar & Villar — usar la carpeta de Liogero como referencia
        from_dirs.append(os.path.join(BASE_DIR, TABULADOR_GRUPO_VV_REF))
    # También revisar la carpeta propia del cliente
    client_own = find_dir(BASE_DIR, cliente_folder_name)
    if client_own:
        from_dirs.insert(0, client_own)   # prioridad a la carpeta propia
    tabulador = {}
    for d in from_dirs:
        if os.path.isdir(d):
            for f in os.listdir(d):
                if f.lower().endswith(".pdf"):
                    tabulador.update(parse_tabulador(os.path.join(d, f)))
    return tabulador

def meals_by_distance(km):
    """km debe ser el km_total (distancia real recorrida, ida y vuelta + ciudad)."""
    if km <= 200: return 1
    if km <= 400: return 2
    return 3

# ─── UTILIDADES ───────────────────────────────────────────────────────────────

def nfc(s):
    return unicodedata.normalize("NFC", s)

def safe_name(s):
    """Nombre de carpeta seguro: sin acentos, sin caracteres especiales."""
    s = ''.join(c for c in unicodedata.normalize('NFD', nfc(str(s)))
                if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s\-]', '', s).strip()
    s = re.sub(r'\s+', '_', s)
    return s

def get_upload_dir(usuario, cliente, viaje_key):
    """Carpeta de comprobantes subidos: uploads/{usuario}/{cliente}/{viaje_key}/"""
    d = os.path.join(UPLOADS_DIR,
                     safe_name(usuario),
                     safe_name(cliente.replace("Viáticos ", "")),
                     safe_name(viaje_key))
    os.makedirs(d, exist_ok=True)
    return d

def find_dir(parent, name):
    name_nfc = nfc(name)   # normalizar ambos lados antes de comparar
    for d in os.listdir(parent):
        if nfc(d) == name_nfc and os.path.isdir(os.path.join(parent, d)):
            return os.path.join(parent, d)
    return None

def hp(password):
    return hashlib.sha256(password.encode()).hexdigest()

# ─── USUARIOS ─────────────────────────────────────────────────────────────────

def load_users():
    if not os.path.exists(USUARIOS_F):
        return {}
    import time
    for attempt in range(5):
        try:
            with open(USUARIOS_F, encoding="utf-8") as f:
                return json.load(f)
        except (PermissionError, OSError):
            time.sleep(0.3)
    return {}

# ─── EMPLEADOS QUE REQUIEREN AUTORIZACIÓN DE SOCIA ────────────────────────────

EMPLOYEES_REQUIRING_SOCIA = {"iescobar", "kruiz"}

# ─── EMAIL ────────────────────────────────────────────────────────────────────

# Datos de casetas (costo ida y vuelta) desde Ciudad de México, según CAPUFE.
# Claves en MAYÚSCULAS sin acentos. El admin puede editarlos en /admin/casetas.
CASETAS_DEFAULT = {
    "AGUASCALIENTES":   1060.0,
    "CELAYA":            480.0,
    "CHIHUAHUA":        2000.0,
    "CIUDAD JUAREZ":    2200.0,
    "COLIMA":           1100.0,
    "CULIACAN":         1800.0,
    "DURANGO":          1400.0,
    "GUADALAJARA":       880.0,
    "GUANAJUATO":        600.0,
    "HERMOSILLO":       2200.0,
    "IRAPUATO":          480.0,
    "LEON":              520.0,
    "MAZATLAN":         1600.0,
    "MERIDA":           1400.0,
    "MONTERREY":        1200.0,
    "MORELIA":           560.0,
    "PACHUCA":           160.0,
    "PUEBLA":            280.0,
    "QUERETARO":         400.0,
    "SAN LUIS POTOSI":   800.0,
    "TAMPICO":           900.0,
    "TLAXCALA":          300.0,
    "TOLUCA":            160.0,
    "TORREON":          1400.0,
    "TUXTLA GUTIERREZ": 1200.0,
    "VERACRUZ":          680.0,
    "VILLAHERMOSA":     1200.0,
    "XALAPA":            640.0,
    "ZACATECAS":        1000.0,
}

def load_casetas():
    if not os.path.exists(CASETAS_F):
        return dict(CASETAS_DEFAULT)
    try:
        with open(CASETAS_F, encoding="utf-8") as f:
            data = json.load(f)
        # Completar con defaults para ciudades no guardadas aún
        for k, v in CASETAS_DEFAULT.items():
            data.setdefault(k, v)
        return data
    except Exception:
        return dict(CASETAS_DEFAULT)

def save_casetas(data):
    with open(CASETAS_F, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def lookup_casetas(ciudad):
    """Busca el costo de casetas (ida y vuelta) para una ciudad."""
    casetas = load_casetas()
    city_norm = strip_accents(ciudad.upper().strip())
    for key, val in casetas.items():
        if strip_accents(key) == city_norm or city_norm in strip_accents(key) or strip_accents(key) in city_norm:
            return val
    return 0.0

def load_email_config():
    if not os.path.exists(EMAIL_CONFIG_F):
        return {}
    try:
        with open(EMAIL_CONFIG_F, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def send_email(to_list, subject, body_html):
    """Envía un correo HTML. Retorna True en éxito, False en error. Nunca lanza excepción."""
    cfg = load_email_config()
    required = {"smtp_server", "smtp_port", "smtp_user", "smtp_password", "from_email"}
    if not cfg or not required.issubset(cfg.keys()):
        print(f"[email] Configuración incompleta — no se envió: {subject}")
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{cfg.get('from_name','Viáticos V&V')} <{cfg['from_email']}>"
        msg["To"]      = ", ".join(to_list)
        msg.attach(MIMEText(body_html, "html", "utf-8"))
        with smtplib.SMTP(cfg["smtp_server"], int(cfg["smtp_port"])) as s:
            s.ehlo()
            s.starttls()
            s.login(cfg["smtp_user"], cfg["smtp_password"])
            s.sendmail(cfg["from_email"], to_list, msg.as_string())
        print(f"[email] Enviado OK → {to_list}: {subject}")
        return True
    except Exception as e:
        print(f"[email] Error enviando '{subject}': {e}")
        return False

# ─── TABULADOR ────────────────────────────────────────────────────────────────

def parse_tabulador(pdf_path):
    rates = {}
    if not os.path.exists(pdf_path):
        return rates
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    m = re.match(
                        r"^([A-ZÁÉÍÓÚÑÜ ,\.\-]+?)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+\$([\d,]+\.?\d*)",
                        line.strip())
                    if m:
                        city = m.group(1).strip().upper()
                        rates[city] = {
                            "km":          float(m.group(2).replace(",","")),
                            "km_total":    float(m.group(3).replace(",","")),
                            "combustible": float(m.group(4).replace(",",""))
                        }
    except Exception as e:
        print(f"[tabulador] {e}")
    return rates

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")

def lookup_city(city, tabulador):
    """Busca la ciudad en el tabulador.
    Prioridad: 1) coincidencia exacta, 2) la ciudad buscada es la primera
    parte del nombre (antes de la coma), 3) coincidencia parcial."""
    def norm(s):
        return strip_accents(s.upper().replace(",","").replace(".","").replace(" ","").replace("-",""))

    city_norm = norm(city)
    # Primera parte del nombre de búsqueda (sin estado)
    city_first = norm(city.split(",")[0].strip())

    exact = None
    starts = None
    partial = None

    for key, val in tabulador.items():
        key_norm   = norm(key)
        key_first  = norm(key.split(",")[0].strip())   # ej. "GUANAJUATO" de "GUANAJUATO, GUANAJUATO"

        if key_norm == city_norm:
            exact = val; break
        if key_first == city_first and starts is None:
            starts = val
        if (city_norm in key_norm or key_norm in city_norm) and partial is None:
            partial = val

    return exact or starts or partial

def _city_data_fallback(km_total, casetas_mapa=0.0):
    """Construye un city_data sintético a partir de km calculados por el mapa."""
    if not km_total:
        return None
    combustible = round(km_total * EMPLOYEE_KM_RATE, 2)
    return {
        "km":          round(km_total / 2, 1),
        "km_total":    round(km_total, 1),
        "combustible": combustible,
        "casetas":     casetas_mapa,
    }

# ─── EXTRACCIÓN DE COMPROBANTES ───────────────────────────────────────────────

AMOUNT_PATTERNS = [
    r"(?:Importe Total|Total)\s+MXN\s*\$?\s*([\d,]+\.?\d*)",
    r"(?:PRECIO TOTAL|TOTAL|Total)[/:\s]+\$?\s*([\d,]+\.?\d*)",
    r"\$\s*([\d,]+\.\d{2})\s*(?:\n|$)",
    r"MXN\s*([\d,]+\.?\d{2})",
]
RFC_PAT = re.compile(r"R\.?F\.?C\.?\s*:?\s*([A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3})")

# Ruta al binario OCR (Vision framework de macOS)
OCR_BIN = os.path.join(BASE_DIR, "ocr_pdf")

def _extract_text_ocr(pdf_path):
    """Usa el binario Swift (Vision framework) para OCR en PDFs escaneados."""
    if not os.path.exists(OCR_BIN):
        return ""
    try:
        import subprocess
        r = subprocess.run([OCR_BIN, pdf_path], capture_output=True, text=True, timeout=15)
        return r.stdout
    except Exception as e:
        print(f"[ocr] {os.path.basename(pdf_path)}: {e}")
        return ""

def extract_pdf_data(pdf_path):
    result = {"amount": 0.0, "rfc": ""}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        # Si pdfplumber no extrajo texto, intentar OCR
        if not text.strip():
            text = _extract_text_ocr(pdf_path)
        m = RFC_PAT.search(text)
        if m:
            result["rfc"] = m.group(1)
        amounts = []
        for pat in AMOUNT_PATTERNS:
            for raw in re.findall(pat, text, re.IGNORECASE | re.MULTILINE):
                try: amounts.append(float(raw.replace(",","")))
                except: pass
        if amounts:
            result["amount"] = max(amounts)
    except Exception as e:
        print(f"[pdf] {os.path.basename(pdf_path)}: {e}")
    return result

def detect_category(filename):
    f = filename.lower()
    # Transporte PRIMERO: evita falsos positivos con "hotel" en rutas de traslado
    if any(x in f for x in ["uber","didi","taxi","indrive","cabify",
                             "traslado","transporte"]):               return "Transporte"
    if any(x in f for x in ["autobus","autobús","camion","camión",
                             "pullman","ado"]):                       return "Autobús"
    if any(x in f for x in ["vuelo","aéreo","aereo","boleto avio"]): return "Vuelo"
    if any(x in f for x in ["comida","alimento","restaur","snack",
                             "cena","desayuno"]):                     return "Comida"
    if "caseta" in f:                                                 return "Casetas"
    if any(x in f for x in ["gasolina","combustible"]):              return "Gasolina"
    if any(x in f for x in ["hotel","hotal","hospedaje","motel"]):   return "Hotel"
    return "Otro"

# ─── CALENDARIO ───────────────────────────────────────────────────────────────

def load_calendar(cal_file=None):
    cal_path = cal_file if cal_file else CALENDAR
    if not os.path.exists(cal_path):
        return []
    # Busca el calendario en la carpeta del cliente primero, luego en Liogero
    cal_path = cal_file if cal_file else CALENDAR
    if not os.path.exists(cal_path):
        return []
    df = pd.read_excel(cal_path, header=1)
    # Intentar mapear por nombre de columna primero; fallback por posición
    cols = list(df.columns)
    cols_lower = [str(c).lower() for c in cols]

    def _find_col(keywords):
        for kw in keywords:
            for i, c in enumerate(cols_lower):
                if kw in c:
                    return cols[i]
        return None

    ci  = _find_col(["inicio"])
    cf  = _find_col(["terminac", "fin", "regres"])
    ciu = _find_col(["ciudad"])
    est = _find_col(["estado"])
    mot = _find_col(["motivo"])
    abo = _find_col(["abogado"])
    via = _find_col(["viaje"])

    if all([ci, cf, ciu, est, mot, abo, via]):
        col_map = {ci:"FechaInicio", cf:"FechaFin", ciu:"Ciudad",
                   est:"Estado",     mot:"Motivo",   abo:"Abogado", via:"Viaje"}
    elif len(cols) >= 7:
        col_map = {cols[0]:"FechaInicio", cols[1]:"FechaFin", cols[2]:"Ciudad",
                   cols[3]:"Estado",      cols[4]:"Motivo",   cols[5]:"Abogado",
                   cols[6]:"Viaje"}
    else:
        col_map = {cols[0]:"FechaInicio", cols[1]:"Ciudad", cols[2]:"Estado",
                   cols[3]:"Motivo",      cols[4]:"Abogado", cols[5]:"Viaje"}
    df = df.rename(columns=col_map)
    df = df[df["Viaje"].astype(str).str.strip().str.upper() == "SI"].copy()
    df["FechaInicio"] = pd.to_datetime(df["FechaInicio"], errors="coerce")
    df["FechaFin"]    = pd.to_datetime(df.get("FechaFin", df["FechaInicio"]), errors="coerce")
    records = []
    for _, r in df.iterrows():
        fi = r["FechaInicio"]
        ff = r["FechaFin"] if pd.notna(r.get("FechaFin")) else fi
        records.append({
            "fecha_inicio": fi.strftime("%d/%m/%Y") if pd.notna(fi) else "",
            "fecha_fin":    ff.strftime("%d/%m/%Y") if pd.notna(ff) else "",
            "ciudad":  nfc(str(r["Ciudad"]))  if pd.notna(r["Ciudad"])  else "",
            "estado":  str(r["Estado"])       if pd.notna(r["Estado"])  else "",
            "motivo":  str(r["Motivo"])       if pd.notna(r["Motivo"])  else "",
            "abogado": str(r["Abogado"])      if pd.notna(r["Abogado"]) else "",
        })
    return records

def lookup_calendar(city, cal_records, hint_dates=None):
    """
    Devuelve un dict con datos del calendario para una ciudad.
    hint_dates: lista de strings "DD/MM/YYYY" extraídas de los comprobantes
    del viaje — se usa para desambiguar cuando hay varios viajes a la misma ciudad.
    """
    def parse_fecha(f):
        try: return datetime.strptime(f, "%d/%m/%Y")
        except: return None

    # strip_accents definida a nivel de módulo

    city_plain = strip_accents(nfc(city)).upper()
    matches = [r for r in cal_records
               if city_plain in strip_accents(nfc(r["ciudad"])).upper()
               or strip_accents(nfc(r["ciudad"])).upper() in city_plain]
    if not matches:
        return None

    # Si hay hint_dates, filtrar por el grupo de fechas más cercano
    if hint_dates and len(matches) > 1:
        hint_parsed = [parse_fecha(f) for f in hint_dates]
        hint_parsed = [f for f in hint_parsed if f]
        if hint_parsed:
            hint_min = min(hint_parsed)
            hint_max = max(hint_parsed)

            def overlap_score(r):
                ri = parse_fecha(r["fecha_inicio"])
                rf = parse_fecha(r["fecha_fin"]) or ri
                if not ri: return 999
                # Distancia mínima entre rangos
                if rf < hint_min: return (hint_min - rf).days
                if ri > hint_max: return (ri - hint_max).days
                return 0   # solapan → puntuación 0 (mejor)

            best_score = min(overlap_score(r) for r in matches)
            matches = [r for r in matches if overlap_score(r) == best_score]

    # Agrupar por fecha de inicio para no mezclar viajes distintos
    # (dos entradas con la misma fecha = mismo viaje con múltiples asuntos)
    fechas_inicio_unicas = sorted(set(r["fecha_inicio"] for r in matches if r["fecha_inicio"]))
    if len(fechas_inicio_unicas) > 1:
        # Tomar el grupo de fecha más reciente entre los filtrados
        fecha_principal = fechas_inicio_unicas[-1]
        matches = [r for r in matches if r["fecha_inicio"] == fecha_principal]

    # Fechas del grupo resultante
    fechas_i = [parse_fecha(r["fecha_inicio"]) for r in matches if r["fecha_inicio"]]
    fechas_f = [parse_fecha(r["fecha_fin"])    for r in matches if r["fecha_fin"]]
    fechas_i = [f for f in fechas_i if f]
    fechas_f = [f for f in fechas_f if f] or fechas_i

    salida  = min(fechas_i).strftime("%d/%m/%Y") if fechas_i else ""
    regreso = max(fechas_f).strftime("%d/%m/%Y") if fechas_f else salida

    # Combinar motivos únicos del grupo
    motivos, seen = [], set()
    for r in matches:
        m = r["motivo"].strip()
        if m and m not in seen:
            motivos.append(m); seen.add(m)

    return {
        "salida":      salida,
        "regreso":     regreso,
        "lugar":       f'{matches[0]["ciudad"]}, {matches[0]["estado"]}',
        "abogado":     matches[0]["abogado"],
        "descripcion": "\n".join(motivos),
        "motivos":     motivos,
    }

# ─── CÁLCULO POR RENGLÓN ──────────────────────────────────────────────────────

def calc_row(categoria, amount, city_data, meals_override=None):
    if categoria == "Hotel":
        emp = min(amount, EMPLOYEE_HOTEL_RATE) if amount > 0 else EMPLOYEE_HOTEL_RATE
        cli = min(amount, CLIENT_HOTEL_MAX)    if amount > 0 else CLIENT_HOTEL_MAX
        return emp, cli
    elif categoria == "Comida":
        n = int(meals_override) if meals_override is not None else (
            meals_by_distance(city_data["km_total"]) if city_data else 1)
        emp = min(n * EMPLOYEE_MEAL_RATE, amount) if amount > 0 else n * EMPLOYEE_MEAL_RATE
        return emp, n * CLIENT_MEAL_RATE
    elif categoria == "Gasolina":
        emp = (city_data["km_total"] * EMPLOYEE_KM_RATE) if city_data else 0.0
        cli = city_data["combustible"] if city_data else amount
        return emp, cli
    else:
        return amount, amount

# ─── ESCRITURA DEL EXCEL ──────────────────────────────────────────────────────

def write_excel(output_path, data):
    from openpyxl.styles import Font, Color

    def set_val(cell, value, fmt=None):
        """Escribe valor con fuente negra explícita (sobreescribe color de tema del template)."""
        cell.value = value
        f = cell.font
        cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                         italic=f.italic, underline=f.underline,
                         color=Color(rgb="FF000000", type="rgb"))
        if fmt:
            cell.number_format = fmt

    print(f"[write_excel] gastos recibidos: {len(data.get('gastos',[]))} items")
    for g in data.get('gastos', []):
        print(f"  tipo={g.get('tipo')} con_iva={g.get('con_iva')} sin_iva={g.get('sin_iva')} empleado={g.get('empleado')}")

    shutil.copy2(TEMPLATE_XLS, output_path)
    wb = load_workbook(output_path, keep_links=False)
    fix_external_refs(wb)
    ws = wb["Comprobación y Autorización"]
    wb.active = ws
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    set_val(ws["A8"],  data.get("salida", ""))
    set_val(ws["E8"],  data.get("regreso", ""))
    set_val(ws["B10"], data.get("beneficiario", ""))
    set_val(ws["B12"], data.get("cargo", ""))
    set_val(ws["B14"], data.get("asunto", ""))
    ws["B17"] = 0
    ws["D17"] = 0
    set_val(ws["B19"], data.get("lugar", ""))
    set_val(ws["A21"], "DESCRIPCIÓN DEL VIAJE:\n" + data.get("descripcion", ""))

    wd = wb["Detalle de Gastos"]
    # Limpiar celdas de datos del template
    for r in range(8, 18):
        for col in range(1, 11):
            c = wd.cell(r, col)
            if isinstance(c.value, str) and c.value in ('0', ''):
                c.value = None
            elif c.value == 0:
                c.value = None

    # Calcular totales con TODOS los gastos (sin límite de filas)
    total_con = total_sin = total_ret = total_emp = 0.0
    for g in data.get("gastos", []):
        con  = float(g.get("con_iva",  0) or 0)
        sin_ = float(g.get("sin_iva",  0) or 0)
        emp  = float(g.get("empleado", 0) or 0)
        iva  = round(con * 0.16, 2)
        ret  = round(emp - con - sin_ - iva, 2)
        total_con += con; total_sin += sin_
        total_ret += ret; total_emp += emp

    # Escribir detalle solo en las filas disponibles (8-17 = 10 filas)
    for i, g in enumerate(data.get("gastos", [])[:10]):
        row  = 8 + i
        con  = float(g.get("con_iva",  0) or 0)
        sin_ = float(g.get("sin_iva",  0) or 0)
        emp  = float(g.get("empleado", 0) or 0)
        iva  = round(con * 0.16, 2)
        ret  = round(emp - con - sin_ - iva, 2)
        for col, v in [(1, g.get("tipo","")),
                       (2, g.get("comprobante", g.get("archivo",""))),
                       (3, g.get("rfc","")),
                       (5, g.get("descripcion","")),
                       (6, con  or None),
                       (7, sin_ or None),
                       (8, iva  or None),
                       (9, ret  if ret != 0 else None),
                       (10, round(emp, 2))]:
            set_val(wd.cell(row, col), v,
                    fmt='#,##0.00' if col >= 6 and v is not None else None)

    total_iva = round(total_con * 0.16, 2)
    for coord, v in [("F18", round(total_con,2)), ("G18", round(total_sin,2)),
                     ("H18", total_iva), ("I18", round(total_ret,2)),
                     ("J18", round(total_emp,2))]:
        set_val(wd[coord], v, fmt='#,##0.00')

    for coord, v in [("B28", round(total_con,2)), ("C28", round(total_sin,2)),
                     ("D28", total_iva), ("F28", round(total_ret,2)),
                     ("G28", round(total_emp,2))]:
        set_val(ws[coord], v, fmt='#,##0.00')

    wb.save(output_path)
    return output_path

# ─── SUBMISSIONS ──────────────────────────────────────────────────────────────

def save_submission(data):
    sid = str(uuid.uuid4())[:8]
    data["id"]           = sid
    data["fecha_envio"]  = datetime.now().strftime("%d/%m/%Y %H:%M")
    data["status"]       = "pendiente"
    data["comentario"]   = ""
    path = os.path.join(SUBS_DIR, f"{sid}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return sid

def list_submissions(empleado_filter=None):
    subs = []
    for fname in sorted(os.listdir(SUBS_DIR), reverse=True):
        if not fname.endswith(".json"):
            continue
        try:
            with open(os.path.join(SUBS_DIR, fname), encoding="utf-8") as f:
                s = json.load(f)
            if empleado_filter and s.get("beneficiario") != empleado_filter:
                continue
            subs.append(s)
        except:
            pass
    return subs

def get_submission(sid):
    path = os.path.join(SUBS_DIR, f"{sid}.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def update_submission(sid, updates):
    path = os.path.join(SUBS_DIR, f"{sid}.json")
    data = get_submission(sid)
    if not data:
        return False
    data.update(updates)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return True

# ─────────────────────────────────────────────────────────────────────────────
#  FLASK APP
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = "vv-viaticos-2026-secret"
app.config["PERMANENT_SESSION_LIFETIME"] = __import__("datetime").timedelta(days=30)
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("rol") != "admin":
            return jsonify({"error": "Acceso restringido"}), 403
        return f(*args, **kwargs)
    return decorated

# ─── PÁGINAS ──────────────────────────────────────────────────────────────────

LOGIN_HTML = """<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Viáticos V&V — Acceso</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:linear-gradient(135deg,#1a3a5c,#2e86ab);
     min-height:100vh;display:flex;align-items:center;justify-content:center}
.box{background:white;border-radius:12px;padding:40px;width:360px;
     box-shadow:0 8px 32px rgba(0,0,0,.3)}
h2{color:#1a3a5c;margin-bottom:6px;font-size:20px}
p{color:#777;font-size:13px;margin-bottom:24px}
label{display:block;font-size:12px;color:#555;margin-bottom:4px}
input{width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;
      font-size:14px;margin-bottom:14px}
button{width:100%;padding:12px;background:#1a3a5c;color:white;border:none;
       border-radius:6px;font-size:14px;font-weight:bold;cursor:pointer}
button:hover{background:#2e5484}
.err{background:#fde;color:#c00;padding:10px;border-radius:6px;
     font-size:13px;margin-bottom:14px;text-align:center}
.logo{text-align:center;margin-bottom:20px;font-size:28px}
</style></head>
<body>
<div class="box">
  <div class="logo">⚖️</div>
  <h2>Villar &amp; Villar Abogados</h2>
  <p>Sistema de Viáticos 2026</p>
  {% if error %}<div class="err">{{ error }}</div>{% endif %}
  <form method="POST">
    <label>Usuario</label>
    <input name="usuario" placeholder="ej. bmorales" autocomplete="username" required>
    <label>Contraseña</label>
    <input name="password" type="password" autocomplete="current-password" required>
    <button type="submit">Ingresar →</button>
  </form>
</div>
</body></html>"""

@app.route("/", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        u = request.form.get("usuario","").strip().lower()
        p = request.form.get("password","")
        users = load_users()
        if u in users and users[u]["password"] == hp(p):
            session.permanent = True   # sesión dura 30 días
            session["usuario"] = u
            session["nombre"]  = users[u]["nombre"]
            session["rol"]     = users[u]["rol"]
            return redirect(url_for("dashboard"))
        error = "Usuario o contraseña incorrectos."
    return render_template_string(LOGIN_HTML, error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── DASHBOARD ────────────────────────────────────────────────────────────────

def _contar_viajes_por_status(usuario_filter=None):
    """Cuenta viajes en la carpeta uploads según el status de _aprobacion.json.
    Devuelve dict {pendiente, aprobado, aprobado_parcial, rechazado, total}."""
    conteo = {"pendiente": 0, "aprobado": 0, "aprobado_parcial": 0,
              "rechazado": 0, "total": 0}
    users = load_users()
    for usr in (([usuario_filter] if usuario_filter else list(users.keys()))):
        if users.get(usr, {}).get("rol") in ("admin", "socia"):
            continue
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
        if not os.path.isdir(user_dir):
            continue
        for cli_safe in os.listdir(user_dir):
            cli_dir = os.path.join(user_dir, cli_safe)
            if not os.path.isdir(cli_dir):
                continue
            for vk in os.listdir(cli_dir):
                trip_dir = os.path.join(cli_dir, vk)
                if not os.path.isdir(trip_dir):
                    continue
                archivos = [f for f in os.listdir(trip_dir)
                            if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))]
                if not archivos:
                    continue
                apro = _load_aprobacion(trip_dir)
                st   = apro.get("status", "pendiente")
                conteo[st] = conteo.get(st, 0) + 1
                conteo["total"] += 1
    return conteo


@app.route("/dashboard")
@login_required
def dashboard():
    rol      = session["rol"]
    nombre   = session["nombre"]
    usuario  = session["usuario"]

    # Submissions formales (anticipos y comprobaciones enviadas)
    if rol == "admin":
        subs = list_submissions()
    else:
        subs = list_submissions(empleado_filter=nombre)

    pendientes_sub = [s for s in subs if s["status"] == "pendiente"]
    aprobados_sub  = [s for s in subs if s["status"] == "aprobado"]
    rechazados_sub = [s for s in subs if s["status"] == "rechazado"]

    # Conteo de viajes por status de aprobación (desde uploads)
    if rol == "admin":
        conteo_viajes = _contar_viajes_por_status()
    else:
        conteo_viajes = _contar_viajes_por_status(usuario_filter=usuario)

    # Pendientes de autorización socia
    pendientes_socia = 0
    if rol == "socia":
        users = load_users()
        for usr in EMPLOYEES_REQUIRING_SOCIA:
            user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
            if not os.path.isdir(user_dir):
                continue
            for cli_safe in os.listdir(user_dir):
                cli_dir = os.path.join(user_dir, cli_safe)
                if not os.path.isdir(cli_dir):
                    continue
                for vk in os.listdir(cli_dir):
                    trip_dir = os.path.join(cli_dir, vk)
                    if not os.path.isdir(trip_dir):
                        continue
                    apro = _load_aprobacion(trip_dir)
                    if apro.get("status") in ("aprobado", "aprobado_parcial") and apro.get("status_socia") == "pendiente":
                        pendientes_socia += 1

    return render_template_string(DASHBOARD_HTML,
        nombre=nombre, rol=rol,
        pendientes=pendientes_sub, aprobados=aprobados_sub, rechazados=rechazados_sub,
        conteo_viajes=conteo_viajes, pendientes_socia=pendientes_socia)

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Viáticos V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f4f6f9;font-size:14px;color:#333}
header{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
       justify-content:space-between;align-items:center}
header h1{font-size:16px}
.hright{display:flex;gap:16px;align-items:center;font-size:13px}
.hright a{color:#adc8e6;text-decoration:none}
.hright a:hover{color:white}
.container{max-width:1100px;margin:20px auto;padding:0 16px}
.welcome{background:white;border-radius:8px;padding:16px 20px;margin-bottom:16px;
         box-shadow:0 1px 4px rgba(0,0,0,.08);display:flex;
         justify-content:space-between;align-items:center}
.welcome h2{font-size:17px;color:#1a3a5c}
.welcome p{font-size:13px;color:#777;margin-top:3px}
.btn{padding:9px 18px;border:none;border-radius:6px;cursor:pointer;
     font-size:13px;font-weight:bold;text-decoration:none;display:inline-block}
.btn-primary{background:#1a3a5c;color:white}
.btn-success{background:#27ae60;color:white}
.btn-warning{background:#f39c12;color:white}
.btn-danger {background:#e74c3c;color:white}
.btn-sm{padding:5px 12px;font-size:12px}
.btn:hover{opacity:.88}
.cards{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:20px}
.card-stat{background:white;border-radius:8px;padding:18px;text-align:center;
           box-shadow:0 1px 4px rgba(0,0,0,.08)}
.stat-n{font-size:28px;font-weight:bold;color:#1a3a5c}
.stat-l{font-size:12px;color:#777;margin-top:4px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.08);
      margin-bottom:16px;overflow:hidden}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;
             font-weight:bold;font-size:13px;display:flex;
             justify-content:space-between;align-items:center}
.card-body{padding:16px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f0f4f8;color:#1a3a5c;padding:8px 12px;text-align:left;border-bottom:2px solid #dde}
td{padding:8px 12px;border-bottom:1px solid #f0f0f0}
tr:hover td{background:#fafbfc}
.badge{display:inline-block;padding:3px 10px;border-radius:10px;font-size:11px;font-weight:bold}
.badge-pend{background:#fff3cd;color:#856404}
.badge-apro{background:#d1e7dd;color:#0f5132}
.badge-rech{background:#f8d7da;color:#842029}
.empty{text-align:center;color:#aaa;padding:24px;font-size:13px}
.tag-admin{background:#e8daef;color:#6c3483;padding:2px 8px;border-radius:8px;font-size:11px}
.tag-trab {background:#d6eaf8;color:#1a5276;padding:2px 8px;border-radius:8px;font-size:11px}
</style></head>
<body>
<header>
  <h1>⚖️ Sistema de Viáticos — Villar &amp; Villar Abogados</h1>
  <div class="hright">
    <span>{{ nombre }}
      {% if rol=='admin' %}<span class="tag-admin">Administradora</span>
      {% elif rol=='socia' %}<span class="tag-admin" style="background:#e8daef;color:#6c3483">Socia</span>
      {% else %}<span class="tag-trab">Trabajador</span>{% endif %}
    </span>
    <a href="/logout">Cerrar sesión</a>
  </div>
</header>

<div class="container">
  <div class="welcome">
    <div>
      <h2>Bienvenida, {{ nombre.split()[0] }}</h2>
      <p>{% if rol=='admin' %}Tienes acceso completo: puedes ver, validar y autorizar todos los gastos.
         {% elif rol=='socia' %}Puedes autorizar viáticos aprobados y capturar tus propios gastos.
         {% else %}Puedes capturar tus viáticos y enviarlos para validación.{% endif %}</p>
    </div>
    <div style="display:flex;gap:10px">
      <a href="/anticipo" class="btn" style="background:#f39c12;color:white">💰 Solicitar anticipo</a>
      <a href="/viaje"    class="btn btn-primary">📋 Comprobar viaje</a>
      <a href="/resumen_viaticos" class="btn" style="background:#1a5276;color:white">📊 Resumen de viáticos</a>
      {% if rol == 'admin' %}<a href="/config_email" class="btn" style="background:#95a5a6;color:white">✉ Config correo</a>{% endif %}
      {% if rol == 'admin' %}<a href="/admin/casetas" class="btn" style="background:#7f8c8d;color:white">🛣 Casetas</a>{% endif %}
      {% if rol == 'admin' %}<a href="/reportes" class="btn" style="background:#1a5276;color:white">📊 Reportes</a>{% endif %}
      {% if rol == 'socia' %}<a href="/autorizar_viaje" class="btn" style="background:#6c3483;color:white">🔏 Autorizaciones</a>{% endif %}
    </div>
  </div>

  <div class="cards">
    <div class="card-stat" style="cursor:pointer" onclick="location.href='/resumen_viaticos'">
      <div class="stat-n" style="color:#856404">{{ conteo_viajes.pendiente }}</div>
      <div class="stat-l">Viajes por revisar</div>
    </div>
    <div class="card-stat" style="cursor:pointer" onclick="location.href='/resumen_viaticos'">
      <div class="stat-n" style="color:#0f5132">{{ conteo_viajes.aprobado + conteo_viajes.aprobado_parcial }}</div>
      <div class="stat-l">Viajes aprobados</div>
    </div>
    <div class="card-stat" style="cursor:pointer" onclick="location.href='/resumen_viaticos'">
      <div class="stat-n" style="color:#842029">{{ conteo_viajes.rechazado }}</div>
      <div class="stat-l">Viajes rechazados</div>
    </div>
    <div class="card-stat" style="cursor:pointer" onclick="location.href='/resumen_viaticos'">
      <div class="stat-n" style="color:#1a5276">{{ conteo_viajes.total }}</div>
      <div class="stat-l">Total de viajes</div>
    </div>
  </div>
  {% if conteo_viajes.pendiente > 0 %}
  <div style="background:#fef9e7;border:1px solid #f0c040;border-radius:8px;
              padding:12px 18px;margin-bottom:16px;display:flex;align-items:center;gap:12px">
    <span style="font-size:22px">⏳</span>
    <div>
      <strong>{{ conteo_viajes.pendiente }} viaje{{ 's' if conteo_viajes.pendiente != 1 else '' }} pendiente{{ 's' if conteo_viajes.pendiente != 1 else '' }} de revisión</strong>
      <div style="font-size:12px;color:#777;margin-top:2px">
        <a href="/resumen_viaticos" style="color:#1a5276;font-weight:600">Ver resumen y aprobar →</a>
      </div>
    </div>
  </div>
  {% endif %}

  {% if rol == 'socia' %}
  <div style="background:#f5eef8;border:1px solid #c39bd3;border-radius:8px;
              padding:12px 18px;margin-bottom:16px;display:flex;align-items:center;gap:12px">
    <span style="font-size:22px">🔏</span>
    <div>
      <strong>{{ pendientes_socia }} viaje{{ 's' if pendientes_socia != 1 else '' }} pendiente{{ 's' if pendientes_socia != 1 else '' }} de tu autorización</strong>
      <div style="font-size:12px;color:#777;margin-top:2px">
        <a href="/autorizar_viaje" style="color:#6c3483;font-weight:600">Ver y autorizar →</a>
      </div>
    </div>
  </div>
  {% endif %}

  {% if pendientes %}
  <div class="card">
    <div class="card-header">
      <span>⏳ Pendientes de revisión</span>
    </div>
    <div class="card-body" style="padding:0">
      <table>
        <thead><tr>
          <th>Fecha envío</th><th>Empleado</th><th>Viaje</th><th>Cliente</th>
          <th>Total empleado</th>
          {% if rol=='admin' %}<th>Total cliente</th>{% endif %}
          <th>Acción</th>
        </tr></thead>
        <tbody>
        {% for s in pendientes %}
        <tr>
          <td>{{ s.fecha_envio }}</td>
          <td>{{ s.beneficiario }}</td>
          <td>{{ "💰 Anticipo" if s.get("tipo_solicitud")=="anticipo" else "📋 Comprobación" }} {{ s.viaje or s.ciudad or "" }}</td>
          <td>{{ s.cargo }}</td>
          <td style="font-weight:bold;color:#1a5276">${{ "%.2f"|format(s.total_empleado) }}</td>
          {% if rol=='admin' %}<td style="font-weight:bold;color:#a04000">${{ "%.2f"|format(s.total_cliente) }}</td>{% endif %}
          <td>
            <a href="/revisar/{{ s.id }}" class="btn btn-warning btn-sm">
              {% if rol=='admin' %}Revisar{% else %}Ver detalle{% endif %}
            </a>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endif %}

  {% if aprobados %}
  <div class="card">
    <div class="card-header" style="background:#0f5132">
      <span>✅ Aprobados</span>
    </div>
    <div class="card-body" style="padding:0">
      <table>
        <thead><tr>
          <th>Fecha envío</th><th>Empleado</th><th>Viaje</th><th>Cliente</th>
          <th>Total empleado</th>
          {% if rol=='admin' %}<th>Total cliente</th>{% endif %}
          <th>Acción</th>
        </tr></thead>
        <tbody>
        {% for s in aprobados %}
        <tr>
          <td>{{ s.fecha_envio }}</td>
          <td>{{ s.beneficiario }}</td>
          <td>{{ "💰 Anticipo" if s.get("tipo_solicitud")=="anticipo" else "📋 Comprobación" }} {{ s.viaje or s.ciudad or "" }}</td>
          <td>{{ s.cargo }}</td>
          <td style="font-weight:bold;color:#1a5276">${{ "%.2f"|format(s.total_empleado) }}</td>
          {% if rol=='admin' %}<td style="font-weight:bold;color:#a04000">${{ "%.2f"|format(s.total_cliente) }}</td>{% endif %}
          <td>
            <a href="/revisar/{{ s.id }}" class="btn btn-sm" style="background:#0f5132;color:white">Ver / Excel</a>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endif %}

  {% if rechazados %}
  <div class="card">
    <div class="card-header" style="background:#842029">
      <span>❌ Rechazados</span>
    </div>
    <div class="card-body" style="padding:0">
      <table>
        <thead><tr>
          <th>Fecha envío</th><th>Empleado</th><th>Viaje</th><th>Comentario</th><th>Acción</th>
        </tr></thead>
        <tbody>
        {% for s in rechazados %}
        <tr>
          <td>{{ s.fecha_envio }}</td>
          <td>{{ s.beneficiario }}</td>
          <td>{{ "💰 Anticipo" if s.get("tipo_solicitud")=="anticipo" else "📋 Comprobación" }} {{ s.viaje or s.ciudad or "" }}</td>
          <td style="color:#842029">{{ s.comentario }}</td>
          <td><a href="/viaje?editar={{ s.id }}" class="btn btn-primary btn-sm">Corregir</a></td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endif %}

  {% if not pendientes and not aprobados and not rechazados %}
  <div class="card"><div class="card-body empty">
    No hay viajes registrados aún. Haz clic en <strong>➕ Nuevo viaje</strong> para comenzar.
  </div></div>
  {% endif %}

</div>
</body></html>"""

# ─── CAPTURA DE VIAJE ─────────────────────────────────────────────────────────

def _strip_acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', nfc(str(s)))
                   if unicodedata.category(c) != 'Mn').lower()

def get_cal_trips_for_user(nombre_usuario):
    """Devuelve los viajes del calendario filtrados por el usuario actual."""
    try:
        cal_file = None
        for f in os.listdir(BASE_DIR):
            if "calendario" in nfc(f).lower() and f.lower().endswith(".xlsx"):
                cal_file = os.path.join(BASE_DIR, f)
                break
        if not cal_file and os.path.exists(CALENDAR):
            cal_file = CALENDAR
        if not cal_file:
            return []
        records = load_calendar(cal_file)
        user_plain = _strip_acc(nombre_usuario)
        propios = [r for r in records
                   if _strip_acc(r.get("abogado", "")) == user_plain]
        # Ordenar por fecha_inicio desc (más recientes primero)
        def parse_f(s):
            try: return datetime.strptime(s, "%d/%m/%Y")
            except: return datetime.min
        propios.sort(key=lambda r: parse_f(r["fecha_inicio"]), reverse=True)
        return propios
    except Exception as e:
        print(f"[cal_trips] {e}")
        return []

@app.route("/viaje")
@login_required
def viaje_page():
    clientes = sorted([
        nfc(d) for d in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, d)) and nfc(d).startswith("Viáticos ")
    ])
    cliente_sel = nfc(request.args.get("cliente", ""))
    # viaje_sel = "ciudad|fecha_ini"  (pipe-separated, enviado por el form)
    viaje_raw   = request.args.get("viaje", "")

    nombre_usuario = session["nombre"]

    # Viajes del calendario para este usuario (solo trabajadores)
    cal_trips = []
    if session.get("rol") != "admin":
        cal_trips = get_cal_trips_for_user(nombre_usuario)

    # Parsear la selección de viaje
    ciudad_sel = ""
    fecha_sel  = ""
    viaje_key  = ""
    cal_info   = {}
    if viaje_raw and viaje_raw != "__nuevo__":
        parts = viaje_raw.split("|", 1)
        ciudad_sel = parts[0]
        fecha_sel  = parts[1] if len(parts) > 1 else ""
        viaje_key  = safe_name(ciudad_sel + ("_" + fecha_sel.replace("/","") if fecha_sel else ""))
        # Buscar en calendario
        try:
            cal_file = None
            client_dir = find_dir(BASE_DIR, cliente_sel) if cliente_sel else None
            search_dirs = ([client_dir] if client_dir else []) + [BASE_DIR]
            for sd in search_dirs:
                for f in os.listdir(sd):
                    if "calendario" in nfc(f).lower() and f.lower().endswith(".xlsx"):
                        cal_file = os.path.join(sd, f); break
                if cal_file: break
            if not cal_file and os.path.exists(CALENDAR):
                cal_file = CALENDAR
            records = load_calendar(cal_file) if cal_file else []
            user_plain = _strip_acc(nombre_usuario)
            propios = [r for r in records if _strip_acc(r.get("abogado","")) == user_plain]
            if propios: records = propios
            hint = [fecha_sel] if fecha_sel else None
            info = lookup_calendar(ciudad_sel, records, hint_dates=hint)
            if info:
                cal_info = info
        except Exception as e:
            print(f"[viaje_page cal] {e}")

    anticipos_aprobados = [
        s for s in list_submissions(empleado_filter=nombre_usuario)
        if s.get("status") == "aprobado" and s.get("tipo_solicitud") == "anticipo"
    ]

    # ── Calcular días y comidas ANTES de cargar gastos ───────────────────────
    dias_viaje = 1
    if cal_info and cal_info.get("salida") and cal_info.get("regreso"):
        try:
            fi = datetime.strptime(cal_info["salida"],  "%d/%m/%Y")
            ff = datetime.strptime(cal_info["regreso"], "%d/%m/%Y")
            dias_viaje = max(int((ff - fi).days) + 1, 1)
        except Exception:
            pass

    comidas_por_dia = {}
    for d in range(1, dias_viaje + 1):
        val = request.args.get(f"comidas_{d}", "1")
        try:
            comidas_por_dia[d] = max(1, min(3, int(val)))
        except Exception:
            comidas_por_dia[d] = 1
    total_comidas = sum(comidas_por_dia.values())

    # ── Edición inline: archivo que se está editando ──────────────────────────
    editar_archivo = request.args.get("editar", "")

    # ── Cargar gastos del viaje directamente (sin AJAX) ───────────────────────
    gastos_inicial = []
    if viaje_key and cliente_sel:
        try:
            trip_dir  = get_upload_dir(session["usuario"], cliente_sel, viaje_key)

            # Leer overrides guardados manualmente
            overrides = {}
            overrides_path = os.path.join(trip_dir, "_overrides.json")
            if os.path.exists(overrides_path):
                try:
                    with open(overrides_path, encoding="utf-8") as _f:
                        overrides = json.load(_f)
                except Exception: pass

            tabulador = _get_tabulador(cliente_sel) if cliente_sel else {}
            city_data = lookup_city(ciudad_sel, tabulador) if ciudad_sel else None
            if not city_data:
                _cfg_path = os.path.join(trip_dir, "_config_viaje.json")
                if os.path.exists(_cfg_path):
                    try:
                        _cfg = json.load(open(_cfg_path, encoding="utf-8"))
                        city_data = _city_data_fallback(_cfg.get("km_mapa"), _cfg.get("casetas_mapa", 0))
                    except Exception: pass
            meals_sug = meals_by_distance(city_data["km_total"]) if city_data else 1
            for fname in sorted(f for f in os.listdir(trip_dir)
                                if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))):
                fpath = os.path.join(trip_dir, fname)
                cat   = detect_category(fname)
                if fname.lower().endswith(".pdf"):
                    d = extract_pdf_data(fpath)
                    amount, rfc = d["amount"], d["rfc"]
                else:
                    amount, rfc = 0.0, ""
                causa_iva = cat not in {"Hotel","Autobús","Vuelo","Otro"}
                meals_override = total_comidas if cat == "Comida" else None
                emp, cli  = calc_row(cat, amount, city_data, meals_override or meals_sug)
                row = {
                    "archivo":     fname,
                    "tipo":        cat,
                    "rfc":         rfc,
                    "descripcion": os.path.splitext(fname)[0],
                    "con_iva":     round(amount,2) if causa_iva else 0.0,
                    "sin_iva":     0.0 if causa_iva else round(amount,2),
                    "empleado":    round(emp,2),
                    "cliente":     round(cli,2),
                }
                # Aplicar ediciones manuales guardadas
                if fname in overrides:
                    row.update(overrides[fname])
                gastos_inicial.append(row)
            # Persistir los montos calculados para que el resumen los lea correctamente
            try:
                calc_path = os.path.join(trip_dir, "_gastos_calculados.json")
                with open(calc_path, "w", encoding="utf-8") as _cf:
                    json.dump(gastos_inicial, _cf, ensure_ascii=False, indent=2)
                # Guardar también el número de comidas elegido para que el resumen use el mismo
                config_path = os.path.join(trip_dir, "_config_viaje.json")
                cfg_existing = {}
                if os.path.exists(config_path):
                    try:
                        with open(config_path, encoding="utf-8") as _cfg:
                            cfg_existing = json.load(_cfg)
                    except Exception: pass
                cfg_existing["total_comidas"]   = total_comidas
                cfg_existing["comidas_por_dia"] = comidas_por_dia
                # Guardar datos del calendario para que el Excel los use correctamente
                if cal_info:
                    for k in ("salida","regreso","motivo","descripcion","ciudad"):
                        if k in cal_info:
                            cfg_existing[k] = cal_info[k]
                    # Asegurar que "descripcion" tenga un valor (usar motivo como fallback)
                    if not cfg_existing.get("descripcion"):
                        cfg_existing["descripcion"] = cal_info.get("motivo", "")
                cfg_existing["lugar"] = ciudad_sel
                with open(config_path, "w", encoding="utf-8") as _cfg:
                    json.dump(cfg_existing, _cfg, ensure_ascii=False, indent=2)
            except Exception: pass
        except Exception as e:
            print(f"[viaje_page gastos] {e}")

    return render_template_string(VIAJE_HTML,
        nombre=nombre_usuario, rol=session["rol"],
        clientes=clientes, cliente_sel=cliente_sel,
        cal_trips=cal_trips, viaje_raw=viaje_raw,
        ciudad_sel=ciudad_sel, fecha_sel=fecha_sel,
        viaje_key=viaje_key, cal_info=cal_info,
        gastos_inicial=gastos_inicial,
        dias_viaje=dias_viaje,
        comidas_por_dia=comidas_por_dia,
        editar_archivo=editar_archivo,
        anticipos=anticipos_aprobados)

VIAJE_HTML = """<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Capturar viaje — Viáticos V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f4f6f9;font-size:14px;color:#333}
header{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
       justify-content:space-between;align-items:center}
header h1{font-size:15px}
.hright{display:flex;gap:16px;align-items:center;font-size:13px}
.hright a{color:#adc8e6;text-decoration:none}
.container{max-width:1100px;margin:16px auto;padding:0 16px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);
      margin-bottom:14px;overflow:hidden}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;
             font-weight:bold;font-size:13px}
.card-body{padding:16px}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}
label{display:block;font-size:12px;color:#555;margin-bottom:3px}
input,select,textarea{width:100%;padding:7px 10px;border:1px solid #ccc;
                      border-radius:5px;font-size:13px;font-family:Arial,sans-serif}
textarea{resize:vertical}
.btn{padding:8px 18px;border:none;border-radius:5px;cursor:pointer;
     font-size:13px;font-weight:bold;text-decoration:none;display:inline-block}
.btn-primary{background:#1a3a5c;color:white}
.btn-success{background:#27ae60;color:white}
.btn-secondary{background:#7f8c8d;color:white}
.btn-danger{background:#e74c3c;color:white}
.btn:hover{opacity:.88}
.btn-row{display:flex;gap:10px;flex-wrap:wrap}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:#1a3a5c;color:white;padding:8px 10px;text-align:left;white-space:nowrap}
td{padding:7px 10px;border-bottom:1px solid #eee}
tr:hover td{background:#f8f9fa}
.tag{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}
.tag-Hotel{background:#d6eaf8;color:#1a5276}
.tag-Vuelo{background:#d5f5e3;color:#1e8449}
.tag-Comida{background:#fdebd0;color:#a04000}
.tag-Transporte{background:#e8daef;color:#6c3483}
.tag-Autobús{background:#d0ece7;color:#0e6655}
.tag-Casetas{background:#f2f3f4;color:#555;border:1px solid #bbb}
.tag-Gasolina{background:#fef9e7;color:#7d6608}
.tag-Otro{background:#f2f3f4;color:#555}
.warn-zero{background:#fff3cd !important}
.sum-emp{font-size:20px;font-weight:bold;color:#1a5276}
.meals-box{background:#fef9e7;border:1px solid #f0c040;border-radius:6px;
           padding:10px 14px;display:flex;gap:16px;align-items:flex-end;margin-bottom:12px}
#loading{display:none;position:fixed;inset:0;background:rgba(0,0,0,.35);
         z-index:999;align-items:center;justify-content:center}
.spinner{background:white;border-radius:10px;padding:30px 40px;text-align:center}
.spin{width:40px;height:40px;border:4px solid #ddd;border-top-color:#1a3a5c;
      border-radius:50%;animation:sp .8s linear infinite;margin:0 auto 12px}
@keyframes sp{to{transform:rotate(360deg)}}
.modal{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);
       z-index:100;align-items:center;justify-content:center}
.modal-box{background:white;border-radius:10px;padding:24px;width:480px;
           max-width:95vw;box-shadow:0 8px 32px rgba(0,0,0,.2)}
.modal-title{font-size:15px;font-weight:bold;margin-bottom:16px;color:#1a3a5c}
.mgrid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}
.mfull{grid-column:span 2}
.alert{padding:10px 14px;border-radius:6px;font-size:13px;margin-top:10px}
.alert-ok{background:#d1e7dd;color:#0f5132}
.alert-err{background:#f8d7da;color:#842029}
.status-bar{padding:7px 16px;background:#eaf4fb;border-top:1px solid #d6eaf8;
            font-size:12px;color:#555;position:fixed;bottom:0;left:0;right:0}
</style>
</head>
<body>
<div id="loading"><div class="spinner"><div class="spin"></div>Leyendo comprobantes…</div></div>

<header>
  <h1>⚖️ Nuevo viaje — {{ nombre }}</h1>
  <div class="hright">
    <a href="/resumen_viaticos">📊 Mi resumen</a>
    <a href="/dashboard">← Volver</a>
    <a href="/logout">Salir</a>
  </div>
</header>

<div class="container" style="padding-bottom:40px">

  <!-- PASO 1 -->
  <div class="card">
    <div class="card-header">1. Seleccionar cliente y viaje</div>
    <div class="card-body">
      <form method="GET" action="/viaje" id="frmCliente">
      <div class="grid-3" style="align-items:flex-end">
        <div>
          <label>Cliente</label>
          <select id="selCliente" name="cliente" onchange="this.form.submit()">
            <option value="">— Seleccione cliente —</option>
            {% for c in clientes %}
            <option value="{{ c }}" {% if c==cliente_sel %}selected{% endif %}>{{ c }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Viaje / Destino</label>
          <select id="selViaje" name="viaje" onchange="this.form.submit()">
            <option value="">— Seleccione viaje —</option>
            {% if cal_trips %}
            <optgroup label="📅 Del Calendario de viajes">
              {% for t in cal_trips %}
              <option value="{{ t.ciudad }}|{{ t.fecha_inicio }}"
                      {% if ciudad_sel == t.ciudad and fecha_sel == t.fecha_inicio %}selected{% endif %}>
                {{ t.ciudad }} — {{ t.fecha_inicio }}{% if t.fecha_fin != t.fecha_inicio %} al {{ t.fecha_fin }}{% endif %}
              </option>
              {% endfor %}
            </optgroup>
            {% endif %}
            <optgroup label="✏️ Otro viaje">
              <option value="__nuevo__" {% if viaje_raw == '__nuevo__' %}selected{% endif %}>
                ➕ Escribir destino manualmente…
              </option>
            </optgroup>
          </select>
        </div>
        <div style="align-self:flex-end;display:flex;gap:8px">
          <button type="button" class="btn btn-primary" onclick="cargarViaje()"
                  style="padding:9px 18px;font-size:14px">
            ✓ Seleccionar viaje
          </button>
        </div>
      </div>
      </form>

      <!-- Valores del servidor para usar en JS -->
      <input type="hidden" id="pageViajeKey" value="{{ viaje_key }}">
      <input type="hidden" id="pageCiudad"   value="{{ ciudad_sel }}">
      <input type="hidden" id="pageFechaIni" value="{{ fecha_sel }}">

      <!-- Input para nombre de viaje manual -->
      <div id="nuevoViajeDiv" style="display:{% if viaje_raw == '__nuevo__' %}block{% else %}none{% endif %};margin-top:10px">
        <label>Nombre del viaje (se usará como identificador de sus archivos)</label>
        <div style="display:flex;gap:8px;margin-top:4px">
          <input id="nuevoViajeNombre" type="text" placeholder="ej. Guadalajara"
                 style="flex:1;padding:8px;border:1px solid #ccc;border-radius:5px;font-size:13px">
          <button class="btn btn-primary" onclick="confirmarNuevoViaje()" style="white-space:nowrap">✓ Usar</button>
        </div>
        <p style="font-size:11px;color:#999;margin-top:4px">
          Sus archivos se guardarán en una carpeta con este nombre.
        </p>
      </div>

      <!-- Subir comprobantes — formulario HTML nativo (sin JS) -->
      {% if viaje_key %}
      <form method="POST" action="/subir_comprobantes" enctype="multipart/form-data"
            style="margin-top:14px;padding-top:14px;border-top:1px solid #eee">
        <input type="hidden" name="cliente"   value="{{ cliente_sel }}">
        <input type="hidden" name="viaje_raw" value="{{ viaje_raw }}">
        <input type="hidden" name="viaje_key" value="{{ viaje_key }}">
        <label style="font-weight:bold;color:#1a3a5c;font-size:13px">
          📎 Subir comprobantes (PDF, JPG, PNG)
        </label>
        <div style="display:flex;gap:10px;align-items:center;margin-top:8px;flex-wrap:wrap">
          <input type="file" name="archivos" multiple
                 accept=".pdf,.jpg,.jpeg,.png"
                 style="flex:1;padding:6px;border:1px solid #ccc;border-radius:5px;font-size:13px">
          <button type="submit" class="btn btn-secondary" style="white-space:nowrap">
            ⬆ Subir archivos
          </button>
        </div>
      </form>
      {% else %}
      <div style="margin-top:14px;padding-top:14px;border-top:1px solid #eee;
                  color:#999;font-size:13px">
        📎 Selecciona el viaje arriba para poder subir comprobantes.
      </div>
      {% endif %}
    </div>
  </div>

  <!-- PASO 2 -->
  <div class="card">
    <div class="card-header">2. Información del viaje</div>
    <div class="card-body">
      <!-- Indicador de calendario -->
      {% if cal_info %}
      <div style="padding:8px 14px;border-radius:6px;font-size:12px;margin-bottom:12px;
                  background:#d1e7dd;color:#0f5132;border:1px solid #a3cfbb">
        📅 <strong>Datos del Calendario de viajes:</strong>
        fechas, destino y asunto completados automáticamente. Puede editarlos si es necesario.
      </div>
      {% elif ciudad_sel and ciudad_sel != '__nuevo__' %}
      <div style="padding:8px 14px;border-radius:6px;font-size:12px;margin-bottom:12px;
                  background:#fff3cd;color:#856404;border:1px solid #ffc107">
        ✏️ <strong>Viaje no encontrado en el calendario.</strong>
        Por favor complete manualmente: fechas, destino y asunto.
      </div>
      {% endif %}
      <!-- Indicador de calendario (para JS) -->
      <div id="cal-status" style="display:none;padding:8px 14px;border-radius:6px;
           font-size:12px;margin-bottom:12px;border:1px solid transparent"></div>

      <div class="grid-2" style="margin-bottom:12px">
        <div>
          <label>Beneficiario</label>
          <input id="beneficiario" type="text" value="{{ nombre }}"
                 style="background:#f8f9fa;color:#333;font-weight:bold" readonly>
        </div>
        <div>
          <label>Con cargo a (cliente)</label>
          <input id="cargo" type="text"
                 value="{{ cliente_sel.replace('Viáticos ','') if cliente_sel else '' }}">
        </div>
        <div><label>Fecha de salida</label>
          <input id="salida" type="text" placeholder="DD/MM/AAAA"
                 value="{{ cal_info.salida if cal_info else '' }}">
        </div>
        <div><label>Fecha de regreso</label>
          <input id="regreso" type="text" placeholder="DD/MM/AAAA"
                 value="{{ cal_info.regreso if cal_info else '' }}">
        </div>
        <div><label>Lugar (ciudad destino)</label>
          <input id="lugar" type="text"
                 value="{{ cal_info.lugar if cal_info else ciudad_sel }}">
        </div>

        <!-- ── Calculadora de ruta con mapa ── -->
        <div id="mapaBox" style="background:#eaf4fb;border:1px solid #7fb3d3;border-radius:8px;
                                  padding:12px 14px;margin-top:-4px;margin-bottom:4px">
          <div style="font-weight:bold;color:#1a5276;font-size:13px;margin-bottom:8px">
            🗺 Calculadora de ruta (opcional)
            <span style="font-weight:normal;font-size:11px;color:#888">
              — Usa si la ciudad no está en el tabulador
            </span>
          </div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px">
            <div>
              <label style="font-size:12px;color:#555">Origen</label>
              <input id="mapaOrigen" type="text" placeholder="Ej: Querétaro, Qro."
                     value="Querétaro, Qro."
                     style="width:100%;padding:6px;border:1px solid #aad4ee;border-radius:5px;font-size:12px;box-sizing:border-box">
            </div>
            <div>
              <label style="font-size:12px;color:#555">Destino</label>
              <input id="mapaDestino" type="text" placeholder="Ej: Tlalnepantla, Méx."
                     style="width:100%;padding:6px;border:1px solid #aad4ee;border-radius:5px;font-size:12px;box-sizing:border-box">
            </div>
          </div>
          <div style="margin-bottom:8px">
            <button type="button" onclick="calcularRutaMapa()"
                    style="background:#2980b9;color:white;border:none;border-radius:5px;
                           padding:7px 16px;font-size:13px;cursor:pointer">
              📐 Calcular distancia
            </button>
            <span id="mapaStatus" style="font-size:12px;color:#888;margin-left:10px"></span>
          </div>
          <div id="mapaResultado" style="display:none;background:white;border:1px solid #76b7d4;
                                          border-radius:6px;padding:10px;font-size:13px">
            <div style="margin-bottom:6px;color:#555">
              📍 Distancia calculada (ida y vuelta):
              <strong id="mapaKmOSRM" style="color:#888"></strong>
              <span style="color:#888;font-size:11px"> km (referencia OSRM)</span>
            </div>
            <label style="font-size:12px;color:#333;display:block;margin-bottom:4px">
              ✏️ <strong>Km totales a usar</strong> (puedes corregir si conoces la distancia real):
            </label>
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
              <input id="mapaKmLabel" type="number" min="1" step="1"
                     style="width:90px;padding:5px 8px;border:2px solid #2980b9;border-radius:5px;
                            font-size:15px;font-weight:bold;color:#1a5276;text-align:center"
                     oninput="_mapaKmCalculado=parseInt(this.value)||0">
              <span style="font-size:12px;color:#555">km ida y vuelta</span>
              <label style="font-size:12px;color:#555;display:flex;align-items:center;gap:4px;margin-left:8px">
                <input type="checkbox" id="mapaExtra50" onchange="
                  var base=parseInt(document.getElementById('mapaKmLabel').value)||0;
                  document.getElementById('mapaKmLabel').value = this.checked ? base+50 : base-50;
                  _mapaKmCalculado=parseInt(document.getElementById('mapaKmLabel').value)||0;
                "> +50 km ciudad
              </label>
            </div>
            <br>
            <label style="font-size:12px;color:#555;margin-top:6px;display:block">
              💳 Casetas estimadas (opcional):
              <input id="mapaCasetas" type="number" min="0" step="50" value="0"
                     style="width:100px;padding:4px 6px;border:1px solid #ccc;border-radius:4px;
                            font-size:12px;margin-left:4px">
              <span style="font-size:11px;color:#888"> MXN ida y vuelta</span>
            </label>
            <button type="button" onclick="guardarKmMapa()"
                    style="background:#27ae60;color:white;border:none;border-radius:5px;
                           padding:6px 14px;font-size:12px;cursor:pointer;margin-top:8px">
              ✅ Usar estos km para el cálculo
            </button>
            <span id="mapaGuardadoMsg" style="font-size:12px;color:#27ae60;margin-left:8px;display:none">
              ✓ Guardado
            </span>
          </div>
          <div id="mapContainer" style="display:none;height:250px;border-radius:6px;
                                         overflow:hidden;margin-top:8px;border:1px solid #aad4ee"></div>
        </div>

        <div><label>Asunto / No. de expediente</label>
          <input id="asunto" type="text"
                 value="{{ (cal_info.descripcion or '').split('\n')[0] if cal_info else '' }}">
        </div>
      </div>

      <!-- Vinculación con anticipo -->
      <div style="background:#e8f8f5;border:1px solid #76d7c4;border-radius:8px;
                  padding:12px 14px;margin-bottom:12px">
        <label style="font-weight:bold;color:#0e6655;font-size:13px">
          🔗 Solicitud de anticipo (opcional)
        </label>
        <select id="selAnticipo" onchange="vincularAnticipo()"
                style="width:100%;margin-top:6px;padding:8px;border:1px solid #76d7c4;
                       border-radius:5px;font-size:13px">
          <option value="">— Sin anticipo previo —</option>
          {% for a in anticipos %}
          <option value="{{ a.id }}"
                  data-monto="{{ a.total_empleado }}"
                  data-ciudad="{{ a.ciudad or a.lugar or '' }}"
                  data-cargo="{{ a.cargo }}"
                  data-salida="{{ a.salida }}"
                  data-regreso="{{ a.regreso }}"
                  data-asunto="{{ a.asunto }}"
                  data-descripcion="{{ a.descripcion or a.asunto or '' }}">
            #{{ a.id }} — {{ a.ciudad or a.lugar }} — ${{ "%.2f"|format(a.total_empleado) }}
            (aprobado {{ a.fecha_aprobacion or a.fecha_envio }})
          </option>
          {% endfor %}
        </select>
        <div id="anticipo-info" style="margin-top:8px;font-size:12px;color:#0e6655;
                                       display:none;padding:6px 10px;background:#d1fae5;
                                       border-radius:5px">
          ✅ Anticipo aprobado: <strong id="anticipo-monto-label"></strong>
          — los datos del viaje se llenaron automáticamente
        </div>
        {% if not anticipos %}
        <p style="font-size:12px;color:#999;margin-top:6px">
          No tienes anticipos aprobados. Si solicitaste uno, espera a que Brenda lo apruebe.
        </p>
        {% endif %}
      </div>

      <!-- Comidas por día -->
      <div class="meals-box" style="flex-direction:column;align-items:flex-start">
        <div style="font-weight:bold;color:#7d6608;margin-bottom:8px">
          🍽 Comidas por día
          <span style="font-weight:normal;font-size:11px;color:#999">
            — 1 = menos de 6h &nbsp;|&nbsp; 2 = viaje largo &nbsp;|&nbsp; 3 = todo el día
          </span>
        </div>
        {% if viaje_key %}
        <form method="GET" action="/viaje" id="frmComidas">
          <input type="hidden" name="cliente"   value="{{ cliente_sel }}">
          <input type="hidden" name="viaje"     value="{{ viaje_raw }}">
          <div id="comidas-tabla-viaje" style="display:flex;flex-wrap:wrap;gap:10px">
            {% for d in range(1, dias_viaje + 1) %}
            <div style="display:flex;align-items:center;gap:6px;background:white;
                        border:1px solid #f0c040;border-radius:6px;padding:6px 10px">
              <span style="font-size:12px;color:#7d6608;white-space:nowrap">Día {{ d }}</span>
              <select name="comidas_{{ d }}" id="comidas_dia_{{ d }}"
                      style="width:60px;padding:4px;font-size:13px;font-weight:bold;
                             text-align:center;border:1px solid #ccc;border-radius:4px">
                <option value="1" {% if comidas_por_dia[d] == 1 %}selected{% endif %}>1</option>
                <option value="2" {% if comidas_por_dia[d] == 2 %}selected{% endif %}>2</option>
                <option value="3" {% if comidas_por_dia[d] == 3 %}selected{% endif %}>3</option>
              </select>
              <span style="font-size:11px;color:#aaa">comida(s)</span>
            </div>
            {% endfor %}
          </div>
          <button type="submit" class="btn"
                  style="background:#f0c040;color:#333;margin-top:8px;padding:7px 14px">
            🔄 Recalcular comidas
          </button>
        </form>
        {% else %}
        <div id="comidas-tabla-viaje">
          <span style="color:#999;font-size:12px">Seleccione viaje para ver días</span>
        </div>
        <button type="button" class="btn"
                style="background:#f0c040;color:#333;margin-top:8px;padding:7px 14px"
                disabled>🔄 Recalcular comidas</button>
        {% endif %}
      </div>
      <div><label>Descripción del viaje</label>
        <textarea id="descripcion" rows="2">{{ cal_info.descripcion if cal_info else '' }}</textarea></div>
    </div>
  </div>

  <!-- PASO 3 -->
  <div class="card">
    <div class="card-header">3. Comprobantes</div>
    <div class="card-body" style="padding-bottom:8px">
      <div class="btn-row" style="margin-bottom:10px">
        <button class="btn btn-secondary" onclick="agregarManual()">+ Agregar gasto manual</button>
      </div>
      <div style="overflow-x:auto">
        <table id="tablaGastos">
          <thead><tr>
            <th>#</th><th>Descripción / Archivo</th><th>Tipo</th>
            <th>RFC Emisor</th><th>Monto c/IVA</th><th>Monto s/IVA</th>
            <th>Empleado paga</th><th>Acción</th>
          </tr></thead>
          <tbody id="gastosBody">
            {% if gastos_inicial %}
              {% for g in gastos_inicial %}
              {% if g.archivo == editar_archivo %}
              {# ── FILA EN MODO EDICIÓN ── #}
              <tr style="background:#eaf4fb">
                <td colspan="8" style="padding:0">
                <form method="POST" action="/guardar_edicion" style="display:contents">
                  <input type="hidden" name="cliente"   value="{{ cliente_sel }}">
                  <input type="hidden" name="viaje_raw" value="{{ viaje_raw }}">
                  <input type="hidden" name="viaje_key" value="{{ viaje_key }}">
                  <input type="hidden" name="archivo"   value="{{ g.archivo }}">
                  {% for d, n in comidas_por_dia.items() %}
                  <input type="hidden" name="comidas_{{ d }}" value="{{ n }}">
                  {% endfor %}
                  <table style="width:100%;border-collapse:collapse">
                  <tr style="background:#eaf4fb">
                    <td style="width:30px;padding:4px 6px;color:#666">{{ loop.index }}</td>
                    <td style="padding:4px 4px">
                      <input name="descripcion" value="{{ g.descripcion or g.archivo }}"
                             style="width:100%;padding:4px;font-size:12px;border:1px solid #7fb3d3;border-radius:4px">
                    </td>
                    <td style="padding:4px 4px">
                      <select name="tipo" style="padding:4px;font-size:12px;border:1px solid #7fb3d3;border-radius:4px">
                        {% for t in ['Comida','Hotel','Gasolina','Transporte','Casetas','Autobús','Vuelo','Otro'] %}
                        <option value="{{ t }}" {% if g.tipo == t %}selected{% endif %}>{{ t }}</option>
                        {% endfor %}
                      </select>
                    </td>
                    <td style="padding:4px 4px">
                      <input name="rfc" value="{{ g.rfc or '' }}" placeholder="RFC"
                             style="width:120px;padding:4px;font-size:11px;border:1px solid #7fb3d3;border-radius:4px">
                    </td>
                    <td style="padding:4px 4px">
                      <input name="con_iva" type="number" step="0.01" min="0"
                             value="{{ g.con_iva }}"
                             style="width:80px;padding:4px;font-size:12px;border:1px solid #7fb3d3;border-radius:4px">
                    </td>
                    <td style="padding:4px 4px">
                      <input name="sin_iva" type="number" step="0.01" min="0"
                             value="{{ g.sin_iva }}"
                             style="width:80px;padding:4px;font-size:12px;border:1px solid #7fb3d3;border-radius:4px">
                    </td>
                    <td style="padding:4px 4px">
                      <input name="empleado" type="number" step="0.01" min="0"
                             value="{{ g.empleado }}"
                             style="width:80px;padding:4px;font-size:12px;font-weight:bold;
                                    color:#1a5276;border:1px solid #7fb3d3;border-radius:4px">
                    </td>
                    <td style="padding:4px 6px;white-space:nowrap">
                      <button type="submit" class="btn btn-success btn-sm"
                              style="padding:4px 10px;font-size:11px;background:#27ae60;color:white;border:none;border-radius:4px;cursor:pointer">
                        💾 Guardar
                      </button>
                      &nbsp;
                      <a href="/viaje?cliente={{ cliente_sel|urlencode }}&viaje={{ viaje_raw|urlencode }}{% for d, n in comidas_por_dia.items() %}&comidas_{{ d }}={{ n }}{% endfor %}"
                         class="btn btn-secondary btn-sm"
                         style="padding:4px 10px;font-size:11px;background:#ccc;color:#333;
                                border-radius:4px;text-decoration:none;display:inline-block">
                        ✕ Cancelar
                      </a>
                    </td>
                  </tr>
                  </table>
                </form>
                </td>
              </tr>
              {% else %}
              {# ── FILA NORMAL ── #}
              <tr {% if g.con_iva == 0 and g.sin_iva == 0 and g.tipo not in ['Comida','Gasolina'] %}class="warn-zero"{% endif %}>
                <td>{{ loop.index }}</td>
                <td title="{{ g.archivo }}">{{ g.descripcion or g.archivo }}{% if g.con_iva == 0 and g.sin_iva == 0 and g.tipo not in ['Comida','Gasolina'] %} ⚠️{% endif %}</td>
                <td><span class="tag tag-{{ g.tipo }}">{{ g.tipo }}</span></td>
                <td style="font-size:11px">{{ g.rfc or '—' }}</td>
                <td>{% if g.con_iva %}${{ "%.2f"|format(g.con_iva) }}{% else %}<span style="color:#e67e22">$0.00</span>{% endif %}</td>
                <td>{% if g.sin_iva %}${{ "%.2f"|format(g.sin_iva) }}{% else %}<span style="color:#e67e22">$0.00</span>{% endif %}</td>
                <td style="font-weight:bold;color:#1a5276">${{ "%.2f"|format(g.empleado) }}</td>
                <td style="white-space:nowrap">
                  {# Botón Editar: GET form para activar edición inline #}
                  <form method="GET" action="/viaje" style="display:inline">
                    <input type="hidden" name="cliente" value="{{ cliente_sel }}">
                    <input type="hidden" name="viaje"   value="{{ viaje_raw }}">
                    <input type="hidden" name="editar"  value="{{ g.archivo }}">
                    {% for d, n in comidas_por_dia.items() %}
                    <input type="hidden" name="comidas_{{ d }}" value="{{ n }}">
                    {% endfor %}
                    <button type="submit" class="btn btn-secondary btn-sm"
                            style="padding:3px 10px;font-size:11px">✎ Editar</button>
                  </form>
                  {# Botón Borrar: POST form #}
                  <form method="POST" action="/borrar_gasto" style="display:inline"
                        onsubmit="return confirm('¿Eliminar «{{ g.archivo }}»?')">
                    <input type="hidden" name="cliente"   value="{{ cliente_sel }}">
                    <input type="hidden" name="viaje_raw" value="{{ viaje_raw }}">
                    <input type="hidden" name="viaje_key" value="{{ viaje_key }}">
                    <input type="hidden" name="archivo"   value="{{ g.archivo }}">
                    {% for d, n in comidas_por_dia.items() %}
                    <input type="hidden" name="comidas_{{ d }}" value="{{ n }}">
                    {% endfor %}
                    <button type="submit" class="btn btn-danger btn-sm"
                            style="padding:3px 10px;font-size:11px">✗</button>
                  </form>
                </td>
              </tr>
              {% endif %}
              {% endfor %}
            {% else %}
              <tr><td colspan="8" style="text-align:center;color:#999;padding:20px">
                {% if viaje_key %}Aún no hay comprobantes — suba sus archivos arriba.
                {% else %}Seleccione cliente y viaje para comenzar.{% endif %}
              </td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- PASO 4 -->
  <div class="card">
    <div class="card-header">4. Resumen y envío</div>
    <div class="card-body">
      <div style="margin-bottom:16px;background:#d6eaf8;border-radius:8px;
                  padding:14px 18px;display:inline-block">
        <div style="font-size:12px;color:#555">Total a pagar al empleado</div>
        <div class="sum-emp" id="totalEmpleado">${{ "%.2f"|format(gastos_inicial|sum(attribute='empleado')) if gastos_inicial else '0.00' }}</div>
      </div>

      {% if viaje_key and gastos_inicial %}
      {# ── Descarga del formato de comprobación ── #}
      <div style="margin-bottom:16px">
        <a href="/descargar_comprobacion?cliente={{ cliente_sel|urlencode }}&viaje={{ viaje_raw|urlencode }}{% for d, n in comidas_por_dia.items() %}&comidas_{{ d }}={{ n }}{% endfor %}&salida={{ cal_info.get('salida','')|urlencode }}&regreso={{ cal_info.get('regreso','')|urlencode }}&asunto={{ cal_info.get('motivo','')|urlencode }}&descripcion={{ cal_info.get('descripcion', cal_info.get('motivo',''))|urlencode }}"
           class="btn"
           style="background:#1a5276;color:white;padding:9px 18px;font-size:14px;
                  text-decoration:none;border-radius:6px;display:inline-block">
          📥 Descargar formato de comprobación (.xlsx)
        </a>
        <div style="font-size:11px;color:#777;margin-top:6px">
          Genera el formato V&V-04 + V&V-05 listo para imprimir y firmar.
        </div>
      </div>
      {% endif %}

      <div class="btn-row">
        <button class="btn btn-success" onclick="enviarRevision()">
          📨 Enviar para validación (Brenda Morales)
        </button>
      </div>
      <div id="msgResult"></div>
    </div>
  </div>

</div><!-- /container -->

<!-- Modal edición -->
<div class="modal" id="editModal" onclick="cerrarModal(event)">
  <div class="modal-box">
    <div class="modal-title">✎ Editar gasto</div>
    <input type="hidden" id="editIdx">
    <div class="mgrid">
      <div class="mfull"><label>Descripción / Archivo</label><input id="eDesc" type="text"></div>
      <div><label>Tipo</label>
        <select id="eTipo">
          <option>Hotel</option><option>Vuelo</option><option>Comida</option>
          <option>Transporte</option><option>Casetas</option>
          <option>Gasolina</option><option>Autobús</option><option>Otro</option>
        </select></div>
      <div><label>RFC Emisor</label><input id="eRfc" type="text"></div>
      <div><label>Monto con IVA ($)</label><input id="eConIva" type="number" step="0.01"></div>
      <div><label>Monto sin IVA ($)</label><input id="eSinIva" type="number" step="0.01"></div>
      <div><label>Empleado paga ($)</label><input id="eEmp" type="number" step="0.01"></div>
    </div>
    <div class="btn-row">
      <button class="btn btn-primary" onclick="guardarEdicion()">✓ Guardar</button>
      <button class="btn btn-secondary"
              onclick="document.getElementById('editModal').style.display='none'">Cancelar</button>
    </div>
  </div>
</div>

<div class="status-bar" id="statusBar">Seleccione cliente y viaje para comenzar.</div>

<script>
// Gastos pre-cargados por el servidor al renderizar la página
let gastos = {{ gastos_inicial | tojson }};
let anticipo_id_sel = '';
let anticipo_monto_sel = 0;
let viajeManual = '';  // nombre cuando el usuario elige "Nuevo viaje"

// ── Helpers de viaje ─────────────────────────────────────────────────────────
// Los valores de viaje vienen del servidor (inputs ocultos pre-llenados por Jinja2)
function getViajeKey(){
  const hidden = document.getElementById('pageViajeKey');
  if(hidden && hidden.value) return hidden.value;
  return viajeManual;
}
function getViajeCiudad(){
  const hidden = document.getElementById('pageCiudad');
  if(hidden && hidden.value) return hidden.value;
  return viajeManual;
}
function getViajeValue(){
  return getViajeKey();
}

function calcDias(salida, regreso){
  try{
    const parts_s = salida.split('/');
    const parts_r = regreso.split('/');
    const fi = new Date(parts_s[2], parts_s[1]-1, parts_s[0]);
    const ff = new Date(parts_r[2], parts_r[1]-1, parts_r[0]);
    return Math.max(Math.floor((ff-fi)/(1000*60*60*24))+1, 1);
  }catch(e){ return 1; }
}

function confirmarNuevoViaje(){
  const input = document.getElementById('nuevoViajeNombre');
  if(!input || !input.value.trim()){ alert('Ingrese un nombre para el viaje.'); return; }
  viajeManual = input.value.trim().replace(/\s+/g,'_');
  // Actualizar el input oculto para que getViajeKey() lo use
  const hid = document.getElementById('pageViajeKey');
  if(hid) hid.value = viajeManual;
  const hc = document.getElementById('pageCiudad');
  if(hc) hc.value = input.value.trim();
  document.getElementById('lugar').value = input.value.trim();
  setStatus('✅ Viaje manual: "'+input.value.trim()+'" — suba los comprobantes.');
}

// ── Vinculación con anticipo ──────────────────────────────────────────────────
function vincularAnticipo(){
  const sel = document.getElementById('selAnticipo');
  if(!sel) return;
  const opt = sel.options[sel.selectedIndex];
  anticipo_id_sel    = opt.value;
  anticipo_monto_sel = parseFloat(opt.dataset.monto||0);
  const infoDiv  = document.getElementById('anticipo-info');
  const montoLbl = document.getElementById('anticipo-monto-label');
  if(anticipo_id_sel){
    infoDiv.style.display = 'block';
    montoLbl.textContent  = '$'+anticipo_monto_sel.toLocaleString('es-MX',
      {minimumFractionDigits:2,maximumFractionDigits:2});
    // Auto-llenar TODOS los campos desde el anticipo
    if(opt.dataset.salida)       document.getElementById('salida').value       = opt.dataset.salida;
    if(opt.dataset.regreso)      document.getElementById('regreso').value      = opt.dataset.regreso;
    if(opt.dataset.asunto)       document.getElementById('asunto').value       = opt.dataset.asunto;
    if(opt.dataset.ciudad)       document.getElementById('lugar').value        = opt.dataset.ciudad;
    if(opt.dataset.cargo)        document.getElementById('cargo').value        = opt.dataset.cargo;
    if(opt.dataset.descripcion)  document.getElementById('descripcion').value  = opt.dataset.descripcion;
    setStatus('✅ Anticipo #'+anticipo_id_sel+' vinculado ($'+
      anticipo_monto_sel.toLocaleString('es-MX',{minimumFractionDigits:2})+
      ') — datos del viaje llenados automáticamente');
  } else {
    infoDiv.style.display = 'none';
    anticipo_id_sel = ''; anticipo_monto_sel = 0;
  }
}

// ── Subida de comprobantes ────────────────────────────────────────────────────
function subirArchivos(){
  const cliente = document.getElementById('selCliente').value;
  const viaje   = getViajeValue();
  const files   = document.getElementById('archivosInput').files;
  const status  = document.getElementById('upload-status');

  if(!cliente){ alert('Seleccione el cliente antes de subir archivos.'); return; }
  if(!viaje){ alert('Seleccione o escriba el nombre del viaje antes de subir archivos.'); return; }
  if(!files||files.length===0){ alert('Seleccione al menos un archivo.'); return; }

  const fd = new FormData();
  fd.append('cliente',   cliente);
  fd.append('viaje_key', getViajeKey());
  for(const f of files) fd.append('archivos', f);

  status.textContent = '⬆ Subiendo '+files.length+' archivo(s)...';
  status.style.color = '#555';

  fetch('/api/subir_comprobantes', {method:'POST', body:fd})
    .then(r=>r.json()).then(data=>{
      if(data.error){
        status.textContent = '❌ '+data.error;
        status.style.color = 'red';
        return;
      }
      status.textContent = '✅ Subidos: '+data.guardados.join(', ')+' — recargando...';
      status.style.color = '#0f5132';
      document.getElementById('archivosInput').value = '';
      // Recargar la página para que el servidor incluya los nuevos comprobantes
      setTimeout(()=>window.location.reload(), 800);
    }).catch(e=>{
      status.textContent = '❌ Error: '+e;
      status.style.color = 'red';
    });
}

function cargarViaje(){
  const cliente   = document.getElementById('selCliente').value;
  const viaje_key = getViajeKey();
  const ciudad    = getViajeCiudad();
  const fecha_ini = (document.getElementById('pageFechaIni')||{}).value || '';
  if(!cliente||!viaje_key){ return; }   // sin viaje seleccionado: no hacer nada
  document.getElementById('loading').style.display = 'flex';
  fetch('/api/cargar', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({cliente, viaje_key, ciudad, fecha_ini})
  })
  .then(r=>{
    if(r.redirected || r.status===302){ window.location='/'; return null; }
    if(!r.ok) throw new Error('HTTP '+r.status);
    return r.json();
  })
  .then(data=>{
    if(!data) return;
    document.getElementById('loading').style.display = 'none';
    if(data.error){
      setStatus('❌ Error al cargar: '+data.error, 'red');
      document.getElementById('cal-status').style.display='none';
      return;
    }
    const i = data.info;
    // Beneficiario = usuario logueado (readonly, no se sobreescribe)
    const cargoActual = document.getElementById('cargo').value;
    if(!cargoActual && i.cargo) document.getElementById('cargo').value = i.cargo;
    if(i.salida||i.fecha)  document.getElementById('salida').value  = i.salida||i.fecha||'';
    if(i.regreso||i.fecha) document.getElementById('regreso').value = i.regreso||i.fecha||'';
    if(i.lugar)            document.getElementById('lugar').value   = i.lugar||'';
    if(i.motivo){
      document.getElementById('asunto').value      = (i.motivo||'').split('\n')[0]||'';
      document.getElementById('descripcion').value = i.motivo||'';
    }

    // Indicador de calendario
    const calDiv = document.getElementById('cal-status');
    if(calDiv){
      if(i.cal_found){
        calDiv.innerHTML = '📅 <strong>Datos del Calendario de viajes:</strong> fechas, destino y asunto cargados automáticamente. Puede editarlos si es necesario.';
        calDiv.style.background='#d1e7dd'; calDiv.style.color='#0f5132';
        calDiv.style.border='1px solid #a3cfbb';
      } else {
        calDiv.innerHTML = '✏️ <strong>Viaje no encontrado en el calendario.</strong> Por favor complete manualmente: fechas, destino y asunto.';
        calDiv.style.background='#fff3cd'; calDiv.style.color='#856404';
        calDiv.style.border='1px solid #ffc107';
      }
      calDiv.style.display='block';
    }

    // Construir tabla de comidas por día
    generarTablaComidasViaje(i.dias||1, i.meals_sugeridas||1);
    gastos = data.gastos;
    renderTabla();
    calcTotal();
    if(gastos.length===0){
      setStatus('⚠️ Aún no hay comprobantes. Suba sus archivos con el botón ⬆ Subir archivos.', 'orange');
    } else {
      setStatus('✅ '+gastos.length+' comprobante(s) cargados. Revise los montos en $0 (amarillo).', '#0f5132');
    }
  })
  .catch(e=>{
    document.getElementById('loading').style.display='none';
    setStatus('❌ Error de conexión: '+e, 'red');
  });
}

const TIPOS_CSS = {Hotel:'Hotel',Vuelo:'Vuelo',Comida:'Comida',Transporte:'Transporte',
                   'Autobús':'Autobús',Casetas:'Casetas',Gasolina:'Gasolina',Otro:'Otro'};

function renderTabla(){
  const tbody = document.getElementById('gastosBody');
  if(!gastos.length){
    tbody.innerHTML='<tr><td colspan="8" style="text-align:center;color:#999;padding:20px">Sin comprobantes</td></tr>';
    return;
  }
  tbody.innerHTML = gastos.map((g,i)=>{
    const esZero = (g.con_iva===0 && g.sin_iva===0 && g.tipo!=='Comida' && g.tipo!=='Gasolina');
    const rowCls = esZero ? 'class="warn-zero"' : '';
    const warn   = esZero ? ' ⚠️' : '';
    return `<tr ${rowCls}>
      <td>${i+1}</td>
      <td title="${g.archivo}">${g.descripcion||g.archivo}${warn}</td>
      <td><span class="tag tag-${TIPOS_CSS[g.tipo]||'Otro'}">${g.tipo}</span></td>
      <td style="font-size:11px">${g.rfc||'—'}</td>
      <td${esZero?' style="color:#e74c3c;font-weight:bold"':''}>${fmt(g.con_iva)}</td>
      <td>${fmt(g.sin_iva)}</td>
      <td style="font-weight:bold;color:#1a5276">${fmt(g.empleado)}</td>
      <td>
        <button class="btn btn-secondary btn-sm" style="padding:3px 10px;font-size:11px"
                onclick="editarFila(${i})">✎ Editar</button>
        <button class="btn btn-danger btn-sm" style="padding:3px 10px;font-size:11px"
                onclick="eliminarFila(${i})">✗</button>
      </td>
    </tr>`;
  }).join('');
}

function fmt(n){
  if(n===0||n===undefined||n===null) return '<span style="color:#e67e22">$0.00</span>';
  return '$'+Number(n).toLocaleString('es-MX',{minimumFractionDigits:2,maximumFractionDigits:2});
}

function calcTotal(){
  const emp = gastos.reduce((a,g)=>a+Number(g.empleado||0),0);
  document.getElementById('totalEmpleado').textContent =
    '$'+Number(emp).toLocaleString('es-MX',{minimumFractionDigits:2,maximumFractionDigits:2});
  return emp;
}

// ── Tabla de comidas por día (comprobación de viaje) ─────────────────────────
function generarTablaComidasViaje(dias, nDefault){
  const tabla = document.getElementById('comidas-tabla-viaje');
  if(!tabla) return;
  let html = '<div style="display:flex;flex-wrap:wrap;gap:10px">';
  for(let d=1; d<=dias; d++){
    html += `<div style="display:flex;align-items:center;gap:6px;
                         background:white;border:1px solid #f0c040;
                         border-radius:6px;padding:6px 10px">
      <span style="font-size:12px;color:#7d6608;white-space:nowrap">Día ${d}</span>
      <select id="comidas_dia_${d}" style="width:60px;padding:4px;font-size:13px;
                                           font-weight:bold;text-align:center;
                                           border:1px solid #ccc;border-radius:4px">
        <option value="1" ${nDefault==1?'selected':''}>1</option>
        <option value="2" ${nDefault==2?'selected':''}>2</option>
        <option value="3" ${nDefault==3?'selected':''}>3</option>
      </select>
      <span style="font-size:11px;color:#aaa">comida(s)</span>
    </div>`;
  }
  html += '</div>';
  tabla.innerHTML = html;
}

function getMealsPorDiaViaje(){
  const meals = [];
  let d = 1;
  while(document.getElementById('comidas_dia_'+d)){
    meals.push(parseInt(document.getElementById('comidas_dia_'+d).value)||1);
    d++;
  }
  return meals.length > 0 ? meals : null;
}

function recalcularComidas(){
  const meals = getMealsPorDiaViaje();
  if(!meals){ setStatus('Cargue un viaje primero.','orange'); return; }
  const totalMeals = meals.reduce((a,b)=>a+b, 0);
  let changed = 0;
  gastos.forEach(g=>{
    if(g.tipo==='Comida'){
      g.empleado = totalMeals * 290;
      g.cliente  = totalMeals * 370;
      changed++;
    }
  });
  renderTabla(); calcTotal();
  const detalle = meals.map((n,i)=>`Día ${i+1}: ${n}`).join(' | ');
  setStatus(changed>0
    ? `Recalculado: ${detalle} → Total ${totalMeals} comida(s) | Empleado: $${(totalMeals*290).toFixed(2)}`
    : 'No hay renglones de Comida para recalcular.');
}

function agregarManual(){
  gastos.push({archivo:'Manual',tipo:'Otro',rfc:'',descripcion:'',
               con_iva:0,sin_iva:0,empleado:0,cliente:0});
  renderTabla(); editarFila(gastos.length-1);
}

function editarFila(i){
  const g = gastos[i];
  document.getElementById('editIdx').value  = i;
  document.getElementById('eDesc').value    = g.descripcion||g.archivo;
  document.getElementById('eTipo').value    = g.tipo;
  document.getElementById('eRfc').value     = g.rfc||'';
  document.getElementById('eConIva').value  = g.con_iva||0;
  document.getElementById('eSinIva').value  = g.sin_iva||0;
  document.getElementById('eEmp').value     = g.empleado||0;
  document.getElementById('editModal').style.display = 'flex';
}

function guardarEdicion(){
  const i = parseInt(document.getElementById('editIdx').value);
  const conIva = parseFloat(document.getElementById('eConIva').value)||0;
  const sinIva = parseFloat(document.getElementById('eSinIva').value)||0;
  const emp    = parseFloat(document.getElementById('eEmp').value)||0;
  // Para transporte: cliente cobra lo mismo que empleado paga
  gastos[i] = {
    archivo:     gastos[i].archivo||'Manual',
    descripcion: document.getElementById('eDesc').value,
    tipo:        document.getElementById('eTipo').value,
    rfc:         document.getElementById('eRfc').value,
    con_iva:     conIva,
    sin_iva:     sinIva,
    empleado:    emp,
    cliente:     emp,   // por defecto cliente = mismo monto
  };
  document.getElementById('editModal').style.display = 'none';
  renderTabla(); calcTotal();
}

function cerrarModal(e){
  if(e.target===document.getElementById('editModal'))
    document.getElementById('editModal').style.display='none';
}
function eliminarFila(i){ gastos.splice(i,1); renderTabla(); calcTotal(); }

function enviarRevision(){
  const cliente = document.getElementById('selCliente').value;
  const viaje   = getViajeValue();
  if(!cliente||!viaje||!gastos.length){ alert('Cargue un viaje primero.'); return; }

  const zeroCount = gastos.filter(g=>g.con_iva===0&&g.sin_iva===0&&
    g.tipo!=='Comida'&&g.tipo!=='Gasolina').length;
  if(zeroCount>0){
    if(!confirm(`Hay ${zeroCount} comprobante(s) con monto $0.00 (marcados en amarillo). ¿Desea enviarlo de todas formas?`)) return;
  }

  const emp = gastos.reduce((a,g)=>a+Number(g.empleado||0),0);
  const cli = gastos.reduce((a,g)=>a+Number(g.cliente||0),0);
  // Diferencia: positivo = empresa debe al empleado, negativo = empleado reembolsa
  const diferencia = Math.round((emp - anticipo_monto_sel)*100)/100;

  fetch('/api/enviar', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({
      tipo_solicitud: 'comprobacion',
      cliente, viaje,
      salida:         document.getElementById('salida').value,
      regreso:        document.getElementById('regreso').value,
      beneficiario:   document.getElementById('beneficiario').value,
      cargo:          document.getElementById('cargo').value,
      asunto:         document.getElementById('asunto').value,
      lugar:          document.getElementById('lugar').value,
      descripcion:    document.getElementById('descripcion').value,
      gastos,
      total_empleado: emp,
      total_cliente:  cli,
      anticipo_id:    anticipo_id_sel||null,
      anticipo_monto: anticipo_monto_sel||0,
      diferencia:     diferencia,
    })
  }).then(r=>r.json()).then(data=>{
    if(data.error){
      document.getElementById('msgResult').innerHTML=
        '<div class="alert alert-err">Error: '+data.error+'</div>'; return;
    }
    let msg = '✅ Enviado para revisión a Brenda Morales. ID: <strong>'+data.id+'</strong>';
    if(anticipo_id_sel){
      const fmt = n=>Math.abs(n).toLocaleString('es-MX',{minimumFractionDigits:2});
      if(diferencia > 0)
        msg += '<br>💚 Diferencia a pagar al colaborador: <strong>$'+fmt(diferencia)+'</strong>';
      else if(diferencia < 0)
        msg += '<br>🔴 Colaborador debe reembolsar a la Firma: <strong>$'+fmt(diferencia)+'</strong>';
      else
        msg += '<br>✓ Sin diferencia — comprobación cuadra exactamente.';
    }
    document.getElementById('msgResult').innerHTML='<div class="alert alert-ok">'+msg+'</div>';
    setStatus('Solicitud enviada. ID: '+data.id);
  });
}

function setStatus(msg, color='#555'){
  const s = document.getElementById('statusBar');
  s.textContent = msg; s.style.color = color;
}

// ── Inicialización al abrir la página ────────────────────────────────────────
(function initPage(){
  const vk = document.getElementById('pageViajeKey');
  if(!vk || !vk.value) return;   // sin viaje: nada que mostrar
  // Tabla de comidas
  const s = document.getElementById('salida').value;
  const r = document.getElementById('regreso').value;
  if(s) generarTablaComidasViaje(calcDias(s, r||s), 1);
  // Los gastos ya vienen del servidor — solo renderizar
  renderTabla();
  calcTotal();
  if(gastos.length === 0){
    setStatus('⚠️ Aún no hay comprobantes. Suba sus archivos con el botón ⬆ Subir archivos.', 'orange');
  } else {
    setStatus('✅ ' + gastos.length + ' comprobante(s) cargados. Revise los montos en $0 (amarillo).', '#0f5132');
  }
})();

// ── Calculadora de ruta con mapa ───────────────────────────────────────────
var _leafletLoaded = false;
var _mapaMap = null;
var _mapaRouteLayer = null;
var _mapaKmCalculado = 0;

function _loadLeaflet(cb){
  if(_leafletLoaded){ cb(); return; }
  var lnk = document.createElement('link');
  lnk.rel='stylesheet';
  lnk.href='https://unpkg.com/leaflet@1.9.4/dist/leaflet.css';
  document.head.appendChild(lnk);
  var s = document.createElement('script');
  s.src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js';
  s.onload=function(){ _leafletLoaded=true; cb(); };
  document.head.appendChild(s);
}

async function geocodeMX(query){
  var url='https://nominatim.openstreetmap.org/search?format=json&countrycodes=mx&limit=1&q='+encodeURIComponent(query);
  var r = await fetch(url, {headers:{'Accept-Language':'es'}});
  var data = await r.json();
  if(data && data.length>0) return {lat:parseFloat(data[0].lat), lon:parseFloat(data[0].lon), display:data[0].display_name};
  return null;
}

async function calcularRutaMapa(){
  var origen  = document.getElementById('mapaOrigen').value.trim();
  var destino = document.getElementById('mapaDestino').value.trim();
  var status  = document.getElementById('mapaStatus');
  if(!origen || !destino){ status.textContent='⚠️ Ingresa origen y destino'; return; }

  status.textContent = '🔍 Buscando coordenadas…';
  var o = await geocodeMX(origen);
  var d = await geocodeMX(destino);
  if(!o){ status.textContent='❌ No encontré: '+origen; return; }
  if(!d){ status.textContent='❌ No encontré: '+destino; return; }

  status.textContent = '📐 Calculando ruta…';
  var url = 'https://router.project-osrm.org/route/v1/driving/'+o.lon+','+o.lat+';'+d.lon+','+d.lat+'?overview=full&geometries=geojson';
  var resp = await fetch(url);
  var rdata = await resp.json();
  if(!rdata.routes || rdata.routes.length===0){ status.textContent='❌ No se pudo calcular la ruta'; return; }

  var km_ida   = Math.round(rdata.routes[0].distance / 1000);
  var km_total = km_ida * 2;
  _mapaKmCalculado = km_total;

  // Mostrar resultado
  document.getElementById('mapaKmOSRM').textContent = km_total;
  document.getElementById('mapaKmLabel').value = km_total;
  document.getElementById('mapaExtra50').checked = false;
  document.getElementById('mapaResultado').style.display = 'block';
  status.textContent = '';

  // Copiar destino al campo lugar
  var lugarParts = destino.split(',');
  document.getElementById('lugar').value = lugarParts[0].trim() || destino;

  // Mostrar mapa
  var mapDiv = document.getElementById('mapContainer');
  mapDiv.style.display = 'block';
  _loadLeaflet(function(){
    if(!_mapaMap){
      _mapaMap = L.map('mapContainer');
      L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',{
        attribution:'© OpenStreetMap'}).addTo(_mapaMap);
    }
    if(_mapaRouteLayer) _mapaMap.removeLayer(_mapaRouteLayer);
    var geo = rdata.routes[0].geometry;
    _mapaRouteLayer = L.geoJSON(geo, {style:{color:'#2980b9',weight:4}}).addTo(_mapaMap);
    L.marker([o.lat,o.lon]).addTo(_mapaMap).bindPopup(origen);
    L.marker([d.lat,d.lon]).addTo(_mapaMap).bindPopup(destino).openPopup();
    _mapaMap.fitBounds(_mapaRouteLayer.getBounds(), {padding:[20,20]});
  });
}

async function guardarKmMapa(){
  var cliente = document.getElementById('selCliente')?.value || '';
  var viaje   = document.getElementById('selViaje')?.value  || '';
  if(!viaje || viaje==='__nuevo__'){
    document.getElementById('mapaGuardadoMsg').textContent = '⚠️ Guarda primero el viaje';
    document.getElementById('mapaGuardadoMsg').style.display='inline';
    return;
  }
  var casetas = parseFloat(document.getElementById('mapaCasetas').value)||0;
  var lugar   = document.getElementById('lugar').value;
  var resp = await fetch('/api/guardar_km_mapa', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({cliente:cliente, viaje:viaje,
                          km_total:_mapaKmCalculado, casetas:casetas, lugar:lugar})
  });
  var data = await resp.json();
  var msg = document.getElementById('mapaGuardadoMsg');
  if(data.ok){
    msg.textContent = '✓ Guardado — recargando…';
    msg.style.color = '#27ae60';
    msg.style.display='inline';
    setTimeout(function(){ location.reload(); }, 800);
  } else {
    msg.textContent = '❌ ' + (data.error||'Error');
    msg.style.color = '#e74c3c';
    msg.style.display='inline';
  }
}
</script>
</body></html>"""

# ─── REVISIÓN / APROBACIÓN ────────────────────────────────────────────────────

@app.route("/revisar/<sid>")
@login_required
def revisar(sid):
    s = get_submission(sid)
    if not s:
        return "No encontrado", 404
    rol = session["rol"]
    nombre = session["nombre"]
    # Trabajador solo puede ver sus propias solicitudes
    if rol != "admin" and s.get("beneficiario") != nombre:
        return redirect(url_for("dashboard"))
    return render_template_string(REVISAR_HTML, s=s, rol=rol)

REVISAR_HTML = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><title>Revisión — Viáticos V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f4f6f9;font-size:14px;color:#333}
header{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
       justify-content:space-between;align-items:center;font-size:15px;font-weight:bold}
.hright a{color:#adc8e6;text-decoration:none;font-size:13px}
.container{max-width:1000px;margin:20px auto;padding:0 16px 40px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);
      margin-bottom:14px;overflow:hidden}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;font-weight:bold;font-size:13px}
.card-body{padding:16px}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.field{margin-bottom:8px}
.field label{font-size:11px;color:#777;display:block}
.field span{font-size:14px;font-weight:bold;color:#222}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1a3a5c;color:white;padding:8px 12px;text-align:left}
td{padding:8px 12px;border-bottom:1px solid #eee}
.tag{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}
.tag-Hotel{background:#d6eaf8;color:#1a5276}
.tag-Vuelo{background:#d5f5e3;color:#1e8449}
.tag-Comida{background:#fdebd0;color:#a04000}
.tag-Transporte{background:#e8daef;color:#6c3483}
.tag-Autobús{background:#d0ece7;color:#0e6655}
.tag-Casetas,.tag-Otro{background:#f2f3f4;color:#555;border:1px solid #bbb}
.tag-Gasolina{background:#fef9e7;color:#7d6608}
.sum-box{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px}
.sum{flex:1;min-width:180px;border-radius:8px;padding:14px;text-align:center}
.sum-emp{background:#d6eaf8}
.sum-cli{background:#fdebd0}
.sum-label{font-size:11px;color:#666;margin-bottom:4px}
.sum-amt{font-size:22px;font-weight:bold;color:#1a3a5c}
.btn{padding:9px 20px;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:bold}
.btn-success{background:#27ae60;color:white}
.btn-danger{background:#e74c3c;color:white}
.btn-primary{background:#1a3a5c;color:white}
.btn-secondary{background:#7f8c8d;color:white}
.btn:hover{opacity:.88}
.btn-row{display:flex;gap:10px}
textarea{width:100%;padding:8px;border:1px solid #ccc;border-radius:5px;
         font-size:13px;font-family:Arial,sans-serif;resize:vertical}
.badge{display:inline-block;padding:4px 12px;border-radius:12px;font-size:12px;font-weight:bold}
.badge-pend{background:#fff3cd;color:#856404}
.badge-apro{background:#d1e7dd;color:#0f5132}
.badge-rech{background:#f8d7da;color:#842029}
.alert{padding:10px 14px;border-radius:6px;font-size:13px;margin-top:10px}
.alert-ok{background:#d1e7dd;color:#0f5132}
.alert-err{background:#f8d7da;color:#842029}
</style></head>
<body>
<header>
  ⚖️ Revisión de viaje: {{ "💰 Anticipo" if s.get("tipo_solicitud")=="anticipo" else "📋 Comprobación" }} {{ s.viaje or s.ciudad or "" }}
  <div class="hright"><a href="/dashboard">← Dashboard</a></div>
</header>
<div class="container">

  <!-- Info general -->
  <div class="card">
    <div class="card-header">
      Solicitud #{{ s.id }} &nbsp;
      <span class="badge badge-{{ 'pend' if s.status=='pendiente' else ('apro' if s.status=='aprobado' else 'rech') }}">
        {{ s.status.upper() }}
      </span>
      &nbsp; Enviado: {{ s.fecha_envio }}
    </div>
    <div class="card-body">
      <div class="grid-2">
        <div>
          <div class="field"><label>Beneficiario</label><span>{{ s.beneficiario }}</span></div>
          <div class="field"><label>Con cargo a</label><span>{{ s.cargo }}</span></div>
          <div class="field"><label>Lugar</label><span>{{ s.lugar }}</span></div>
        </div>
        <div>
          <div class="field"><label>Fecha salida</label><span>{{ s.salida }}</span></div>
          <div class="field"><label>Fecha regreso</label><span>{{ s.regreso }}</span></div>
          <div class="field"><label>Asunto</label><span>{{ s.asunto }}</span></div>
        </div>
      </div>
      <div class="field" style="margin-top:8px">
        <label>Descripción</label><span>{{ s.descripcion }}</span></div>
    </div>
  </div>

  <!-- Conceptos / Gastos -->
  <div class="card">
    <div class="card-header">Desglose de conceptos solicitados</div>
    <div class="card-body" style="padding:0;overflow-x:auto">
      <table style="min-width:500px">
        <thead><tr>
          <th>#</th>
          <th>Concepto</th>
          <th style="text-align:right">Monto solicitado</th>
          {% if rol=='admin' %}
          <th style="text-align:right">Monto aprobado</th>
          {% endif %}
        </tr></thead>
        <tbody>
        {% set items = s.conceptos if s.conceptos else s.gastos %}
        {% for g in items %}
        <tr>
          <td>{{ loop.index }}</td>
          <td>{{ g.concepto if g.concepto is defined else (g.descripcion or g.archivo) }}</td>
          <td style="text-align:right;font-weight:bold;color:#1a5276">
            ${{ "%.2f"|format(g.monto if g.monto is defined else g.empleado) }}
          </td>
          {% if rol=='admin' %}
          <td style="text-align:right">
            {% if s.status == 'pendiente' %}
            <input type="number" step="0.01" min="0"
                   id="aprobado_{{ loop.index0 }}"
                   value="{{ "%.2f"|format(g.monto if g.monto is defined else g.empleado) }}"
                   style="width:100px;text-align:right;padding:4px 6px;border:1px solid #ccc;border-radius:4px;font-size:13px">
            {% else %}
            <strong>${{ "%.2f"|format(g.get('aprobado', g.get('monto', g.get('empleado',0)))) }}</strong>
            {% endif %}
          </td>
          {% endif %}
        </tr>
        {% endfor %}
        </tbody>
        <tfoot>
          <tr style="background:#f0f4f8;font-weight:bold">
            <td colspan="2" style="text-align:right;padding:10px 12px">TOTAL SOLICITADO:</td>
            <td style="text-align:right;padding:10px 12px;color:#1a5276;font-size:15px">
              ${{ "%.2f"|format(s.total_empleado) }}
            </td>
            {% if rol=='admin' %}
            <td style="text-align:right;padding:10px 12px;font-size:15px" id="total_aprobado_cell">
              {% if s.status == 'pendiente' %}
              ${{ "%.2f"|format(s.total_empleado) }}
              {% else %}
              ${{ "%.2f"|format(s.get('total_aprobado', s.total_empleado)) }}
              {% endif %}
            </td>
            {% endif %}
          </tr>
        </tfoot>
      </table>
    </div>
  </div>

  <!-- Totales -->
  <div class="card">
    <div class="card-header">Resumen</div>
    <div class="card-body">
      <div class="sum-box">
        <div class="sum sum-emp">
          <div class="sum-label">Total gastos del empleado</div>
          <div class="sum-amt">${{ "%.2f"|format(s.total_empleado) }}</div>
        </div>
        {% if rol=='admin' and s.get('tipo_solicitud') != 'anticipo' %}
        <div class="sum sum-cli">
          <div class="sum-label">Total a cobrar al cliente</div>
          <div class="sum-amt">${{ "%.2f"|format(s.total_cliente) }}</div>
        </div>
        {% endif %}
        {% if s.get('anticipo_id') %}
        <div class="sum" style="background:#fef9e7;border:1px solid #f0c040">
          <div class="sum-label">Anticipo aprobado (#{{ s.anticipo_id }})</div>
          <div class="sum-amt" style="color:#7d6608">${{ "%.2f"|format(s.anticipo_monto or 0) }}</div>
        </div>
        {% set dif = s.total_empleado - (s.anticipo_monto or 0) %}
        <div class="sum" style="background:{% if dif > 0 %}#d1e7dd{% elif dif < 0 %}#f8d7da{% else %}#f2f3f4{% endif %}">
          <div class="sum-label">
            {% if dif > 0 %}💚 La Firma paga al colaborador{% elif dif < 0 %}🔴 Colaborador reembolsa{% else %}✓ Sin diferencia{% endif %}
          </div>
          <div class="sum-amt" style="color:{% if dif > 0 %}#0f5132{% elif dif < 0 %}#842029{% else %}#555{% endif %}">
            ${{ "%.2f"|format(dif|abs) }}
          </div>
        </div>
        {% endif %}
      </div>

      {% if rol=='admin' and s.status=='pendiente' %}
      <div style="margin-bottom:12px">
        <label style="font-size:12px;color:#555;display:block;margin-bottom:4px">
          Comentario (requerido si se rechaza)
        </label>
        <textarea id="comentario" rows="2" placeholder="Observaciones de la revisión..."></textarea>
      </div>
      <div style="font-size:12px;color:#555;margin-bottom:8px">
        ℹ️ Puede modificar el monto aprobado por concepto en la tabla de arriba antes de aprobar.
      </div>
      <div class="btn-row">
        <button class="btn btn-success" onclick="accion('aprobar')">✅ Aprobar</button>
        <button class="btn btn-danger"  onclick="accion('rechazar')">❌ Rechazar</button>
        <a href="/dashboard" class="btn btn-secondary">Volver</a>
      </div>
      {% elif rol=='admin' %}
      <div class="btn-row">
        <a href="/api/excel/{{ s.id }}" class="btn btn-primary">📄 Descargar Excel</a>
        <a href="/dashboard" class="btn btn-secondary">Volver</a>
      </div>
      {% else %}
      <div class="btn-row">
        {% if s.status=='aprobado' %}
        <a href="/api/excel/{{ s.id }}" class="btn btn-primary">📄 Descargar Excel</a>
        {% endif %}
        <a href="/dashboard" class="btn btn-secondary">Volver</a>
      </div>
      {% endif %}

      <div id="msgResult"></div>
    </div>
  </div>

</div>
{% if rol=='admin' and s.status=='pendiente' %}
<script>
// Actualizar total aprobado al cambiar los inputs
document.querySelectorAll('input[id^="aprobado_"]').forEach(inp => {
  inp.addEventListener('input', calcTotal);
});
function calcTotal(){
  let t = 0;
  document.querySelectorAll('input[id^="aprobado_"]').forEach(inp => {
    t += parseFloat(inp.value) || 0;
  });
  const cell = document.getElementById('total_aprobado_cell');
  if(cell) cell.textContent = '$' + t.toFixed(2);
}
calcTotal();

function accion(tipo){
  const comentario = document.getElementById('comentario').value;
  if(tipo==='rechazar' && !comentario.trim()){
    alert('Indique el motivo del rechazo en el comentario.'); return;
  }
  // Recopilar montos aprobados por concepto
  const montos = {};
  document.querySelectorAll('input[id^="aprobado_"]').forEach(inp => {
    const idx = inp.id.replace('aprobado_','');
    montos[idx] = parseFloat(inp.value) || 0;
  });
  fetch('/api/'+tipo+'/{{ s.id }}', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({comentario, montos_aprobados: montos})
  }).then(r=>r.json()).then(data=>{
    if(data.ok){
      document.getElementById('msgResult').innerHTML =
        '<div class="alert alert-ok">'+
        (tipo==='aprobar'?'✅ Aprobado. ':'❌ Rechazado. ')+
        data.msg+'</div>';
      setTimeout(()=>window.location='/dashboard', 2000);
    } else {
      document.getElementById('msgResult').innerHTML =
        '<div class="alert alert-err">Error: '+data.error+'</div>';
    }
  });
}
</script>
{% endif %}
</body></html>"""

# ─── API ENDPOINTS ────────────────────────────────────────────────────────────

@app.route("/api/clientes")
@login_required
def api_clientes():
    clients = sorted([
        nfc(d) for d in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, d)) and nfc(d).startswith("Viáticos ")
    ])
    return jsonify(clients)

@app.route("/api/viajes")
@login_required
def api_viajes():
    cliente = request.args.get("cliente","")
    client_dir = find_dir(BASE_DIR, cliente)
    if not client_dir:
        return jsonify([])
    trips = sorted([
        nfc(d) for d in os.listdir(client_dir)
        if os.path.isdir(os.path.join(client_dir, d)) and nfc(d).startswith("Viaje")
    ])
    return jsonify(trips)

@app.route("/api/cargar", methods=["POST"])
@login_required
def api_cargar():
    body      = request.json
    cliente   = body.get("cliente","")
    viaje_key = body.get("viaje_key","")   # clave única del viaje (para carpeta de uploads)
    ciudad    = body.get("ciudad","")      # ciudad destino (para el calendario)
    fecha_ini = body.get("fecha_ini","")   # hint de fecha para desambiguar

    usuario = session["usuario"]

    # ── Tabulador (grupo correcto según cliente) ─────────────────────────────
    tabulador = _get_tabulador(cliente) if cliente else {}

    # ── Metadatos del viaje desde el calendario ───────────────────────────────
    info = {"cargo": cliente.replace("Viáticos ",""), "lugar": ciudad, "cal_found": False}
    if ciudad:
        try:
            cal_file = None
            search_dirs = ([client_dir] if client_dir else []) + [BASE_DIR]
            for sd in search_dirs:
                for f in os.listdir(sd):
                    if "calendario" in nfc(f).lower() and f.lower().endswith(".xlsx"):
                        cal_file = os.path.join(sd, f); break
                if cal_file: break
            if not cal_file and os.path.exists(CALENDAR):
                cal_file = CALENDAR
            cal_records = load_calendar(cal_file) if cal_file else []
            nombre_usuario = session.get("nombre","")
            if nombre_usuario and session.get("rol") != "admin":
                user_plain = _strip_acc(nombre_usuario)
                propios = [r for r in cal_records if _strip_acc(r.get("abogado","")) == user_plain]
                if propios:
                    cal_records = propios
            hint = [fecha_ini] if fecha_ini else None
            cal_info = lookup_calendar(ciudad, cal_records, hint_dates=hint)
            if cal_info:
                info.update({
                    "abogado":   cal_info["abogado"],
                    "salida":    cal_info["salida"],
                    "regreso":   cal_info["regreso"],
                    "lugar":     cal_info["lugar"],
                    "motivo":    cal_info["descripcion"],
                    "cal_found": True,
                })
        except Exception as e:
            print(f"[cal] {e}")

    # ── Comidas sugeridas ─────────────────────────────────────────────────────
    city_data = lookup_city(ciudad, tabulador) if ciudad else None
    if not city_data:
        city_data = _city_data_fallback(body.get("km_mapa"), body.get("casetas_mapa", 0))
    meals_sug = meals_by_distance(city_data["km_total"]) if city_data else 1
    info["meals_sugeridas"] = meals_sug
    try:
        from datetime import datetime as _dt
        fi = _dt.strptime(info.get("salida",""), "%d/%m/%Y")
        ff = _dt.strptime(info.get("regreso",""), "%d/%m/%Y")
        info["dias"] = max(int((ff-fi).days)+1, 1)
    except:
        info["dias"] = 1

    # ── Comprobantes subidos por el usuario ───────────────────────────────────
    trip_dir = get_upload_dir(usuario, cliente, viaje_key)
    gastos = []
    for fname in sorted(f for f in os.listdir(trip_dir)
                        if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))):
        fpath = os.path.join(trip_dir, fname)
        cat   = detect_category(fname)
        if fname.lower().endswith(".pdf"):
            d = extract_pdf_data(fpath)
            amount, rfc = d["amount"], d["rfc"]
        else:
            amount, rfc = 0.0, ""
        causa_iva = cat not in {"Hotel","Autobús","Vuelo","Otro"}
        emp, cli  = calc_row(cat, amount, city_data, meals_sug)
        gastos.append({
            "archivo":     fname,
            "tipo":        cat,
            "rfc":         rfc,
            "descripcion": os.path.splitext(fname)[0],
            "con_iva":     round(amount,2) if causa_iva else 0.0,
            "sin_iva":     0.0 if causa_iva else round(amount,2),
            "empleado":    round(emp,2),
            "cliente":     round(cli,2),
        })
    return jsonify({"info": info, "gastos": gastos})

@app.route("/api/enviar", methods=["POST"])
@login_required
def api_enviar():
    data = request.json
    sid  = save_submission(data)
    return jsonify({"ok": True, "id": sid})

@app.route("/api/aprobar/<sid>", methods=["POST"])
@login_required
def api_aprobar(sid):
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permiso"}), 403
    body = request.json or {}
    montos = body.get("montos_aprobados", {})

    # Guardar montos aprobados por concepto en los gastos de la solicitud
    s = get_submission(sid)
    if s and montos:
        gastos_upd = s.get("gastos", [])
        for i, g in enumerate(gastos_upd):
            k = str(i)
            if k in montos:
                g["aprobado"] = float(montos[k])
        total_aprobado = sum(g.get("aprobado", g.get("empleado",0)) for g in gastos_upd)
        update_submission(sid, {"gastos": gastos_upd, "total_aprobado": round(total_aprobado, 2)})

    ok = update_submission(sid, {"status": "aprobado",
                                  "comentario": body.get("comentario",""),
                                  "aprobado_por": session["nombre"],
                                  "fecha_aprobacion": datetime.now().strftime("%d/%m/%Y %H:%M")})
    if ok:
        return jsonify({"ok": True, "msg": "Puede descargar el Excel desde el detalle."})
    return jsonify({"error": "No encontrado"})

@app.route("/api/rechazar/<sid>", methods=["POST"])
@login_required
def api_rechazar(sid):
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permiso"}), 403
    body = request.json or {}
    ok = update_submission(sid, {"status": "rechazado",
                                  "comentario": body.get("comentario","")})
    if ok:
        return jsonify({"ok": True, "msg": "Solicitud rechazada."})
    return jsonify({"error": "No encontrado"})

@app.route("/api/excel/<sid>")
@login_required
def api_excel(sid):
    s = get_submission(sid)
    if not s:
        return "No encontrado", 404
    if session["rol"] != "admin" and s.get("beneficiario") != session["nombre"]:
        return "Sin permiso", 403

    client_dir = find_dir(BASE_DIR, s.get("cliente",""))
    trip_dir   = find_dir(client_dir, s.get("viaje","")) if client_dir else None
    if not trip_dir:
        trip_dir = BASE_DIR

    fecha_hoy = datetime.now().strftime("%Y%m%d")
    out_name  = f"Comprobación {s.get('viaje','')} - {fecha_hoy}.xlsx"
    out_path  = os.path.join(trip_dir, out_name)
    write_excel(out_path, s)
    return send_file(out_path, as_attachment=True, download_name=out_name)

# ─── NÚMERO A LETRAS (pesos mexicanos) ───────────────────────────────────────

def numero_a_letras(n):
    """Convierte un número a su representación en letras en español (pesos M.N.)"""
    from math import floor
    unidades  = ["","UN","DOS","TRES","CUATRO","CINCO","SEIS","SIETE","OCHO","NUEVE",
                 "DIEZ","ONCE","DOCE","TRECE","CATORCE","QUINCE","DIECISÉIS",
                 "DIECISIETE","DIECIOCHO","DIECINUEVE"]
    decenas   = ["","DIEZ","VEINTE","TREINTA","CUARENTA","CINCUENTA",
                 "SESENTA","SETENTA","OCHENTA","NOVENTA"]
    centenas  = ["","CIEN","DOSCIENTOS","TRESCIENTOS","CUATROCIENTOS","QUINIENTOS",
                 "SEISCIENTOS","SETECIENTOS","OCHOCIENTOS","NOVECIENTOS"]

    def bloque(num):
        if num == 0: return ""
        if num == 100: return "CIEN"
        c = num // 100
        resto = num % 100
        cent_str = ("CIENTO" if c == 1 and resto > 0 else centenas[c])
        r = cent_str + (" " if c and resto else "")
        if resto < 20:
            r += unidades[resto]
        else:
            d = resto // 10
            u = resto % 10
            r += decenas[d]
            if u: r += " Y " + unidades[u]
        return r.strip()

    n = round(float(n), 2)
    entero  = int(floor(n))
    centavos = round((n - entero) * 100)

    if entero == 0:
        texto = "CERO"
    elif entero < 1000:
        texto = bloque(entero)
    elif entero < 1000000:
        miles = entero // 1000
        resto = entero % 1000
        texto = ("MIL" if miles == 1 else bloque(miles) + " MIL")
        if resto: texto += " " + bloque(resto)
    else:
        millones = entero // 1000000
        resto    = entero % 1000000
        texto = ("UN MILLÓN" if millones == 1 else bloque(millones) + " MILLONES")
        if resto:
            miles = resto // 1000
            r2    = resto % 1000
            if miles: texto += " " + ("MIL" if miles==1 else bloque(miles)+" MIL")
            if r2:    texto += " " + bloque(r2)

    return f"{texto} PESOS {centavos:02d}/100 M.N."

# ─── CÁLCULO DE ANTICIPO ──────────────────────────────────────────────────────

def calcular_anticipo(ciudad, fecha_salida_str, fecha_regreso_str,
                      modo_transporte, costo_transporte_estimado,
                      incluye_hotel, cliente, tabulador,
                      traslados=None, meals_por_dia=None,
                      casetas_monto=None, km_mapa=None, casetas_mapa=0.0):
    """
    Calcula el desglose del anticipo a solicitar.
    Devuelve dict con conceptos y total.
    """
    from datetime import datetime as dt

    def parse_f(s):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try: return dt.strptime(s, fmt)
            except: pass
        return None

    fi = parse_f(fecha_salida_str)
    ff = parse_f(fecha_regreso_str)
    if not fi or not ff:
        return None

    dias   = max((ff - fi).days + 1, 1)   # días fuera incluyendo ambos extremos
    noches = max((ff - fi).days, 0)
    # Si marcó hotel, garantizar al menos 1 noche aunque las fechas sean iguales
    if incluye_hotel and noches == 0:
        noches = 1

    city_data = lookup_city(ciudad, tabulador)
    if not city_data and km_mapa:
        city_data = _city_data_fallback(km_mapa, casetas_mapa)
    km       = city_data["km"]       if city_data else None
    km_total = city_data["km_total"] if city_data else None

    conceptos = []
    total     = 0.0

    # — Transporte —
    if modo_transporte == "automovil" and km_total:
        km_total = city_data["km_total"]
        monto_trans = round(km_total * EMPLOYEE_KM_RATE, 2)
        conceptos.append({
            "concepto": f"Gasolina (automóvil, {km_total:.0f} km × ${EMPLOYEE_KM_RATE})",
            "monto": monto_trans
        })
        total += monto_trans
        # — Casetas —
        if casetas_monto is None:
            casetas_monto = lookup_casetas(ciudad)
        casetas_monto = round(float(casetas_monto), 2)
        if casetas_monto > 0:
            conceptos.append({
                "concepto": "Casetas (ida y vuelta)",
                "monto": casetas_monto,
                "editable": True,
            })
            total += casetas_monto
    elif modo_transporte in ("avion","autobus","otro") and costo_transporte_estimado:
        monto_trans = float(costo_transporte_estimado)
        label = {"avion":"Vuelo (estimado)","autobus":"Autobús (estimado)"}.get(
            modo_transporte, "Transporte (estimado)")
        conceptos.append({"concepto": label, "monto": monto_trans})
        total += monto_trans

    # — Traslados locales (aplica principalmente en viajes de avión) —
    if traslados:
        for t in traslados:
            monto_t = float(t.get("monto", 0))
            if monto_t > 0:
                conceptos.append({"concepto": t.get("descripcion","Traslado"),
                                   "monto": round(monto_t, 2)})
                total += monto_t

    # — Comidas por día —
    # meals_por_dia: lista con el nº de comidas de cada día, ej. [3, 1]
    # Si no se pasa, se calcula automáticamente por distancia para todos los días
    n_default = meals_by_distance(km_total) if km_total else 1
    if not meals_por_dia or len(meals_por_dia) == 0:
        meals_por_dia = [n_default] * dias
    # Ajustar si la lista tiene menos elementos que los días del viaje
    while len(meals_por_dia) < dias:
        meals_por_dia.append(n_default)
    meals_por_dia = meals_por_dia[:dias]   # recortar si sobran

    # Agrupar días con el mismo nº de comidas para presentación compacta
    from itertools import groupby as _gb
    grupos = {}
    for idx, n in enumerate(meals_por_dia):
        grupos.setdefault(n, []).append(idx + 1)

    for n_meals, dias_idx in sorted(grupos.items()):
        nd = len(dias_idx)
        monto = nd * n_meals * EMPLOYEE_MEAL_RATE
        if len(grupos) == 1:
            label = f"Comidas ({nd} día(s) × {n_meals} comida(s)/día × ${EMPLOYEE_MEAL_RATE:.0f})"
        else:
            dias_str = ", ".join(str(d) for d in dias_idx)
            label = f"Comidas día(s) {dias_str} × {n_meals} comida(s) × ${EMPLOYEE_MEAL_RATE:.0f}"
        conceptos.append({"concepto": label, "monto": round(monto, 2)})
        total += monto

    # — Hotel —
    if incluye_hotel and noches > 0:
        monto_hotel = noches * EMPLOYEE_HOTEL_RATE
        conceptos.append({
            "concepto": f"Hospedaje ({noches} noche(s) × ${EMPLOYEE_HOTEL_RATE:.0f})",
            "monto": round(monto_hotel, 2)
        })
        total += monto_hotel

    return {
        "conceptos":       conceptos,
        "total":           round(total, 2),
        "dias":            dias,
        "noches":          noches,
        "km":              km,
        "meals_por_dia_out": meals_por_dia,
        "meals_dia":       meals_por_dia[0] if meals_por_dia else 1,
    }

# ─── ESCRITURA V&V-02 (SOLICITUD DE ANTICIPO) ────────────────────────────────

def fix_external_refs(wb):
    """Elimina referencias a libros externos [N] dejando solo la referencia interna."""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and re.search(r"\[\d+\]", cell.value):
                    cell.value = re.sub(r"='\[(\d+)\]([^']+)'!", r"='\2'!", cell.value)

def write_anticipo_excel(output_path, data):
    shutil.copy2(TEMPLATE_XLS, output_path)
    wb = load_workbook(output_path, keep_links=False)
    fix_external_refs(wb)  # limpiar referencias externas rotas
    ws = wb["Solicitud de Anticipo"]

    ws["C9"]  = data.get("beneficiario", "")      # BENEFICIARIO
    ws["C11"] = data.get("total", 0)               # MONTO DEL ANTICIPO
    ws["C13"] = numero_a_letras(data.get("total", 0))  # IMPORTE CON LETRA
    ws["A22"] = data.get("salida", "")             # FECHA DE SALIDA
    ws["G22"] = data.get("regreso", "")            # FECHA DE REGRESO
    ws["D25"] = "✓"                                # CON CARGO A: CLIENTE
    ws["B29"] = data.get("cargo", "")              # NOMBRE DEL CLIENTE
    ws["G29"] = data.get("expediente", "")         # NÚM DE EXPEDIENTE
    ws["B31"] = data.get("asunto", "")             # ASUNTO
    ws["B33"] = data.get("lugar", "")              # LUGAR
    ws["C35"] = data.get("descripcion", "")        # DESCRIPCIÓN

    wb.save(output_path)
    return output_path

# ─── RUTA: SOLICITUD DE ANTICIPO ─────────────────────────────────────────────

@app.route("/anticipo")
@login_required
def anticipo_page():
    clientes = sorted([
        nfc(d) for d in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, d)) and nfc(d).startswith("Viáticos ")
    ])
    return render_template_string(ANTICIPO_HTML,
        nombre=session["nombre"], rol=session["rol"], clientes=clientes)

@app.route("/api/casetas")
@login_required
def api_casetas():
    ciudad = request.args.get("ciudad","")
    monto  = lookup_casetas(ciudad) if ciudad else 0.0
    return jsonify({"ciudad": ciudad, "monto": monto})

@app.route("/api/guardar_km_mapa", methods=["POST"])
@login_required
def api_guardar_km_mapa():
    """Guarda km y casetas calculados desde el mapa en _config_viaje.json del viaje activo."""
    body    = request.get_json(force=True) or {}
    cliente = body.get("cliente", "")
    viaje   = body.get("viaje", "")
    km_total = float(body.get("km_total", 0))
    casetas  = float(body.get("casetas", 0))
    lugar    = body.get("lugar", "")
    if not cliente or not viaje:
        return jsonify({"ok": False, "error": "cliente/viaje requeridos"}), 400
    trip_dir = get_upload_dir(session["usuario"], cliente, viaje)
    os.makedirs(trip_dir, exist_ok=True)
    cfg_path = os.path.join(trip_dir, "_config_viaje.json")
    cfg = {}
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, encoding="utf-8") as f: cfg = json.load(f)
        except Exception: pass
    cfg["km_mapa"]      = km_total
    cfg["casetas_mapa"] = casetas
    if lugar: cfg["lugar"] = lugar
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True})

@app.route("/api/calcular_anticipo", methods=["POST"])
@login_required
def api_calcular_anticipo():
    body = request.json
    cliente  = body.get("cliente", "")
    ciudad   = body.get("ciudad", "")
    tabulador = _get_tabulador(cliente) if cliente else {}

    # casetas: si el cliente manda un valor explícito lo usamos; si manda null buscamos en config
    casetas_input = body.get("casetas_monto")  # None = auto-lookup, 0 = sin casetas, >0 = manual
    resultado = calcular_anticipo(
        ciudad              = ciudad,
        fecha_salida_str    = body.get("salida", ""),
        fecha_regreso_str   = body.get("regreso", ""),
        modo_transporte     = body.get("modo_transporte", "automovil"),
        costo_transporte_estimado = body.get("costo_transporte", 0),
        incluye_hotel       = body.get("incluye_hotel", False),
        cliente             = cliente,
        tabulador           = tabulador,
        traslados           = body.get("traslados", []),
        meals_por_dia       = body.get("meals_por_dia", []),
        casetas_monto       = casetas_input,
        km_mapa             = body.get("km_mapa"),
        casetas_mapa        = body.get("casetas_mapa", 0.0),
    )
    if resultado:
        resultado["meals_por_dia"] = resultado.get("meals_por_dia_out", [])
    if not resultado:
        return jsonify({"error": "Fechas inválidas"})

    city_data = lookup_city(ciudad, tabulador)
    resultado["en_tabulador"] = city_data is not None
    resultado["ciudad_tabulador"] = list(tabulador.keys())[
        list(tabulador.values()).index(city_data)] if city_data else ""
    # Siempre devolver el valor de casetas sugerido para que el form lo muestre
    resultado["casetas_sugeridas"] = lookup_casetas(ciudad)
    return jsonify(resultado)

@app.route("/api/generar_anticipo", methods=["POST"])
@login_required
def api_generar_anticipo():
    body     = request.json
    cliente  = body.get("cliente", "")
    ciudad   = body.get("ciudad", "").replace(" ","_")
    fecha    = datetime.now().strftime("%Y%m%d")
    out_name = f"Solicitud Anticipo {ciudad} - {fecha}.xlsx"

    client_dir = find_dir(BASE_DIR, cliente)
    out_dir    = client_dir if client_dir else BASE_DIR
    out_path   = os.path.join(out_dir, out_name)

    try:
        write_anticipo_excel(out_path, body)
    except Exception as e:
        return jsonify({"error": str(e)})

    return jsonify({"ok": True, "path": out_path, "filename": out_name})

@app.route("/api/descargar_anticipo")
@login_required
def api_descargar_anticipo():
    path = request.args.get("path", "")
    if not os.path.exists(path):
        return "Archivo no encontrado", 404
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path))

ANTICIPO_HTML = """<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Solicitud de Anticipo — Viáticos V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f4f6f9;font-size:14px;color:#333}
header{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
       justify-content:space-between;align-items:center;font-size:15px;font-weight:bold}
.hright a{color:#adc8e6;text-decoration:none;font-size:13px;margin-left:16px}
.container{max-width:800px;margin:20px auto;padding:0 16px 40px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);
      margin-bottom:14px;overflow:hidden}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;
             font-weight:bold;font-size:13px}
.card-body{padding:18px}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
label{display:block;font-size:12px;color:#555;margin-bottom:4px;font-weight:bold}
input,select,textarea{width:100%;padding:9px 11px;border:1px solid #ccc;
                      border-radius:5px;font-size:13px;font-family:Arial,sans-serif}
input:focus,select:focus{outline:none;border-color:#1a3a5c;
                          box-shadow:0 0 0 2px rgba(26,58,92,.15)}
.radio-group{display:flex;gap:12px;flex-wrap:wrap;margin-top:4px}
.radio-group label{display:flex;align-items:center;gap:5px;font-weight:normal;
                   cursor:pointer;padding:7px 14px;border:1px solid #ccc;
                   border-radius:20px;font-size:13px;transition:all .2s}
.radio-group input[type=radio]{width:auto;margin:0}
.radio-group label:has(input:checked){background:#1a3a5c;color:white;border-color:#1a3a5c}
.btn{padding:10px 22px;border:none;border-radius:6px;cursor:pointer;
     font-size:13px;font-weight:bold;text-decoration:none;display:inline-block}
.btn-primary{background:#1a3a5c;color:white}
.btn-success{background:#27ae60;color:white}
.btn-secondary{background:#7f8c8d;color:white}
.btn:hover{opacity:.88}
.btn-row{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px}
.resumen{background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:16px}
.concepto-row{display:flex;justify-content:space-between;padding:7px 0;
              border-bottom:1px solid #eee;font-size:13px}
.concepto-row:last-child{border-bottom:none}
.concepto-label{color:#555}
.concepto-monto{font-weight:bold;color:#1a3a5c}
.total-row{display:flex;justify-content:space-between;padding:10px 0;
           margin-top:8px;border-top:2px solid #1a3a5c}
.total-label{font-weight:bold;font-size:15px;color:#1a3a5c}
.total-monto{font-weight:bold;font-size:20px;color:#1a3a5c}
.total-letras{font-size:12px;color:#777;margin-top:6px;font-style:italic}
.alert{padding:10px 14px;border-radius:6px;font-size:13px;margin-top:12px}
.alert-ok{background:#d1e7dd;color:#0f5132}
.alert-warn{background:#fff3cd;color:#856404}
.alert-err{background:#f8d7da;color:#842029}
.hidden{display:none}
#loading{display:none;position:fixed;inset:0;background:rgba(0,0,0,.3);
         z-index:99;align-items:center;justify-content:center}
.spin-box{background:white;border-radius:10px;padding:28px 36px;text-align:center}
.spin{width:36px;height:36px;border:4px solid #ddd;border-top-color:#1a3a5c;
      border-radius:50%;animation:sp .8s linear infinite;margin:0 auto 10px}
@keyframes sp{to{transform:rotate(360deg)}}
</style>
</head>
<body>
<div id="loading"><div class="spin-box"><div class="spin"></div>Calculando…</div></div>

<header>
  ⚖️ Solicitud de Anticipo de Viáticos
  <div class="hright">
    <a href="/dashboard">← Dashboard</a>
    <a href="/logout">Salir</a>
  </div>
</header>

<div class="container">

  <!-- DATOS DEL VIAJE -->
  <div class="card">
    <div class="card-header">Datos del viaje</div>
    <div class="card-body">
      <div class="grid-2" style="margin-bottom:14px">
        <div>
          <label>Solicitante</label>
          <input id="beneficiario" value="{{ nombre }}" readonly
                 style="background:#f8f9fa;color:#555">
        </div>
        <div>
          <label>Cliente (con cargo a)</label>
          <select id="selCliente" onchange="cargarTabulador()">
            <option value="">— Seleccione cliente —</option>
            {% for c in clientes %}
            <option value="{{ c }}">{{ c }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Ciudad destino</label>
          <input id="ciudad" type="text" placeholder="ej. Guadalajara"
                 oninput="limpiarResumen(); buscarCasetas()">
          <div style="margin-top:4px">
            <button type="button" onclick="toggleMapaAnticipo()"
                    style="background:none;border:none;color:#2980b9;font-size:12px;
                           cursor:pointer;padding:0;text-decoration:underline">
              🗺 Calcular distancia con mapa
            </button>
          </div>
        </div>
        <div id="mapaAnticipo" style="display:none;grid-column:1/-1;background:#eaf4fb;
              border:1px solid #7fb3d3;border-radius:8px;padding:10px">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px">
            <div>
              <label style="font-size:12px">Origen</label>
              <input id="mapaAntOrigen" type="text" value="Querétaro, Qro."
                     style="width:100%;padding:5px;border:1px solid #aad4ee;border-radius:4px;font-size:12px;box-sizing:border-box">
            </div>
            <div>
              <label style="font-size:12px">Destino</label>
              <input id="mapaAntDestino" type="text" placeholder="Ciudad destino"
                     style="width:100%;padding:5px;border:1px solid #aad4ee;border-radius:4px;font-size:12px;box-sizing:border-box">
            </div>
          </div>
          <button type="button" onclick="calcularRutaAnticipo()"
                  style="background:#2980b9;color:white;border:none;border-radius:4px;
                         padding:6px 14px;font-size:12px;cursor:pointer">
            📐 Calcular
          </button>
          <span id="mapaAntStatus" style="font-size:12px;color:#888;margin-left:8px"></span>
          <div id="mapaAntResultado" style="display:none;margin-top:8px;font-size:13px">
            Distancia ida y vuelta: <strong id="mapaAntKm" style="color:#1a5276"></strong> km
            <button type="button" onclick="usarKmAnticipo()"
                    style="background:#27ae60;color:white;border:none;border-radius:4px;
                           padding:4px 10px;font-size:12px;cursor:pointer;margin-left:8px">
              ✅ Usar estos km
            </button>
          </div>
        </div>
        <div>
          <label>Asunto / No. de expediente</label>
          <input id="asunto" type="text" placeholder="ej. Baja del C. ...">
        </div>
        <div>
          <label>Fecha de salida</label>
          <input id="salida" type="date" onchange="limpiarResumen()">
        </div>
        <div>
          <label>Fecha de regreso</label>
          <input id="regreso" type="date" onchange="limpiarResumen(); if(document.querySelector('input[name=modo]:checked').value==='avion') actualizarTraslados()">
        </div>
      </div>

      <div style="margin-bottom:14px">
        <label>Modo de transporte</label>
        <div class="radio-group">
          <label><input type="radio" name="modo" value="automovil" checked
                        onchange="toggleTransporte()"> 🚗 Automóvil</label>
          <label><input type="radio" name="modo" value="avion"
                        onchange="toggleTransporte()"> ✈️ Avión</label>
          <label><input type="radio" name="modo" value="autobus"
                        onchange="toggleTransporte()"> 🚌 Autobús</label>
          <label><input type="radio" name="modo" value="otro"
                        onchange="toggleTransporte()"> 🚕 Otro</label>
        </div>
      </div>

      <div id="costoTransDiv" class="hidden" style="margin-bottom:14px">
        <label>Costo estimado del boleto ($)</label>
        <input id="costoTransporte" type="number" step="0.01" placeholder="0.00"
               style="max-width:200px" oninput="actualizarTraslados()">
      </div>

      <!-- Casetas — visible solo para automóvil -->
      <div id="casetasDiv" style="margin-bottom:14px">
        <label>🛣 Casetas (ida y vuelta)
          <span style="font-weight:normal;font-size:11px;color:#777" id="casetasFuente"></span>
        </label>
        <input id="casetasMonto" type="number" step="0.01" min="0" placeholder="0.00"
               style="max-width:200px" oninput="limpiarResumen()"
               title="Costo total de casetas ida y vuelta">
      </div>

      <!-- Traslados locales — solo visible para avión -->
      <div id="traslados-section" class="hidden" style="margin-bottom:14px">
        <div style="background:#eaf4fb;border:1px solid #aed6f1;border-radius:8px;padding:14px">
          <div style="font-weight:bold;color:#1a5276;margin-bottom:10px;font-size:13px">
            🚕 Traslados locales estimados
            <span style="font-weight:normal;color:#777;font-size:11px">
              — Edite los montos según la ciudad
            </span>
          </div>
          <div id="traslados-lista"></div>
        </div>
      </div>

      <div style="margin-bottom:6px">
        <label style="display:flex;align-items:center;gap:8px;font-weight:normal;cursor:pointer">
          <input id="incluyeHotel" type="checkbox" style="width:auto"
                 onchange="limpiarResumen(); if(document.querySelector('input[name=modo]:checked').value==='avion') actualizarTraslados()">
          <span style="font-weight:bold">¿El viaje requiere hospedaje?</span>
          <span style="font-size:11px;color:#777">(se agrega $2,100 por noche)</span>
        </label>
      </div>

      <div class="btn-row">
        <button class="btn btn-primary" onclick="calcular()">
          🧮 Calcular anticipo
        </button>
      </div>
    </div>
  </div>

  <!-- RESUMEN -->
  <div class="card" id="cardResumen" style="display:none">
    <div class="card-header">Desglose del anticipo</div>
    <div class="card-body">
      <div class="resumen">
        <div id="conceptosDiv"></div>
        <div class="total-row">
          <span class="total-label">TOTAL A ANTICIPAR</span>
          <span class="total-monto" id="totalMonto">$0.00</span>
        </div>
        <div class="total-letras" id="totalLetras"></div>
      </div>
      <div id="alertaNoTabulador" class="alert alert-warn hidden">
        ⚠️ La ciudad no está en el tabulador. Se estimaron las comidas según distancia estándar.
        Puede ajustar el número de comidas manualmente.
      </div>

      <!-- Comidas por día — tabla editable -->
      <div style="margin-top:14px;background:#fef9e7;border:1px solid #f0c040;
                  border-radius:6px;padding:14px">
        <div style="font-weight:bold;color:#7d6608;margin-bottom:10px">
          🍽 Comidas por día
          <span style="font-weight:normal;font-size:11px;color:#999">
            — 1 = menos de 6h &nbsp;|&nbsp; 2 = viaje largo &nbsp;|&nbsp; 3 = todo el día
          </span>
        </div>
        <div id="comidas-tabla"></div>
        <button class="btn" style="background:#f0c040;color:#333;margin-top:10px;padding:7px 14px"
                onclick="recalcularConComidas()">🔄 Recalcular totales</button>
      </div>

      <div class="btn-row" style="margin-top:16px">
        <button class="btn btn-success" onclick="generarExcel()">
          📄 Generar Excel (V&V-02)
        </button>
        <button class="btn" style="background:#e67e22;color:white" onclick="enviarAnticipo()">
          📨 Enviar a Administración (Brenda)
        </button>
      </div>
      <div id="msgResult"></div>
    </div>
  </div>

</div><!-- /container -->

<script>
let ultimoCalculo = null;

// Clientes ya vienen del servidor — solo utilidades de fetch para viajes y cálculos
function fetchJson(url, opts){
  return fetch(url, opts).then(r=>{
    if(r.status===302||r.redirected){ window.location='/'; return null; }
    if(!r.ok) throw new Error('Error '+r.status);
    return r.json();
  });
}

// Montos sugeridos por tipo de traslado
const MONTOS_DEFAULT = {
  'casa_aeropuerto':    350,
  'aeropuerto_hotel':   300,
  'aeropuerto_reunion': 300,
  'hotel_reunion':      200,
  'reunion_hotel':      200,
  'reunion_aeropuerto': 300,
  'hotel_aeropuerto':   300,
  'aeropuerto_casa':    350,
};

function toggleTransporte(){
  const modo = document.querySelector('input[name=modo]:checked').value;
  document.getElementById('costoTransDiv').className       = (modo==='automovil') ? 'hidden' : '';
  document.getElementById('casetasDiv').style.display      = (modo==='automovil') ? '' : 'none';
  document.getElementById('traslados-section').className   = (modo==='avion')     ? '' : 'hidden';
  if(modo==='avion') actualizarTraslados();
  limpiarResumen();
}

function actualizarTraslados(){
  const hotel = document.getElementById('incluyeHotel').checked;
  const salida  = document.getElementById('salida').value;
  const regreso = document.getElementById('regreso').value;
  let dias = 1;
  if(salida && regreso){
    const fi = new Date(salida), ff = new Date(regreso);
    dias = Math.max(Math.floor((ff-fi)/(1000*60*60*24))+1, 1);
  }
  const noches = Math.max(dias-1, 0);

  // Construir lista de traslados según si hay hotel y días
  let traslados = [];

  if(!hotel || noches === 0){
    // Viaje de ida y vuelta en el día
    traslados = [
      {id:'casa_aeropuerto',    desc:'Casa → Aeropuerto (ciudad origen)'},
      {id:'aeropuerto_reunion', desc:'Aeropuerto (destino) → Lugar de reunión'},
      {id:'reunion_aeropuerto', desc:'Lugar de reunión → Aeropuerto (destino)'},
      {id:'aeropuerto_casa',    desc:'Aeropuerto (ciudad origen) → Casa'},
    ];
  } else {
    // Viaje con hotel
    traslados = [
      {id:'casa_aeropuerto',    desc:'Casa → Aeropuerto (ciudad origen)'},
      {id:'aeropuerto_hotel',   desc:'Aeropuerto (destino) → Hotel'},
    ];
    // Días intermedios: hotel ↔ reunión
    for(let d=1; d<=noches; d++){
      traslados.push({id:`hotel_reunion_d${d}`,   desc:`Hotel → Lugar de reunión (día ${d})`});
      traslados.push({id:`reunion_hotel_d${d}`,   desc:`Lugar de reunión → Hotel (día ${d})`});
    }
    traslados.push({id:'hotel_aeropuerto',  desc:'Hotel → Aeropuerto (destino)'});
    traslados.push({id:'aeropuerto_casa',   desc:'Aeropuerto (ciudad origen) → Casa'});
  }

  const lista = document.getElementById('traslados-lista');
  lista.innerHTML = traslados.map(t=>{
    const baseId = t.id.replace(/_d\d+$/, '');
    const defVal = MONTOS_DEFAULT[baseId] || 250;
    return `<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">
      <span style="flex:1;font-size:12px;color:#333">${t.desc}</span>
      <div style="display:flex;align-items:center;gap:4px">
        <span style="font-size:12px;color:#555">$</span>
        <input type="number" step="0.01" value="${defVal}"
               id="traslado_${t.id}"
               style="width:90px;padding:5px 8px;border:1px solid #ccc;
                      border-radius:4px;font-size:13px;text-align:right">
      </div>
    </div>`;
  }).join('');
}

function obtenerTraslados(){
  const lista = document.getElementById('traslados-lista');
  const inputs = lista.querySelectorAll('input[type=number]');
  const traslados = [];
  lista.querySelectorAll('div[style*="margin-bottom"]').forEach(row => {
    const desc  = row.querySelector('span').textContent;
    const input = row.querySelector('input');
    const monto = parseFloat(input.value)||0;
    if(monto > 0) traslados.push({descripcion: desc, monto: monto});
  });
  return traslados;
}

function limpiarResumen(){
  document.getElementById('cardResumen').style.display='none';
}

function cargarTabulador(){ limpiarResumen(); buscarCasetas(); }

let _casetasTimer = null;
function buscarCasetas(){
  const modo = document.querySelector('input[name=modo]:checked').value;
  if(modo !== 'automovil') return;
  const ciudad = document.getElementById('ciudad').value.trim();
  if(!ciudad || ciudad.length < 3) return;
  clearTimeout(_casetasTimer);
  _casetasTimer = setTimeout(()=>{
    fetch('/api/casetas?ciudad='+encodeURIComponent(ciudad))
      .then(r=>r.json()).then(d=>{
        const inp = document.getElementById('casetasMonto');
        const lbl = document.getElementById('casetasFuente');
        if(d.monto > 0){
          inp.value = d.monto.toFixed(2);
          lbl.textContent = '— estimado CAPUFE, ajuste si es necesario';
        } else {
          inp.value = '';
          lbl.textContent = '— ciudad no encontrada, ingrese manualmente';
        }
      }).catch(()=>{});
  }, 400);
}

function calcular(){
  const cliente = document.getElementById('selCliente').value;
  const ciudad  = document.getElementById('ciudad').value.trim();
  const salida  = document.getElementById('salida').value;
  const regreso = document.getElementById('regreso').value;
  const modo    = document.querySelector('input[name=modo]:checked').value;
  const costoT  = parseFloat(document.getElementById('costoTransporte').value)||0;
  const hotel   = document.getElementById('incluyeHotel').checked;
  // Casetas: leer campo si ya tiene valor ingresado; null = dejar que el servidor busque
  const casetasRaw = document.getElementById('casetasMonto').value;
  const casetasMonto = (casetasRaw !== '') ? parseFloat(casetasRaw)||0 : null;

  if(!ciudad||!salida||!regreso){
    alert('Complete ciudad, fecha de salida y fecha de regreso.'); return;
  }
  if(!cliente){
    alert('Seleccione el cliente.'); return;
  }

  // Convertir date input (YYYY-MM-DD) a DD/MM/YYYY
  function toMX(d){ const p=d.split('-'); return p[2]+'/'+p[1]+'/'+p[0]; }

  document.getElementById('loading').style.display='flex';
  const traslados = (modo==='avion') ? obtenerTraslados() : [];

  // Si ya existe tabla de comidas (recálculo), usar esos valores
  const mealsExistentes = getMealsPorDiaAnticipo();

  fetch('/api/calcular_anticipo',{
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({
      cliente, ciudad,
      salida:  toMX(salida),
      regreso: toMX(regreso),
      modo_transporte:  modo,
      costo_transporte: costoT,
      incluye_hotel:    hotel,
      traslados:        traslados,
      meals_por_dia:    mealsExistentes.length>0 ? mealsExistentes : [],
      casetas_monto:    casetasMonto,
    })
  }).then(r=>r.json()).then(data=>{
    document.getElementById('loading').style.display='none';
    if(data.error){ alert('Error: '+data.error); return; }
    ultimoCalculo = data;
    // Pre-llenar casetas con el valor sugerido si el campo estaba vacío
    const casetasInput = document.getElementById('casetasMonto');
    if(casetasInput.value === '' && data.casetas_sugeridas > 0){
      casetasInput.value = data.casetas_sugeridas.toFixed(2);
    }
    // Generar tabla de comidas por día
    generarTablaComidasAnticipo(data.dias||1, data.meals_por_dia||[], data.meals_dia||1);
    mostrarResumen(data);
  }).catch(e=>{ document.getElementById('loading').style.display='none'; alert('Error: '+e); });
}

function mostrarResumen(data){
  const div = document.getElementById('conceptosDiv');
  div.innerHTML = (data.conceptos||[]).map(c=>
    `<div class="concepto-row">
       <span class="concepto-label">${c.concepto}</span>
       <span class="concepto-monto">$${fmt(c.monto)}</span>
     </div>`
  ).join('');

  document.getElementById('totalMonto').textContent = '$'+fmt(data.total);
  document.getElementById('totalLetras').textContent = data.letras||'';

  const alertaDiv = document.getElementById('alertaNoTabulador');
  alertaDiv.className = 'alert alert-warn' + (data.en_tabulador ? ' hidden' : '');

  document.getElementById('cardResumen').style.display='block';

  // Pedir letras al servidor
  fetch('/api/numero_letras?n='+data.total).then(r=>r.json()).then(d=>{
    document.getElementById('totalLetras').textContent = d.letras;
  });
}

// ── Tabla de comidas por día (anticipo) ──────────────────────────────────────
function generarTablaComidasAnticipo(dias, mealsArr, nDefault){
  const tabla = document.getElementById('comidas-tabla');
  if(!tabla) return;
  let html = '<div style="display:flex;flex-wrap:wrap;gap:10px">';
  for(let d=1; d<=dias; d++){
    const val = (mealsArr && mealsArr[d-1]) ? mealsArr[d-1] : nDefault;
    html += `<div style="display:flex;align-items:center;gap:6px;
                         background:white;border:1px solid #f0c040;
                         border-radius:6px;padding:6px 10px">
      <span style="font-size:12px;color:#7d6608;white-space:nowrap">Día ${d}</span>
      <select id="ant_comidas_dia_${d}" style="width:60px;padding:4px;font-size:13px;
                                               font-weight:bold;text-align:center;
                                               border:1px solid #ccc;border-radius:4px">
        <option value="1" ${val==1?'selected':''}>1</option>
        <option value="2" ${val==2?'selected':''}>2</option>
        <option value="3" ${val==3?'selected':''}>3</option>
      </select>
      <span style="font-size:11px;color:#aaa">comida(s)</span>
    </div>`;
  }
  html += '</div>';
  tabla.innerHTML = html;
}

function getMealsPorDiaAnticipo(){
  const meals = [];
  let d = 1;
  while(document.getElementById('ant_comidas_dia_'+d)){
    meals.push(parseInt(document.getElementById('ant_comidas_dia_'+d).value)||1);
    d++;
  }
  return meals;
}

function recalcularConComidas(){
  if(!ultimoCalculo) return;
  const meals = getMealsPorDiaAnticipo();
  if(!meals.length) return;
  // Recalcular manteniendo todos los otros conceptos y reemplazando comidas
  const sinComidas = ultimoCalculo.conceptos.filter(c=>!c.concepto.startsWith('Comidas'));

  // Agrupar días por nº de comidas
  const grupos = {};
  meals.forEach((n,i)=>{ grupos[n] = grupos[n]||[]; grupos[n].push(i+1); });
  const nuevasComidas = Object.entries(grupos).sort().map(([n,dias])=>{
    const nd = dias.length, nm = parseInt(n);
    const monto = nd * nm * 290;
    const label = Object.keys(grupos).length===1
      ? `Comidas (${nd} día(s) × ${nm} comida(s)/día × $290)`
      : `Comidas día(s) ${dias.join(',')} × ${nm} comida(s) × $290`;
    return {concepto: label, monto};
  });

  const nuevosConceptos = [...sinComidas.filter(c=>c.concepto.indexOf('Comidas')<0), ...nuevasComidas,
    ...sinComidas.filter(c=>c.concepto.indexOf('Comidas')>=0)];
  // Reordenar: vuelo/gasolina → traslados → comidas → hotel
  const ordenados = [
    ...sinComidas,
    ...nuevasComidas,
  ];
  const nuevoTotal = ordenados.reduce((a,c)=>a+c.monto, 0);
  ultimoCalculo = {...ultimoCalculo, conceptos:ordenados, total:Math.round(nuevoTotal*100)/100,
                   meals_por_dia: meals};
  mostrarResumen(ultimoCalculo);
}

function generarExcel(){
  if(!ultimoCalculo) return;
  function toMX(d){ if(!d) return ''; const p=d.split('-'); return p[2]+'/'+p[1]+'/'+p[0]; }
  const payload = {
    cliente:      document.getElementById('selCliente').value,
    ciudad:       document.getElementById('ciudad').value,
    beneficiario: document.getElementById('beneficiario').value,
    cargo:        document.getElementById('selCliente').value.replace('Viáticos ',''),
    asunto:       document.getElementById('asunto').value,
    lugar:        document.getElementById('ciudad').value,
    salida:       toMX(document.getElementById('salida').value),
    regreso:      toMX(document.getElementById('regreso').value),
    descripcion:  document.getElementById('asunto').value,
    total:        ultimoCalculo.total,
    conceptos:    ultimoCalculo.conceptos,
  };
  fetch('/api/generar_anticipo',{
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(payload)
  }).then(r=>r.json()).then(data=>{
    const msg = document.getElementById('msgResult');
    if(data.error){
      msg.innerHTML='<div class="alert alert-err">Error: '+data.error+'</div>'; return;
    }
    msg.innerHTML=`<div class="alert alert-ok">
      ✅ Solicitud generada: <strong>${data.filename}</strong><br>
      <a href="/api/descargar_anticipo?path=${encodeURIComponent(data.path)}"
         style="color:#0f5132;font-weight:bold" download>⬇ Descargar aquí</a>
    </div>`;
  });
}

function enviarAnticipo(){
  if(!ultimoCalculo){ alert('Calcule el anticipo primero.'); return; }
  function toMX(d){ if(!d) return ''; const p=d.split('-'); return p[2]+'/'+p[1]+'/'+p[0]; }
  const payload = {
    tipo_solicitud:  'anticipo',
    cliente:         document.getElementById('selCliente').value,
    ciudad:          document.getElementById('ciudad').value,
    beneficiario:    document.getElementById('beneficiario').value,
    cargo:           document.getElementById('selCliente').value.replace('Viáticos ',''),
    asunto:          document.getElementById('asunto').value,
    lugar:           document.getElementById('ciudad').value,
    salida:          toMX(document.getElementById('salida').value),
    regreso:         toMX(document.getElementById('regreso').value),
    descripcion:     document.getElementById('asunto').value,
    total_empleado:  ultimoCalculo.total,
    total_cliente:   0,
    conceptos:       ultimoCalculo.conceptos,
    gastos:          [],
  };
  fetch('/api/enviar', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(payload)
  }).then(r=>r.json()).then(data=>{
    const msg = document.getElementById('msgResult');
    if(data.error){
      msg.innerHTML='<div class="alert alert-err">Error: '+data.error+'</div>'; return;
    }
    msg.innerHTML=`<div class="alert alert-ok">
      ✅ Solicitud de anticipo enviada a <strong>Brenda Morales</strong> para aprobación.<br>
      Monto solicitado: <strong>$${fmt(ultimoCalculo.total)}</strong> &nbsp;|&nbsp; ID: <strong>${data.id}</strong>
    </div>`;
  });
}

function fmt(n){
  return Number(n||0).toLocaleString('es-MX',{minimumFractionDigits:2,maximumFractionDigits:2});
}

// ── Mapa anticipo ──────────────────────────────────────────────────────────
var _antKmMapa = 0;
function toggleMapaAnticipo(){
  var d = document.getElementById('mapaAnticipo');
  d.style.display = d.style.display==='none' ? 'block' : 'none';
  var dest = document.getElementById('ciudad').value.trim();
  if(dest) document.getElementById('mapaAntDestino').value = dest;
}
async function geocodeMXant(q){
  var r = await fetch('https://nominatim.openstreetmap.org/search?format=json&countrycodes=mx&limit=1&q='+encodeURIComponent(q),{headers:{'Accept-Language':'es'}});
  var d = await r.json(); return d&&d.length?{lat:parseFloat(d[0].lat),lon:parseFloat(d[0].lon)}:null;
}
async function calcularRutaAnticipo(){
  var o=document.getElementById('mapaAntOrigen').value.trim();
  var d=document.getElementById('mapaAntDestino').value.trim();
  var st=document.getElementById('mapaAntStatus');
  if(!o||!d){st.textContent='⚠️ Ingresa origen y destino';return;}
  st.textContent='🔍 Calculando…';
  var oc=await geocodeMXant(o); var dc=await geocodeMXant(d);
  if(!oc){st.textContent='❌ No encontré: '+o;return;}
  if(!dc){st.textContent='❌ No encontré: '+d;return;}
  var url='https://router.project-osrm.org/route/v1/driving/'+oc.lon+','+oc.lat+';'+dc.lon+','+dc.lat+'?overview=false';
  var resp=await fetch(url); var rd=await resp.json();
  if(!rd.routes||!rd.routes.length){st.textContent='❌ Sin ruta';return;}
  var km_ida=Math.round(rd.routes[0].distance/1000);
  _antKmMapa=km_ida*2;
  document.getElementById('mapaAntKm').textContent=_antKmMapa;
  document.getElementById('mapaAntResultado').style.display='block';
  st.textContent='';
}
function usarKmAnticipo(){
  // Actualiza el campo ciudad con el destino y guarda km_mapa en variable global
  var dest=document.getElementById('mapaAntDestino').value.trim();
  if(dest) document.getElementById('ciudad').value=dest.split(',')[0].trim();
  window._kmMapaAnticipo=_antKmMapa;
  document.getElementById('mapaAnticipo').style.display='none';
  limpiarResumen();
  alert('✅ Se usarán '+_antKmMapa+' km para el cálculo del anticipo.');
}
// Inyectar km_mapa en la solicitud de cálculo de anticipo
var _origCalcAnticipo = typeof calcularAnticipo==='function' ? calcularAnticipo : null;
document.addEventListener('DOMContentLoaded',function(){
  // Patch del fetch al calcular anticipo para incluir km_mapa
  var _origFetch = window.fetch;
  window.fetch = function(url, opts){
    if(typeof url==='string' && url.includes('/api/calcular_anticipo') && window._kmMapaAnticipo){
      try{
        var b = JSON.parse(opts.body||'{}');
        b.km_mapa = window._kmMapaAnticipo;
        opts = Object.assign({},opts,{body:JSON.stringify(b)});
      }catch(e){}
    }
    return _origFetch.call(this,url,opts);
  };
});
</script>
</body></html>"""

@app.route("/api/numero_letras")
@login_required
def api_numero_letras():
    n = request.args.get("n", 0)
    try: return jsonify({"letras": numero_a_letras(float(n))})
    except: return jsonify({"letras": ""})

# ─── SUBIDA DE COMPROBANTES ──────────────────────────────────────────────────

@app.route("/api/subir_comprobantes", methods=["POST"])
@login_required
def api_subir_comprobantes():
    """Recibe archivos del browser y los guarda en la carpeta del usuario."""
    cliente   = nfc(request.form.get("cliente",   ""))
    viaje_key = nfc(request.form.get("viaje_key", ""))  # clave única del viaje

    if not cliente or not viaje_key:
        return jsonify({"error": "Seleccione cliente y viaje antes de subir archivos"})

    usuario  = session["usuario"]
    trip_dir = get_upload_dir(usuario, cliente, viaje_key)

    archivos = request.files.getlist("archivos")
    if not archivos:
        return jsonify({"error": "No se seleccionaron archivos"})

    guardados = []
    for f in archivos:
        if f.filename:
            fname = nfc(os.path.basename(f.filename.replace("\\", "/")))
            dest  = os.path.join(trip_dir, fname)
            f.save(dest)
            guardados.append(fname)

    return jsonify({"ok": True, "guardados": guardados, "n": len(guardados)})


@app.route("/subir_comprobantes", methods=["POST"])
@login_required
def subir_comprobantes_form():
    """Upload via formulario HTML normal (sin JS). Redirige de vuelta al viaje."""
    cliente   = nfc(request.form.get("cliente",   ""))
    viaje_raw = nfc(request.form.get("viaje_raw", ""))
    viaje_key = nfc(request.form.get("viaje_key", ""))

    if cliente and viaje_key:
        usuario  = session["usuario"]
        trip_dir = get_upload_dir(usuario, cliente, viaje_key)
        for f in request.files.getlist("archivos"):
            if f.filename:
                fname = nfc(os.path.basename(f.filename.replace("\\", "/")))
                f.save(os.path.join(trip_dir, fname))

    return redirect(url_for("viaje_page",
                             cliente=cliente,
                             viaje=viaje_raw))


def _comidas_params_from_form():
    """Extrae comidas_N del formulario actual para preservarlos en la redirección."""
    return {k: v for k, v in request.form.items() if k.startswith("comidas_")}


@app.route("/borrar_gasto", methods=["POST"])
@login_required
def borrar_gasto():
    """Borra un archivo de comprobante y su override (si existe)."""
    cliente   = nfc(request.form.get("cliente", ""))
    viaje_raw = nfc(request.form.get("viaje_raw", ""))
    viaje_key = nfc(request.form.get("viaje_key", ""))
    archivo   = request.form.get("archivo", "")
    if cliente and viaje_key and archivo:
        trip_dir = get_upload_dir(session["usuario"], cliente, viaje_key)
        fpath = os.path.join(trip_dir, archivo)
        if os.path.isfile(fpath):
            os.remove(fpath)
        # Limpiar override si existe
        overrides_path = os.path.join(trip_dir, "_overrides.json")
        if os.path.exists(overrides_path):
            try:
                with open(overrides_path, encoding="utf-8") as f:
                    ovr = json.load(f)
                if archivo in ovr:
                    del ovr[archivo]
                    with open(overrides_path, "w", encoding="utf-8") as f:
                        json.dump(ovr, f, ensure_ascii=False)
            except Exception: pass
    params = {"cliente": cliente, "viaje": viaje_raw}
    params.update(_comidas_params_from_form())
    return redirect(url_for("viaje_page", **params))


@app.route("/guardar_edicion", methods=["POST"])
@login_required
def guardar_edicion():
    """Guarda los valores editados de un gasto en _overrides.json."""
    cliente   = nfc(request.form.get("cliente", ""))
    viaje_raw = nfc(request.form.get("viaje_raw", ""))
    viaje_key = nfc(request.form.get("viaje_key", ""))
    archivo   = request.form.get("archivo", "")
    if cliente and viaje_key and archivo:
        trip_dir = get_upload_dir(session["usuario"], cliente, viaje_key)
        overrides_path = os.path.join(trip_dir, "_overrides.json")
        ovr = {}
        if os.path.exists(overrides_path):
            try:
                with open(overrides_path, encoding="utf-8") as f:
                    ovr = json.load(f)
            except Exception: pass
        def _flt(key, default=0.0):
            try: return float(request.form.get(key, default) or default)
            except: return default
        ovr[archivo] = {
            "descripcion": request.form.get("descripcion", ""),
            "tipo":        request.form.get("tipo", "Otro"),
            "rfc":         request.form.get("rfc", ""),
            "con_iva":     _flt("con_iva"),
            "sin_iva":     _flt("sin_iva"),
            "empleado":    _flt("empleado"),
            "cliente":     _flt("empleado"),  # cliente = mismo monto por defecto
        }
        with open(overrides_path, "w", encoding="utf-8") as f:
            json.dump(ovr, f, ensure_ascii=False, indent=2)
    params = {"cliente": cliente, "viaje": viaje_raw}
    params.update(_comidas_params_from_form())
    return redirect(url_for("viaje_page", **params))


@app.route("/exportar_viaje_excel")
@login_required
def exportar_viaje_excel():
    """Admin/socia: exporta desglose completo de un viaje a Excel (.xlsx)."""
    if session.get("rol") not in ("admin", "socia"):
        return "Sin permiso", 403
    usuario   = request.args.get("usuario", "")
    cliente   = nfc(request.args.get("cliente", ""))
    viaje_key = nfc(request.args.get("viaje_key", ""))
    if not usuario or not cliente or not viaje_key:
        return "Parámetros incompletos.", 400

    users = load_users()
    nombre_emp = users.get(usuario, {}).get("nombre", usuario)

    gastos, total_emp, total_cli, aprobacion = _gastos_para_viaje(usuario, cliente, viaje_key)

    # Leer config del viaje
    trip_dir = None
    user_dir = os.path.join(UPLOADS_DIR, safe_name(usuario))
    if os.path.isdir(user_dir):
        for cf in os.listdir(user_dir):
            cand = os.path.join(user_dir, cf, safe_name(viaje_key))
            if os.path.isdir(cand):
                trip_dir = cand; break
    cfg = {}
    if trip_dir:
        cfg_path = os.path.join(trip_dir, "_config_viaje.json")
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path, encoding="utf-8") as f:
                    cfg = json.load(f)
            except Exception: pass

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Desglose de viaje"

    AZUL  = "1A5276"
    VERDE = "1E8449"
    GRIS  = "EBF5FB"
    NEG   = "000000"
    BLANCO = "FFFFFF"
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, txt, bg=AZUL, fg=BLANCO, bold=True, sz=10, aln="center"):
        cell.value = txt
        cell.font  = Font(bold=bold, color=fg, size=sz, name="Calibri")
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=True)
        cell.border = brd

    def val(cell, v, fmt=None, bold=False, aln="left", bg=None):
        cell.value = v
        cell.font  = Font(color=NEG, size=10, bold=bold, name="Calibri")
        cell.alignment = Alignment(horizontal=aln, vertical="center", wrap_text=True)
        cell.border = brd
        if fmt: cell.number_format = fmt
        if bg:  cell.fill = PatternFill("solid", fgColor=bg)

    # ── Encabezado ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    hdr(ws["A1"], f"DESGLOSE DE VIAJE — {viaje_key.replace('_',' ').upper()}", sz=13)
    ws.row_dimensions[1].height = 22

    info_rows = [
        ("Colaborador:", nombre_emp),
        ("Cliente:",     cliente),
        ("Destino:",     cfg.get("lugar", viaje_key.replace("_"," "))),
        ("Salida:",      cfg.get("salida","")),
        ("Regreso:",     cfg.get("regreso","")),
        ("Asunto:",      cfg.get("motivo") or cfg.get("descripcion","")),
    ]
    for i, (lbl, v_) in enumerate(info_rows, 2):
        ws.merge_cells(f"A{i}:B{i}")
        val(ws[f"A{i}"], lbl, bold=True, bg="FFEBF5FB" if i%2==0 else None)
        ws.merge_cells(f"C{i}:J{i}")
        val(ws[f"C{i}"], v_)

    # ── Cabecera tabla ───────────────────────────────────────────────────────────
    r = len(info_rows) + 3
    ws.row_dimensions[r].height = 32
    cols = ["#","Tipo","Archivo / Descripción","Comprobante $",
            "Con IVA $","Sin IVA $","IVA $",
            "A pagar\nEmpleado $","A cobrar\nCliente $","Estado"]
    for c, lbl in enumerate(cols, 1):
        hdr(ws.cell(r, c), lbl)

    # ── Filas de gastos ──────────────────────────────────────────────────────────
    fmt_p = '#,##0.00'
    sum_comp = sum_con = sum_sin = sum_emp2 = sum_cli2 = 0.0
    for idx, g in enumerate(gastos, 1):
        dr = r + idx
        con  = float(g.get("con_iva",  0) or 0)
        sin_ = float(g.get("sin_iva",  0) or 0)
        emp  = float(g.get("empleado", 0) or 0)
        cli  = float(g.get("cliente",  0) or 0)
        comp = con + sin_
        iva  = round(con * 0.16, 2)
        est  = g.get("status_apro","pendiente")
        est_txt = {"aprobado":"✔ Aprobado","parcial":"⚡ Parcial",
                   "rechazado":"✗ Rechazado","pendiente":"⏳ Pendiente"}.get(est, est)
        bg_row = "F9EBEA" if est=="rechazado" else ("EAFAF1" if est=="aprobado" else None)

        for c, v_ in [(1, idx), (2, g.get("tipo","")),
                      (3, g.get("descripcion", g.get("archivo",""))),
                      (4, comp or None), (5, con or None), (6, sin_ or None),
                      (7, iva or None), (8, emp), (9, cli), (10, est_txt)]:
            cell = ws.cell(dr, c)
            fmt = fmt_p if c in (4,5,6,7,8,9) and v_ is not None else None
            val(cell, v_, fmt=fmt, aln="right" if c in (4,5,6,7,8,9) else "left",
                bg=("FFFAF0F0" if est=="rechazado" else ("FFECF9F0" if est=="aprobado" else None)))

        sum_comp += comp; sum_con += con; sum_sin += sin_
        sum_emp2 += emp;  sum_cli2 += cli

    # ── Fila de totales ──────────────────────────────────────────────────────────
    tr = r + len(gastos) + 1
    ws.merge_cells(f"A{tr}:C{tr}")
    hdr(ws[f"A{tr}"], "TOTALES", bg=AZUL, sz=10)
    for c, v_ in [(4, sum_comp), (5, sum_con), (6, sum_sin),
                  (7, round(sum_con*0.16,2)), (8, sum_emp2), (9, sum_cli2)]:
        cell = ws.cell(tr, c)
        cell.value = round(v_, 2)
        cell.font  = Font(bold=True, color=BLANCO, size=10, name="Calibri")
        cell.fill  = PatternFill("solid", fgColor=AZUL)
        cell.number_format = fmt_p
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = brd

    # ── Anchos de columna ────────────────────────────────────────────────────────
    for c, w in [(1,4),(2,12),(3,38),(4,13),(5,13),(6,13),(7,10),(8,14),(9,14),(10,13)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    import tempfile
    out_name = f"Desglose_{viaje_key}_{safe_name(nombre_emp)}.xlsx"
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, out_name)
        wb.save(out_path)
        return send_file(out_path, as_attachment=True, download_name=out_name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/descargar_anticipo_admin")
@login_required
def descargar_anticipo_admin():
    """Admin descarga el Excel de anticipo de cualquier colaborador."""
    if session.get("rol") not in ("admin", "socia"):
        return "Sin permiso", 403
    usuario    = request.args.get("usuario", "")
    viaje_key  = nfc(request.args.get("viaje_key", ""))
    cliente_s  = nfc(request.args.get("cliente", ""))
    if not usuario or not viaje_key:
        return "Parámetros incompletos.", 400

    users_a    = load_users()
    nombre_usr = users_a.get(usuario, {}).get("nombre", usuario)

    # Buscar submission de anticipo del colaborador para este viaje
    ciudad_key = viaje_key.rsplit("_", 1)[0].replace("_", " ")
    sub = None
    for s in list_submissions(empleado_filter=nombre_usr):
        if (s.get("tipo_solicitud") == "anticipo"
                and strip_accents(s.get("ciudad","")).upper() == strip_accents(ciudad_key).upper()):
            sub = s
            break

    import tempfile
    if sub:
        out_name = f"Anticipo {ciudad_key} - {nombre_usr}.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, out_name)
            write_anticipo_excel(out_path, sub)
            return send_file(out_path, as_attachment=True, download_name=out_name,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        # Sin submission: intentar reconstruir desde _config_viaje.json
        trip_dir_a = os.path.join(UPLOADS_DIR, safe_name(usuario), safe_name(cliente_s), safe_name(viaje_key))
        cfg_a = {}
        cfg_pa = os.path.join(trip_dir_a, "_config_viaje.json")
        if os.path.exists(cfg_pa):
            try:
                with open(cfg_pa, encoding="utf-8") as f:
                    cfg_a = json.load(f)
            except Exception:
                pass
        if not cfg_a:
            return "No se encontró información de anticipo para este viaje.", 404
        data_a = {
            "beneficiario": nombre_usr,
            "cargo":        users_a.get(usuario, {}).get("cargo", ""),
            "cliente":      cliente_s,
            "ciudad":       ciudad_key,
            "salida":       cfg_a.get("salida", ""),
            "regreso":      cfg_a.get("regreso", ""),
            "asunto":       cfg_a.get("motivo") or cfg_a.get("descripcion", ""),
            "descripcion":  cfg_a.get("descripcion") or cfg_a.get("motivo", ""),
            "conceptos":    [],
            "total":        0,
        }
        out_name = f"Anticipo {ciudad_key} - {nombre_usr}.xlsx"
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = os.path.join(tmpdir, out_name)
            write_anticipo_excel(out_path, data_a)
            return send_file(out_path, as_attachment=True, download_name=out_name,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/descargar_comprobacion")
@login_required
def descargar_comprobacion():
    """Genera y descarga el Excel de comprobación de gastos para el viaje actual."""
    cliente_sel      = nfc(request.args.get("cliente", ""))
    viaje_raw        = nfc(request.args.get("viaje", ""))
    viaje_key_direct = nfc(request.args.get("viaje_key_direct", ""))
    if not cliente_sel or (not viaje_raw and not viaje_key_direct) or viaje_raw == "__nuevo__":
        return "Seleccione un viaje primero.", 400
    if viaje_key_direct:
        viaje_key  = viaje_key_direct
        parts      = viaje_key.rsplit("_", 1)
        ciudad_sel = parts[0].replace("_", " ")
        fecha_sel  = parts[1] if len(parts) > 1 else ""
    else:
        parts      = viaje_raw.split("|", 1)
        ciudad_sel = parts[0]
        fecha_sel  = parts[1] if len(parts) > 1 else ""
        viaje_key  = safe_name(ciudad_sel + ("_" + fecha_sel.replace("/","") if fecha_sel else ""))

    # Admin puede descargar el formato de cualquier colaborador
    if session.get("rol") in ("admin", "socia") and request.args.get("usuario"):
        usuario = request.args.get("usuario")
        users_tmp = load_users()
        nombre_usuario = users_tmp.get(usuario, {}).get("nombre", usuario)
    else:
        usuario        = session["usuario"]
        nombre_usuario = session["nombre"]

    # ── Datos del calendario (misma lógica que viaje_page) ─────────────────────
    cal_info = {}
    try:
        cal_file = None
        client_dir_c = find_dir(BASE_DIR, cliente_sel) if cliente_sel else None
        search_dirs  = ([client_dir_c] if client_dir_c else []) + [BASE_DIR]
        for sd in search_dirs:
            for f in os.listdir(sd):
                if "calendario" in nfc(f).lower() and f.lower().endswith(".xlsx"):
                    cal_file = os.path.join(sd, f); break
            if cal_file: break
        if not cal_file and os.path.exists(CALENDAR):
            cal_file = CALENDAR
        records = load_calendar(cal_file) if cal_file else []
        user_plain = _strip_acc(nombre_usuario)
        propios = [r for r in records if _strip_acc(r.get("abogado","")) == user_plain]
        if propios: records = propios
        hint = [fecha_sel] if fecha_sel else None
        info = lookup_calendar(ciudad_sel, records, hint_dates=hint)
        if info:
            cal_info = info
    except Exception as e:
        print(f"[descargar_comprobacion cal] {e}")

    # ── Cargar gastos usando la misma fuente que el sistema (consistencia garantizada)
    gastos_raw, _, _, _ = _gastos_para_viaje(usuario, cliente_sel, viaje_key)
    gastos = gastos_raw

    # ── Leer datos guardados del calendario desde _config_viaje.json ────────────
    trip_dir_dl = os.path.join(UPLOADS_DIR, safe_name(usuario),
                               safe_name(cliente_sel), safe_name(viaje_key))
    cfg_dl = {}
    cfg_path_dl = os.path.join(trip_dir_dl, "_config_viaje.json")
    if os.path.exists(cfg_path_dl):
        try:
            with open(cfg_path_dl, encoding="utf-8") as _f:
                cfg_dl = json.load(_f)
        except Exception:
            pass
    # Parámetros enviados directamente desde el botón de descarga (más confiables)
    url_salida  = nfc(request.args.get("salida", ""))
    url_regreso = nfc(request.args.get("regreso", ""))
    url_asunto  = nfc(request.args.get("asunto", ""))
    url_desc    = nfc(request.args.get("descripcion", ""))

    # Prioridad: URL params → _config_viaje.json → calendario en vivo → fallback
    salida_dl  = url_salida  or cfg_dl.get("salida",  cal_info.get("salida",  fecha_sel))
    regreso_dl = url_regreso or cfg_dl.get("regreso", cal_info.get("regreso", ""))
    # asunto: puede estar en cfg como "motivo" o "descripcion"
    asunto_dl  = (url_asunto
                  or cfg_dl.get("motivo")
                  or cfg_dl.get("descripcion")
                  or cal_info.get("motivo", "")
                  or cal_info.get("descripcion", ""))
    lugar_dl   = cfg_dl.get("lugar") or ciudad_sel
    desc_dl    = (url_desc
                  or cfg_dl.get("descripcion")
                  or cfg_dl.get("motivo")
                  or cal_info.get("descripcion")
                  or cal_info.get("motivo", "")
                  or asunto_dl)

    # ── Armar datos para write_excel ────────────────────────────────────────────
    cargo_usuario = ""
    try:
        users = load_users()
        cargo_usuario = users.get(usuario, {}).get("cargo", "")
    except Exception: pass

    data = {
        "salida":       salida_dl,
        "regreso":      regreso_dl,
        "beneficiario": nombre_usuario,
        "cargo":        cargo_usuario,
        "asunto":       asunto_dl,
        "lugar":        lugar_dl,
        "descripcion":  desc_dl,
        "gastos": [{
            "tipo":        g["tipo"],
            "comprobante": g["archivo"],
            "rfc":         g.get("rfc",""),
            "descripcion": g.get("descripcion", os.path.splitext(g["archivo"])[0]),
            "con_iva":     g.get("con_iva",0),
            "sin_iva":     g.get("sin_iva",0),
            "empleado":    g.get("empleado",0),
        } for g in gastos],
    }

    # ── Generar el Excel en carpeta temporal y enviarlo ─────────────────────────
    import tempfile
    fecha_hoy = datetime.now().strftime("%d%m%Y")
    out_name  = f"Comprobación {ciudad_sel} {fecha_sel.replace('/','')}.xlsx"
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, out_name)
        write_excel(out_path, data)
        import shutil as _sh
        _sh.copy2(out_path, "/tmp/ultima_comprobacion.xlsx")
        return send_file(out_path, as_attachment=True, download_name=out_name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── RESUMEN POR USUARIO ──────────────────────────────────────────────────────

def _load_aprobacion(trip_dir):
    """Lee _aprobacion.json del viaje; devuelve {} si no existe."""
    path = os.path.join(trip_dir, "_aprobacion.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_aprobacion(trip_dir, data):
    path = os.path.join(trip_dir, "_aprobacion.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _gastos_para_viaje(usuario, cliente_safe, viaje_key):
    """Devuelve (lista_gastos, total_empleado, total_cliente, aprobacion_dict).
    Prioridad: _gastos_calculados.json (guardado por viaje_page) → recálculo manual."""
    trip_dir = os.path.join(UPLOADS_DIR, safe_name(usuario),
                            safe_name(cliente_safe), safe_name(viaje_key))
    if not os.path.isdir(trip_dir):
        # Buscar en todas las carpetas del usuario si el viaje_key existe en alguna
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usuario))
        found = None
        if os.path.isdir(user_dir):
            for cli_folder in os.listdir(user_dir):
                candidate = os.path.join(user_dir, cli_folder, safe_name(viaje_key))
                if os.path.isdir(candidate):
                    found = candidate
                    break
        if not found:
            return [], 0.0, 0.0, {}
        trip_dir = found

    aprobacion = _load_aprobacion(trip_dir)
    det        = aprobacion.get("detalle", {})

    # ── Leer número de comidas guardado por viaje_page ───────────────────────
    total_comidas_cfg = None
    config_path = os.path.join(trip_dir, "_config_viaje.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, encoding="utf-8") as f:
                cfg = json.load(f)
            total_comidas_cfg = cfg.get("total_comidas")
        except Exception:
            pass

    # ── Intentar leer gastos ya calculados por viaje_page ─────────────────────
    calc_path = os.path.join(trip_dir, "_gastos_calculados.json")
    if os.path.exists(calc_path):
        try:
            with open(calc_path, encoding="utf-8") as f:
                gastos_base = json.load(f)
            gastos = []
            for row in gastos_base:
                fname     = row.get("archivo", "")
                info_apro = det.get(fname, {})
                # Recalcular comidas si hay un número guardado diferente al cacheado
                if row.get("tipo") == "Comida" and total_comidas_cfg is not None:
                    amount = row.get("con_iva", 0.0) + row.get("sin_iva", 0.0)
                    emp, cli = calc_row("Comida", amount, None, total_comidas_cfg)
                    row["empleado"] = round(emp, 2)
                    row["cliente"]  = round(cli, 2)
                # Aplicar override de monto cliente si el admin lo ajustó
                if "cliente_override" in info_apro:
                    row["cliente"] = info_apro["cliente_override"]
                # Solo hay monto aprobado cuando el admin lo revisó explícitamente
                item_status = info_apro.get("status", "pendiente")
                row["status_apro"]     = item_status
                row["comentario_apro"] = info_apro.get("comentario", "")
                if item_status in ("aprobado", "parcial"):
                    row["aprobado"] = float(info_apro.get("aprobado", row.get("empleado", 0)))
                else:
                    row["aprobado"] = 0.0
                gastos.append(row)
            total_emp = sum(g.get("empleado", 0) for g in gastos)
            total_cli = sum(g.get("cliente",  0) for g in gastos)
            return gastos, round(total_emp, 2), round(total_cli, 2), aprobacion
        except Exception:
            pass  # caer al recálculo manual si el archivo está corrupto

    # ── Recálculo manual (sin viaje_page previo) ──────────────────────────────
    overrides = {}
    ovr_path  = os.path.join(trip_dir, "_overrides.json")
    if os.path.exists(ovr_path):
        try:
            with open(ovr_path, encoding="utf-8") as f:
                overrides = json.load(f)
        except Exception:
            pass

    # ── Cargar tabulador del grupo correcto para cobro al cliente ─────────────
    # viaje_key tiene forma "Ciudad_DDMMAAAA"; extraer ciudad para lookup
    ciudad_fallback = viaje_key.replace("_", " ").split()[0] if viaje_key else ""
    # cliente_safe puede ser "Liogero" o "INSECOM"; reconstruir nombre de carpeta
    cliente_folder = next(
        (nfc(d) for d in os.listdir(BASE_DIR)
         if os.path.isdir(os.path.join(BASE_DIR, d))
         and nfc(d).startswith("Viáticos ")
         and safe_name(nfc(d)) == safe_name(nfc(cliente_safe))),
        None)
    tabulador_fb = _get_tabulador(cliente_folder) if cliente_folder else {}
    city_data_fb = lookup_city(ciudad_fallback, tabulador_fb) if ciudad_fallback else None

    # ── Si no hay _config_viaje.json y la ciudad no está en tabulador,
    #    intentar recuperar el nº de comidas desde _overrides.json ────────────
    if total_comidas_cfg is None and city_data_fb is None:
        # Buscar un override de tipo "Comida" con valor empleado > 0
        for _fn, _ov in overrides.items():
            if _ov.get("tipo") == "Comida" and _ov.get("empleado", 0) > 0:
                backcalc = round(_ov["empleado"] / EMPLOYEE_MEAL_RATE)
                total_comidas_cfg = max(1, min(3, backcalc))
                break

    archivos = sorted(f for f in os.listdir(trip_dir)
                      if f.lower().endswith((".pdf",".jpg",".jpeg",".png")))
    gastos   = []
    for fname in archivos:
        if fname in overrides:
            row = dict(overrides[fname])
            row.setdefault("archivo", fname)
            row.setdefault("tipo",    detect_category(fname))
            row.setdefault("cliente", row.get("empleado", 0))
        else:
            fpath     = os.path.join(trip_dir, fname)
            cat       = detect_category(fname)
            amount    = 0.0
            rfc       = ""
            if fname.lower().endswith(".pdf"):
                try:
                    d_pdf  = extract_pdf_data(fpath)
                    amount = d_pdf["amount"]
                    rfc    = d_pdf["rfc"]
                except Exception:
                    pass
            causa_iva = cat not in {"Hotel","Autobús","Vuelo","Otro"}
            meals_fb  = total_comidas_cfg if (cat == "Comida" and total_comidas_cfg) else 1
            emp, cli  = calc_row(cat, amount, city_data_fb, meals_fb)
            row = {
                "archivo":  fname,
                "tipo":     cat,
                "rfc":      rfc,
                "con_iva":  round(amount, 2) if causa_iva else 0.0,
                "sin_iva":  0.0 if causa_iva else round(amount, 2),
                "empleado": round(emp, 2),
                "cliente":  round(cli, 2),
            }
        info_apro          = det.get(fname, {})
        # Aplicar override de monto cliente si el admin lo ajustó
        if "cliente_override" in info_apro:
            row["cliente"] = info_apro["cliente_override"]
        item_status        = info_apro.get("status", "pendiente")
        row["status_apro"] = item_status
        row["comentario_apro"] = info_apro.get("comentario", "")
        # Solo hay monto aprobado cuando el admin lo revisó explícitamente
        if item_status in ("aprobado", "parcial"):
            row["aprobado"] = float(info_apro.get("aprobado", row.get("empleado", 0)))
        else:
            row["aprobado"] = 0.0
        gastos.append(row)

    total_emp = sum(g.get("empleado", 0) for g in gastos)
    total_cli = sum(g.get("cliente",  0) for g in gastos)
    return gastos, round(total_emp, 2), round(total_cli, 2), aprobacion


def _resumen_usuario(usuario, nombre):
    """Construye el resumen completo de un usuario: anticipos + comprobaciones."""
    user_dir = os.path.join(UPLOADS_DIR, safe_name(usuario))
    viajes   = []

    if os.path.isdir(user_dir):
        for cliente_safe in sorted(os.listdir(user_dir)):
            cliente_dir = os.path.join(user_dir, cliente_safe)
            if not os.path.isdir(cliente_dir):
                continue
            for viaje_key in sorted(os.listdir(cliente_dir)):
                trip_dir = os.path.join(cliente_dir, viaje_key)
                if not os.path.isdir(trip_dir):
                    continue
                archivos = [f for f in os.listdir(trip_dir)
                            if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))]
                if not archivos:
                    continue
                gastos, total_comp, total_cli, aprobacion = \
                    _gastos_para_viaje(usuario, cliente_safe, viaje_key)
                viaje_status = aprobacion.get("status", "pendiente")
                # Solo hay monto aprobado cuando el admin revisó el viaje
                if viaje_status == "pendiente":
                    total_aprobado = 0.0
                else:
                    total_aprobado = sum(
                        g.get("aprobado", 0) for g in gastos
                        if g.get("status_apro") in ("aprobado", "parcial")
                    )
                viajes.append({
                    "cliente":        cliente_safe,
                    "viaje_key":      viaje_key,
                    "num_archivos":   len(archivos),
                    "gastos":         gastos,
                    "comprobado":     total_comp,
                    "total_cliente":  round(total_cli, 2),
                    "aprobado":       round(total_aprobado, 2),
                    "aprobacion":     aprobacion,
                    "status_apro":    viaje_status,
                    "comentario_gral":aprobacion.get("comentario_general", ""),
                })

    anticipos = [s for s in list_submissions(empleado_filter=nombre)
                 if s.get("tipo_solicitud") == "anticipo"
                 and s.get("status") == "aprobado"]
    anticipo_total = sum(float(a.get("monto", 0) or 0) for a in anticipos)

    total_comprobado = sum(v["comprobado"]    for v in viajes)
    total_aprobado   = sum(v["aprobado"]      for v in viajes)
    total_cliente    = sum(v["total_cliente"] for v in viajes)
    saldo            = round(anticipo_total - total_aprobado, 2)

    return {
        "usuario":          usuario,
        "nombre":           nombre,
        "viajes":           viajes,
        "anticipo_total":   anticipo_total,
        "anticipos":        anticipos,
        "total_comprobado": total_comprobado,
        "total_aprobado":   round(total_aprobado, 2),
        "total_cliente":    round(total_cliente, 2),
        "saldo":            saldo,
    }


RESUMEN_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Resumen de Viáticos — V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#1a5276;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#aed6f1;text-decoration:none;font-size:13px}
.topbar a:hover{color:white}
.container{max-width:1160px;margin:24px auto;padding:0 16px}
h1{font-size:22px;color:#1a5276;margin-bottom:4px}
.subtitle{color:#777;font-size:13px;margin-bottom:18px}

/* filtro */
.filtro-bar{display:flex;gap:8px;align-items:center;margin-bottom:18px;flex-wrap:wrap}
.filtro-bar select,.filtro-bar button{padding:7px 12px;border-radius:6px;
  border:1px solid #ccc;font-size:13px;cursor:pointer}
.filtro-bar button{background:#1a5276;color:white;border:none}

/* tarjeta de empleado */
.user-card{background:white;border-radius:10px;
           box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:24px;overflow:hidden}
.user-header{background:#1a5276;color:white;padding:11px 18px;
             display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.user-header h2{font-size:15px;font-weight:600}
.chips{display:flex;gap:8px;flex-wrap:wrap;align-items:center}
.chip{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700}
.chip-a{background:#f0b429;color:#333}
.chip-c{background:#27ae60;color:white}
.chip-ap{background:#2980b9;color:white}
.chip-cli{background:#8e44ad;color:white}
.chip-debe{background:#e74c3c;color:white}
.chip-recibe{background:#d5f5e3;color:#155724}
.chip-ok{background:#d5f5e3;color:#155724}
.chip-pend{background:#fef9e7;color:#7d6608;border:1px solid #f0c040}
.chip-parcial{background:#fde8cc;color:#7d4e00}

/* tabla de gastos */
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{background:#ebf5fb;color:#1a5276;padding:7px 10px;text-align:left;
         border-bottom:2px solid #d6eaf8;font-weight:600;white-space:nowrap}
th.num,td.num{text-align:right}
tbody tr:nth-child(even){background:#fafcff}
tbody tr:hover{background:#ebf5fb}
td{padding:6px 10px;border-bottom:1px solid #eee;vertical-align:middle}
.tag{display:inline-block;padding:2px 7px;border-radius:4px;font-size:11px;font-weight:600}
.tag-Hotel{background:#d5e8d4;color:#1e5736}
.tag-Comida{background:#fff3cd;color:#856404}
.tag-Transporte{background:#dce3f8;color:#1a3a7c}
.tag-Gasolina{background:#ffe8cc;color:#7d4e00}
.tag-Casetas{background:#e8e8e8;color:#555}
.tag-Vuelo{background:#d0eaff;color:#0d5c99}
.tag-Autobús{background:#e6d5f8;color:#4a1d7a}
.tag-Otro{background:#f3f3f3;color:#777}
.apro-ok{color:#27ae60;font-weight:700}
.apro-rechazado{color:#e74c3c;font-weight:700;text-decoration:line-through}
.apro-parcial{color:#e67e22;font-weight:700}
.apro-pend{color:#7f8c8d}
.comentario-cell{font-size:11px;color:#c0392b;font-style:italic;max-width:200px}

/* separador de viaje */
.viaje-sep td{background:#eaf0f8;font-weight:700;color:#1a5276;
              font-size:12px;padding:7px 10px;border-top:2px solid #c8d8ea}
.viaje-actions{float:right;display:flex;gap:6px}

/* subtotal */
.subtotal-row td{font-weight:700;background:#f0f8ff;border-top:2px solid #aed6f1;font-size:12px}

/* barra de resumen inferior */
.resumen-bar{display:flex;gap:16px;flex-wrap:wrap;padding:14px 18px;
             background:#f8fbff;border-top:1px solid #d6eaf8}
.ri{text-align:center;min-width:90px}
.ri .val{font-size:18px;font-weight:700;color:#1a5276}
.ri .lbl{font-size:10px;color:#999;margin-top:2px;text-transform:uppercase;letter-spacing:.4px}

/* sección de clientes */
.sec-title{font-size:17px;font-weight:700;color:#1a5276;margin:28px 0 12px;
           border-bottom:2px solid #d6eaf8;padding-bottom:6px}
.client-card{background:white;border-radius:8px;padding:16px 20px;margin-bottom:12px;
             box-shadow:0 1px 5px rgba(0,0,0,.07);display:flex;
             justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px}
.client-name{font-size:15px;font-weight:700;color:#1a5276}
.client-detail{font-size:12px;color:#777;margin-top:3px}
.client-amount{text-align:right}
.client-amount .val{font-size:22px;font-weight:700;color:#8e44ad}
.client-amount .lbl{font-size:11px;color:#aaa}

.empty{padding:20px;color:#aaa;text-align:center;font-style:italic;font-size:13px}
.no-data{text-align:center;padding:60px;color:#bbb}
.badge-status{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700}
.bs-pend{background:#fef9e7;color:#7d6608;border:1px solid #f0c040}
.bs-apro{background:#d5f5e3;color:#155724}
.bs-parc{background:#fde8cc;color:#7d4e00}
.bs-rech{background:#fde8e8;color:#c0392b}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700;font-size:16px">📊 Resumen de Viáticos — Villar &amp; Villar</span>
  <div style="display:flex;gap:16px;align-items:center">
    {% if rol == 'admin' %}<a href="/dashboard">⚙ Panel Admin</a><a href="/config_email">✉ Config correo</a>{% endif %}
    {% if rol == 'socia' %}<a href="/autorizar_viaje">🔏 Autorizaciones</a>{% endif %}
    <a href="/viaje">📋 Capturar viaje</a>
    <a href="/logout">Salir</a>
  </div>
</div>

<div class="container">
  <h1>Resumen de Viáticos por Empleado</h1>
  <p class="subtitle">Comprobantes subidos · Monto comprobado vs aprobado · Saldo pendiente{% if rol=='admin' %} · A cobrar al cliente{% endif %}</p>

  {% if rol == 'admin' %}
  <div class="filtro-bar">
    <form method="GET" action="/resumen_viaticos" style="display:flex;gap:8px;flex-wrap:wrap">
      <select name="usuario" onchange="this.form.submit()">
        <option value="">— Todos los empleados —</option>
        {% for uid, unombre in usuarios %}
        <option value="{{ uid }}" {% if filtro_usuario == uid %}selected{% endif %}>{{ unombre }}</option>
        {% endfor %}
      </select>
    </form>
    <a href="/resumen_clientes" style="background:#8e44ad;color:white;padding:7px 14px;
       border-radius:6px;text-decoration:none;font-size:13px;font-weight:600">
      🏢 Resumen por cliente
    </a>
  </div>
  {% endif %}

  {% if not resumenes %}
  <div class="no-data">
    <div style="font-size:48px;margin-bottom:12px">📂</div>
    <div>No hay comprobantes registrados aún.</div>
  </div>
  {% endif %}

  {% for r in resumenes %}
  <div class="user-card">
    <div class="user-header">
      <h2>👤 {{ r.nombre }} <span style="opacity:.7;font-weight:400;font-size:12px">({{ r.usuario }})</span></h2>
      <div class="chips">
        {% if r.anticipo_total > 0 %}
        <span class="chip chip-a">💰 Anticipo ${{ "%.2f"|format(r.anticipo_total) }}</span>
        {% endif %}
        <span class="chip chip-c">✅ Comprobado ${{ "%.2f"|format(r.total_comprobado) }}</span>
        {% if r.total_aprobado > 0 and r.total_aprobado != r.total_comprobado %}
        <span class="chip chip-ap">✔ Aprobado ${{ "%.2f"|format(r.total_aprobado) }}</span>
        {% endif %}
        {% if rol == 'admin' and r.total_cliente > 0 %}
        <span class="chip chip-cli">🏢 A cliente ${{ "%.2f"|format(r.total_cliente) }}</span>
        {% endif %}
        {% if r.total_aprobado > 0 %}
          {% if r.anticipo_total > 0 %}
            {% if r.saldo > 0.01 %}<span class="chip chip-debe">⚠ Devolver ${{ "%.2f"|format(r.saldo) }}</span>
            {% elif r.saldo < -0.01 %}<span class="chip chip-recibe">📤 Reembolsar ${{ "%.2f"|format(-r.saldo) }}</span>
            {% else %}<span class="chip chip-ok">✔ Liquidado</span>{% endif %}
          {% else %}
          <span class="chip chip-recibe">📤 A reembolsar ${{ "%.2f"|format(r.total_aprobado) }}</span>
          {% endif %}
        {% endif %}
      </div>
    </div>

    {% if r.viajes %}
    <div style="overflow-x:auto">
    <table style="min-width:750px">
      <thead>
        <tr>
          <th style="width:26%">Archivo / Concepto</th>
          <th>Tipo</th>
          <th class="num">Monto comprobante</th>
          <th class="num">Solicitado</th>
          <th class="num">Aprobado</th>
          {% if rol == 'admin' %}<th class="num">A cliente</th>{% endif %}
          <th>Observación</th>
        </tr>
      </thead>
      <tbody>
      {% for v in r.viajes %}
        {% set apro_status = v.status_apro %}
        <tr class="viaje-sep">
          <td colspan="{% if rol=='admin' %}7{% else %}6{% endif %}">
            📁 <strong>{{ v.cliente }}</strong> — {{ v.viaje_key.replace('_',' ') }}
            <span style="font-weight:400;color:#6a8daa;font-size:11px">({{ v.num_archivos }} comprobante{{ 's' if v.num_archivos!=1 else '' }})</span>
            &nbsp;
            {% if apro_status == 'aprobado' %}<span class="badge-status bs-apro">✔ Aprobado</span>
            {% elif apro_status == 'aprobado_parcial' %}<span class="badge-status bs-parc">⚡ Aprobado parcial</span>
            {% elif apro_status == 'rechazado' %}<span class="badge-status bs-rech">✗ Rechazado</span>
            {% else %}<span class="badge-status bs-pend">⏳ Pendiente de revisión</span>{% endif %}
            {% set st_socia = v.aprobacion.get('status_socia','') %}
            {% if st_socia == 'pendiente' %}<span class="badge-status" style="background:#f5eef8;color:#6c3483;border:1px solid #c39bd3">⏳ Pendiente autorización socia</span>
            {% elif st_socia == 'autorizado' %}<span class="badge-status" style="background:#d5f5e3;color:#155724">✅ Autorizado por socia</span>
            {% elif st_socia == 'rechazado' %}<span class="badge-status" style="background:#fde8e8;color:#c0392b">❌ Rechazado por socia</span>
            {% endif %}
            {% if rol == 'admin' %}
            <span class="viaje-actions">
              <a href="/aprobar_viaje?usuario={{ r.usuario }}&cliente={{ v.cliente|urlencode }}&viaje_key={{ v.viaje_key|urlencode }}"
                 style="background:#e67e22;color:white;padding:3px 10px;border-radius:4px;
                        text-decoration:none;font-size:11px;font-weight:600">
                ✏ Revisar / Aprobar
              </a>
              <a href="/descargar_anticipo_admin?usuario={{ r.usuario }}&cliente={{ v.cliente|urlencode }}&viaje_key={{ v.viaje_key|urlencode }}"
                 style="background:#117a65;color:white;padding:3px 10px;border-radius:4px;
                        text-decoration:none;font-size:11px;font-weight:600">
                📋 Anticipo
              </a>
              <a href="/descargar_comprobacion?usuario={{ r.usuario }}&cliente={{ v.cliente|urlencode }}&viaje_key_direct={{ v.viaje_key|urlencode }}"
                 style="background:#1a5276;color:white;padding:3px 10px;border-radius:4px;
                        text-decoration:none;font-size:11px;font-weight:600">
                📥 Comprobación
              </a>
              <a href="/exportar_viaje_excel?usuario={{ r.usuario }}&cliente={{ v.cliente|urlencode }}&viaje_key={{ v.viaje_key|urlencode }}"
                 style="background:#1e8449;color:white;padding:3px 10px;border-radius:4px;
                        text-decoration:none;font-size:11px;font-weight:600">
                📊 Exportar Excel
              </a>
              <a href="/comprobantes_viaje?usuario={{ r.usuario }}&cliente={{ v.cliente|urlencode }}&viaje_key={{ v.viaje_key|urlencode }}"
                 style="background:#6c3483;color:white;padding:3px 10px;border-radius:4px;
                        text-decoration:none;font-size:11px;font-weight:600">
                📎 Comprobantes
              </a>
            </span>
            {% endif %}
            {% if v.comentario_gral %}
            <div style="font-size:11px;color:#c0392b;margin-top:3px">💬 {{ v.comentario_gral }}</div>
            {% endif %}
          </td>
        </tr>
        {% for g in v.gastos %}
        {% set st = g.status_apro %}
        <tr {% if st == 'rechazado' %}style="opacity:.6;background:#fff8f8"{% endif %}>
          <td style="font-size:12px" title="{{ g.archivo }}">
            {{ g.get('descripcion', g.archivo) or g.archivo }}
          </td>
          <td><span class="tag tag-{{ g.tipo }}">{{ g.tipo }}</span></td>
          <td class="num">
            {% set monto = (g.con_iva or 0)+(g.sin_iva or 0) %}
            {% if monto > 0 %}${{ "%.2f"|format(monto) }}{% else %}<span style="color:#ccc">—</span>{% endif %}
          </td>
          <td class="num">
            {% if g.empleado > 0 %}${{ "%.2f"|format(g.empleado) }}{% else %}<span style="color:#ccc">—</span>{% endif %}
          </td>
          <td class="num">
            {% if st == 'rechazado' %}
              <span class="apro-rechazado">$0.00</span>
            {% elif st == 'parcial' %}
              <span class="apro-parcial">${{ "%.2f"|format(g.aprobado) }}</span>
            {% elif st == 'aprobado' %}
              <span class="apro-ok">${{ "%.2f"|format(g.aprobado) }}</span>
            {% else %}
              <span class="apro-pend">${{ "%.2f"|format(g.empleado) }}</span>
            {% endif %}
          </td>
          {% if rol == 'admin' %}
          <td class="num" style="color:#8e44ad;font-weight:600">
            {% if g.cliente > 0 %}${{ "%.2f"|format(g.cliente) }}{% else %}<span style="color:#ccc">—</span>{% endif %}
          </td>
          {% endif %}
          <td class="comentario-cell">
            {% if g.comentario_apro %}💬 {{ g.comentario_apro }}{% endif %}
          </td>
        </tr>
        {% endfor %}
        <tr class="subtotal-row">
          <td colspan="{% if rol=='admin' %}3{% else %}2{% endif %}" style="text-align:right;color:#555;font-size:11px">
            Subtotal {{ v.viaje_key.replace('_',' ') }}
          </td>
          <td class="num">${{ "%.2f"|format(v.comprobado) }}</td>
          <td class="num" style="color:#2980b9">${{ "%.2f"|format(v.aprobado) }}</td>
          {% if rol == 'admin' %}
          <td class="num" style="color:#8e44ad">${{ "%.2f"|format(v.total_cliente) }}</td>
          {% endif %}
          <td></td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>

    <div class="resumen-bar">
      <div class="ri"><div class="val">${{ "%.2f"|format(r.total_comprobado) }}</div><div class="lbl">Comprobado</div></div>
      {% if r.total_aprobado > 0 %}
      <div class="ri"><div class="val" style="color:#2980b9">${{ "%.2f"|format(r.total_aprobado) }}</div><div class="lbl">Aprobado</div></div>
      {% if r.total_comprobado - r.total_aprobado > 0.01 %}
      <div class="ri"><div class="val" style="color:#e74c3c">${{ "%.2f"|format(r.total_comprobado - r.total_aprobado) }}</div><div class="lbl">No aprobado</div></div>
      {% endif %}
      {% if r.anticipo_total > 0 %}
      <div class="ri"><div class="val" style="color:#b7950b">${{ "%.2f"|format(r.anticipo_total) }}</div><div class="lbl">Anticipo</div></div>
      <div class="ri">
        {% if r.saldo > 0.01 %}<div class="val" style="color:#c0392b">${{ "%.2f"|format(r.saldo) }}</div><div class="lbl">A devolver</div>
        {% elif r.saldo < -0.01 %}<div class="val" style="color:#27ae60">${{ "%.2f"|format(-r.saldo) }}</div><div class="lbl">A reembolsar</div>
        {% else %}<div class="val" style="color:#7f8c8d">$0.00</div><div class="lbl">Liquidado ✓</div>{% endif %}
      </div>
      {% else %}
      <div class="ri"><div class="val" style="color:#27ae60">${{ "%.2f"|format(r.total_aprobado) }}</div><div class="lbl">A reembolsar</div></div>
      {% endif %}
      {% endif %}
      {% if rol == 'admin' and r.total_cliente > 0 %}
      <div class="ri"><div class="val" style="color:#8e44ad">${{ "%.2f"|format(r.total_cliente) }}</div><div class="lbl">A cobrar a cliente</div></div>
      {% endif %}
      <div class="ri"><div class="val" style="color:#888">{{ r.viajes|length }}</div><div class="lbl">Viaje{{ 's' if r.viajes|length!=1 else '' }}</div></div>
    </div>
    {% else %}
    <div class="empty">Sin comprobantes subidos aún.</div>
    {% endif %}
  </div>
  {% endfor %}

  {% if rol == 'admin' and resumenes|length > 1 %}
  <div style="background:white;border-radius:10px;padding:20px 24px;
              box-shadow:0 2px 8px rgba(0,0,0,.08);margin-top:4px">
    <div class="sec-title" style="margin-top:0">📊 Totales globales</div>
    <div style="display:flex;gap:32px;flex-wrap:wrap">
      <div class="ri" style="text-align:left">
        <div class="val">${{ "%.2f"|format(resumenes|sum(attribute='total_comprobado')) }}</div>
        <div class="lbl">Total comprobado</div>
      </div>
      <div class="ri" style="text-align:left">
        <div class="val" style="color:#2980b9">${{ "%.2f"|format(resumenes|sum(attribute='total_aprobado')) }}</div>
        <div class="lbl">Total aprobado</div>
      </div>
      <div class="ri" style="text-align:left">
        <div class="val" style="color:#27ae60">${{ "%.2f"|format(resumenes|sum(attribute='total_aprobado')) }}</div>
        <div class="lbl">A reembolsar (neto)</div>
      </div>
      <div class="ri" style="text-align:left">
        <div class="val" style="color:#8e44ad">${{ "%.2f"|format(resumenes|sum(attribute='total_cliente')) }}</div>
        <div class="lbl">A cobrar a clientes</div>
      </div>
    </div>
  </div>
  {% endif %}
</div>
</body></html>
"""



@app.route("/resumen_viaticos")
@login_required
def resumen_viaticos():
    nombre_usuario = session["nombre"]
    rol            = session["rol"]
    usuario_actual = session["usuario"]

    users = load_users()

    if rol == "admin":
        # Admin puede filtrar por usuario o ver todos
        filtro = request.args.get("usuario", "")
        if filtro and filtro in users:
            lista = [(filtro, users[filtro].get("nombre", filtro))]
            filtro_nombre = users[filtro].get("nombre", filtro)
        else:
            lista = [(u, d.get("nombre", u))
                     for u, d in users.items() if d.get("rol") != "admin"]
            filtro = ""
            filtro_nombre = ""
    else:
        lista         = [(usuario_actual, nombre_usuario)]
        filtro        = usuario_actual
        filtro_nombre = nombre_usuario

    resumenes = []
    for usr, nom in lista:
        r = _resumen_usuario(usr, nom)
        if r["viajes"] or r["anticipo_total"] > 0:
            resumenes.append(r)

    return render_template_string(RESUMEN_HTML,
        rol=rol,
        resumenes=resumenes,
        filtro_usuario=filtro,
        filtro_nombre=filtro_nombre,
        usuarios=[(u, d.get("nombre",u)) for u, d in users.items()
                  if d.get("rol") != "admin"])


# ─── APROBACIÓN DE VIAJES ─────────────────────────────────────────────────────

APROBAR_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Aprobar viaje — V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#1a5276;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#aed6f1;text-decoration:none;font-size:13px}
.container{max-width:900px;margin:28px auto;padding:0 16px}
h1{font-size:20px;color:#1a5276;margin-bottom:4px}
.subtitle{color:#888;font-size:13px;margin-bottom:20px}
.card{background:white;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);
      margin-bottom:20px;overflow:hidden}
.card-header{background:#1a5276;color:white;padding:10px 18px;font-weight:600;font-size:14px}
.card-body{padding:18px}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#ebf5fb;color:#1a5276;padding:8px 10px;text-align:left;
         border-bottom:2px solid #d6eaf8;font-weight:600}
th.num{text-align:right}
td{padding:8px 10px;border-bottom:1px solid #eee;vertical-align:middle}
td.num{text-align:right}
tbody tr:hover{background:#f8fbff}
.tag{display:inline-block;padding:2px 7px;border-radius:4px;font-size:11px;font-weight:600}
.tag-Hotel{background:#d5e8d4;color:#1e5736}
.tag-Comida{background:#fff3cd;color:#856404}
.tag-Transporte{background:#dce3f8;color:#1a3a7c}
.tag-Gasolina{background:#ffe8cc;color:#7d4e00}
.tag-Casetas{background:#e8e8e8;color:#555}
.tag-Vuelo{background:#d0eaff;color:#0d5c99}
.tag-Autobús{background:#e6d5f8;color:#4a1d7a}
.tag-Otro{background:#f3f3f3;color:#777}
input[type=number],select,textarea{padding:5px 8px;border:1px solid #ccc;border-radius:4px;
  font-size:12px;font-family:inherit}
input[type=number]{width:90px;text-align:right}
select{width:120px}
textarea{width:100%;min-height:48px;resize:vertical}
.comment-input{width:220px;min-height:38px;resize:vertical}
label{font-weight:600;font-size:12px;color:#555;display:block;margin-bottom:4px}
.form-group{margin-bottom:14px}
.btn{padding:9px 20px;border-radius:6px;border:none;cursor:pointer;font-size:14px;
     font-weight:600;text-decoration:none;display:inline-block}
.btn-success{background:#27ae60;color:white}
.btn-secondary{background:#ccc;color:#333}
.btn-row{display:flex;gap:10px;margin-top:20px}
.rechazado-row{opacity:.5;background:#fff8f8}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700">✏ Aprobar viaje — {{ empleado_nombre }}</span>
  <a href="/resumen_viaticos">← Volver al resumen</a>
</div>
<div class="container">
  <h1>{{ viaje_key.replace('_',' ') }} — {{ cliente }}</h1>
  <p class="subtitle">Empleado: <strong>{{ empleado_nombre }}</strong> &nbsp;|&nbsp; Revise cada concepto e indique el monto aprobado y, si aplica, el motivo de rechazo o ajuste.</p>

  {% if montos_estimados %}
  <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:12px 16px;
              margin-bottom:16px;font-size:13px;color:#856404">
    ⚠️ <strong>Los montos mostrados son estimados</strong> — el colaborador aún no ha abierto la
    página de comprobación de este viaje en el sistema, por lo que los montos de "Empleado $"
    podrían no coincidir exactamente con lo que solicitó. Pide al colaborador que abra su viaje
    en el sistema para sincronizar los montos antes de aprobar.
  </div>
  {% endif %}

  <form method="POST" action="/guardar_aprobacion">
    <input type="hidden" name="usuario"    value="{{ usuario }}">
    <input type="hidden" name="cliente"    value="{{ cliente }}">
    <input type="hidden" name="viaje_key"  value="{{ viaje_key }}">

    <div class="card">
      <div class="card-header" style="display:flex;justify-content:space-between;align-items:center">
        <span>Detalle de gastos</span>
        <label style="color:white;font-size:12px;font-weight:600;display:flex;align-items:center;gap:8px;margin:0">
          Comidas del viaje:
          <select name="total_comidas" style="width:60px;padding:4px 6px;font-size:12px;color:#333;border-radius:4px;border:none">
            <option value="1" {% if total_comidas == 1 %}selected{% endif %}>1</option>
            <option value="2" {% if total_comidas == 2 %}selected{% endif %}>2</option>
            <option value="3" {% if total_comidas == 3 %}selected{% endif %}>3</option>
          </select>
        </label>
      </div>
      <div class="card-body" style="padding:0">
        <table>
          <thead>
            <tr>
              <th>Archivo / Concepto</th>
              <th>Tipo</th>
              <th class="num">Comprobante</th>
              <th class="num">Empleado $</th>
              <th class="num">A cliente $</th>
              <th class="num">Aprobado $</th>
              <th>Estado</th>
              <th>Comentario / Motivo</th>
            </tr>
          </thead>
          <tbody>
          {% for g in gastos %}
          <tr id="row_{{ loop.index0 }}" {% if g.status_apro == 'rechazado' %}class="rechazado-row"{% endif %}>
            <td style="font-size:12px" title="{{ g.archivo }}">
              {{ g.get('descripcion', g.archivo) or g.archivo }}
              &nbsp;<a href="/ver_comprobante/{{ usuario }}/{{ cliente|urlencode }}/{{ viaje_key|urlencode }}/{{ g.archivo }}"
                 target="_blank" style="font-size:11px;color:#2980b9;text-decoration:none;font-weight:600">Ver</a>
            </td>
            <td><span class="tag tag-{{ g.tipo }}">{{ g.tipo }}</span></td>
            <td class="num">
              {% set monto = (g.con_iva or 0)+(g.sin_iva or 0) %}
              {% if monto > 0 %}${{ "%.2f"|format(monto) }}{% else %}—{% endif %}
            </td>
            <td class="num">${{ "%.2f"|format(g.empleado) }}</td>
            <td class="num">
              <input type="number" step="0.01" min="0"
                     name="cliente_{{ loop.index0 }}"
                     value="{{ "%.2f"|format(g.cliente) }}"
                     style="width:90px;text-align:right">
            </td>
            <td class="num">
              {% if g.status_apro == 'pendiente' %}
                {% set apro_val = g.empleado %}
              {% elif g.status_apro == 'rechazado' %}
                {% set apro_val = 0 %}
              {% else %}
                {% set apro_val = g.aprobado %}
              {% endif %}
              <input type="number" step="0.01" min="0"
                     name="aprobado_{{ loop.index0 }}"
                     value="{{ "%.2f"|format(apro_val) }}"
                     {% if g.status_apro in ('aprobado', 'rechazado') %}readonly style="background:#f0f0f0;width:90px;text-align:right"{% endif %}>
              <input type="hidden" name="archivo_{{ loop.index0 }}" value="{{ g.archivo }}">
            </td>
            <td>
              <select name="status_{{ loop.index0 }}">
                <option value="pendiente" {% if g.status_apro=='pendiente' %}selected{% endif %}>⏳ Pendiente</option>
                <option value="aprobado"  {% if g.status_apro=='aprobado'  %}selected{% endif %}>✔ Aprobado</option>
                <option value="parcial"   {% if g.status_apro=='parcial'   %}selected{% endif %}>⚡ Parcial</option>
                <option value="rechazado" {% if g.status_apro=='rechazado' %}selected{% endif %}>✗ Rechazado</option>
              </select>
            </td>
            <td>
              <textarea class="comment-input" name="comentario_{{ loop.index0 }}"
                        placeholder="Motivo de ajuste o rechazo…">{{ g.comentario_apro }}</textarea>
            </td>
          </tr>
          {% endfor %}
          <input type="hidden" name="num_gastos" value="{{ gastos|length }}">
          </tbody>
        </table>
      </div>
    </div>

    <div class="card">
      <div class="card-header">Estado general y comentario</div>
      <div class="card-body">
        <div style="display:flex;gap:24px;flex-wrap:wrap">
          <div class="form-group">
            <label>Estado del viaje</label>
            <select name="status_general" style="width:200px;padding:7px">
              <option value="pendiente"       {% if aprobacion.status=='pendiente'       %}selected{% endif %}>⏳ Pendiente de revisión</option>
              <option value="aprobado"        {% if aprobacion.status=='aprobado'        %}selected{% endif %}>✔ Aprobado completo</option>
              <option value="aprobado_parcial"{% if aprobacion.status=='aprobado_parcial'%}selected{% endif %}>⚡ Aprobado parcialmente</option>
              <option value="rechazado"       {% if aprobacion.status=='rechazado'       %}selected{% endif %}>✗ Rechazado</option>
            </select>
          </div>
          <div class="form-group" style="flex:1;min-width:280px">
            <label>Comentario general (visible para el empleado)</label>
            <textarea name="comentario_general" style="width:100%;min-height:60px"
                      placeholder="Observaciones generales sobre este viaje…">{{ aprobacion.comentario_general or '' }}</textarea>
          </div>
        </div>
        <div class="btn-row">
          <button type="submit" class="btn btn-success">💾 Guardar aprobación</button>
          <a href="/resumen_viaticos" class="btn btn-secondary">✕ Cancelar</a>
        </div>
      </div>
    </div>
  </form>
</div>
</body></html>"""


@app.route("/aprobar_viaje")
@login_required
def aprobar_viaje():
    if session.get("rol") != "admin":
        return redirect(url_for("resumen_viaticos"))
    usuario    = request.args.get("usuario", "")
    cliente    = nfc(request.args.get("cliente", ""))
    viaje_key  = nfc(request.args.get("viaje_key", ""))
    if not usuario or not cliente or not viaje_key:
        return "Parámetros incompletos.", 400
    users = load_users()
    empleado_nombre = users.get(usuario, {}).get("nombre", usuario)
    gastos, _, _, aprobacion = _gastos_para_viaje(usuario, cliente, viaje_key)
    # Detectar si los montos son estimados (sin _gastos_calculados.json)
    trip_dir_ap = os.path.join(UPLOADS_DIR, safe_name(usuario),
                               safe_name(cliente), safe_name(viaje_key))
    montos_estimados = not os.path.exists(os.path.join(trip_dir_ap, "_gastos_calculados.json"))
    total_comidas_ap = 1
    cfg_path_ap = os.path.join(trip_dir_ap, "_config_viaje.json")
    if os.path.exists(cfg_path_ap):
        try:
            with open(cfg_path_ap, encoding="utf-8") as _f:
                _cfg = json.load(_f)
            total_comidas_ap = int(_cfg.get("total_comidas", 1))
        except Exception:
            pass
    else:
        # Intentar inferir desde los gastos calculados
        for _g in gastos:
            if _g.get("tipo") == "Comida" and _g.get("empleado", 0) > 0:
                backcalc = round(_g["empleado"] / EMPLOYEE_MEAL_RATE)
                total_comidas_ap = max(1, min(3, backcalc))
                break
    return render_template_string(APROBAR_HTML,
        usuario=usuario, cliente=cliente, viaje_key=viaje_key,
        empleado_nombre=empleado_nombre, gastos=gastos, aprobacion=aprobacion,
        total_comidas=total_comidas_ap, montos_estimados=montos_estimados)


@app.route("/guardar_aprobacion", methods=["POST"])
@login_required
def guardar_aprobacion():
    if session.get("rol") != "admin":
        return redirect(url_for("resumen_viaticos"))
    usuario   = request.form.get("usuario", "")
    cliente   = nfc(request.form.get("cliente", ""))
    viaje_key = nfc(request.form.get("viaje_key", ""))
    if not usuario or not cliente or not viaje_key:
        return "Parámetros incompletos.", 400

    trip_dir = os.path.join(UPLOADS_DIR, safe_name(usuario),
                            safe_name(cliente), safe_name(viaje_key))
    os.makedirs(trip_dir, exist_ok=True)

    n = int(request.form.get("num_gastos", 0))

    # Obtener montos empleado reales para cuando el status sea 'aprobado'
    gastos_actuales, _, _, _ = _gastos_para_viaje(usuario, cliente, viaje_key)
    empleado_por_archivo = {g["archivo"]: g.get("empleado", 0) for g in gastos_actuales}

    detalle = {}
    for i in range(n):
        fname  = request.form.get(f"archivo_{i}", "")
        if not fname:
            continue
        status = request.form.get(f"status_{i}", "pendiente")
        try:   aprobado = float(request.form.get(f"aprobado_{i}", 0) or 0)
        except: aprobado = 0.0
        # Forzar montos correctos según status
        if status == "aprobado":
            aprobado = empleado_por_archivo.get(fname, aprobado)
        elif status == "rechazado":
            aprobado = 0.0
        # Leer cliente override si se envió
        try:   cliente_val = float(request.form.get(f"cliente_{i}", "") or "")
        except: cliente_val = None
        entry = {
            "aprobado":   round(aprobado, 2),
            "status":     status,
            "comentario": request.form.get(f"comentario_{i}", "").strip(),
        }
        if cliente_val is not None:
            entry["cliente_override"] = round(cliente_val, 2)
        detalle[fname] = entry

    status_general = request.form.get("status_general", "pendiente")
    aprobacion = {
        "status":              status_general,
        "comentario_general":  request.form.get("comentario_general", "").strip(),
        "aprobado_por":        session["nombre"],
        "fecha":               datetime.now().strftime("%d/%m/%Y %H:%M"),
        "detalle":             detalle,
    }

    # Flujo de dos niveles: si el viaje es aprobado y el empleado requiere socia, marcar pendiente
    if status_general in ("aprobado", "aprobado_parcial") and usuario in EMPLOYEES_REQUIRING_SOCIA:
        aprobacion["status_socia"] = "pendiente"
        # Notificar a la socia
        users = load_users()
        empleado_nombre = users.get(usuario, {}).get("nombre", usuario)
        gastos_socia, total_emp_socia, _, _ = _gastos_para_viaje(usuario, cliente, viaje_key)
        total_aprobado_socia = sum(
            float(d.get("aprobado", 0)) for d in detalle.values()
            if d.get("status") in ("aprobado", "parcial")
        )
        # Buscar email de socia
        socia_email = next(
            (d.get("email","") for u, d in users.items() if d.get("rol") == "socia"),
            "lmontero@villarabogados.com.mx"
        )
        body_socia = f"""
        <html><body>
        <h2>Solicitud de autorización de viáticos</h2>
        <p>Se requiere tu autorización para el siguiente viaje:</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse">
          <tr><td><b>Empleado</b></td><td>{empleado_nombre}</td></tr>
          <tr><td><b>Cliente</b></td><td>{cliente.replace('_',' ')}</td></tr>
          <tr><td><b>Viaje</b></td><td>{viaje_key.replace('_',' ')}</td></tr>
          <tr><td><b>Monto aprobado</b></td><td>${total_aprobado_socia:,.2f}</td></tr>
        </table>
        <p><a href="http://127.0.0.1:5050/autorizar_viaje">Ver en el sistema →</a></p>
        </body></html>
        """
        send_email([socia_email], "Solicitud de autorización de viáticos", body_socia)

    _save_aprobacion(trip_dir, aprobacion)

    # Guardar total_comidas en _config_viaje.json si el admin lo especificó
    try:
        tc_form = int(request.form.get("total_comidas", "") or "")
        tc_form = max(1, min(3, tc_form))
        cfg_path_ga = os.path.join(trip_dir, "_config_viaje.json")
        cfg_ga = {}
        if os.path.exists(cfg_path_ga):
            try:
                with open(cfg_path_ga, encoding="utf-8") as _f:
                    cfg_ga = json.load(_f)
            except Exception:
                pass
        cfg_ga["total_comidas"] = tc_form
        with open(cfg_path_ga, "w", encoding="utf-8") as _f:
            json.dump(cfg_ga, _f, ensure_ascii=False, indent=2)
    except (ValueError, TypeError):
        pass

    return redirect(url_for("resumen_viaticos", usuario=usuario))


# ─── AUTORIZACIÓN POR SOCIA ───────────────────────────────────────────────────

AUTORIZAR_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Autorización de Viáticos — Socia</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#6c3483;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#d7bde2;text-decoration:none;font-size:13px}
.container{max-width:1000px;margin:28px auto;padding:0 16px}
h1{font-size:20px;color:#6c3483;margin-bottom:4px}
.subtitle{color:#888;font-size:13px;margin-bottom:20px}
.card{background:white;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);
      margin-bottom:20px;overflow:hidden}
.card-header{background:#6c3483;color:white;padding:10px 18px;font-weight:600;font-size:14px}
.card-body{padding:18px}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#f5eef8;color:#6c3483;padding:8px 10px;text-align:left;
         border-bottom:2px solid #e8daef;font-weight:600}
td{padding:8px 10px;border-bottom:1px solid #eee;vertical-align:middle}
.btn{padding:7px 16px;border-radius:6px;border:none;cursor:pointer;font-size:13px;
     font-weight:600;text-decoration:none;display:inline-block}
.btn-success{background:#27ae60;color:white}
.btn-danger{background:#e74c3c;color:white}
.btn-secondary{background:#ccc;color:#333}
.comment-input{width:200px;min-height:34px;resize:vertical;padding:4px 6px;
               border:1px solid #ccc;border-radius:4px;font-size:12px}
.empty{padding:30px;text-align:center;color:#aaa;font-style:italic}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700">🔏 Autorización de Viáticos — Socia</span>
  <div style="display:flex;gap:16px">
    <a href="/dashboard">← Dashboard</a>
    <a href="/logout">Salir</a>
  </div>
</div>
<div class="container">
  <h1>Viajes pendientes de autorización</h1>
  <p class="subtitle">Los viajes ya aprobados por la administradora que requieren tu autorización final.</p>

  {% if not pendientes %}
  <div class="card"><div class="card-body empty">No hay viajes pendientes de autorización.</div></div>
  {% else %}
  <div class="card">
    <div class="card-header">Viajes pendientes ({{ pendientes|length }})</div>
    <div class="card-body" style="padding:0">
      <table>
        <thead>
          <tr>
            <th>Empleado</th>
            <th>Cliente</th>
            <th>Viaje</th>
            <th style="text-align:right">Total aprobado</th>
            <th>Comentario</th>
            <th style="text-align:center">Acción</th>
          </tr>
        </thead>
        <tbody>
        {% for p in pendientes %}
        <tr>
          <td><strong>{{ p.nombre }}</strong><br><span style="font-size:11px;color:#888">{{ p.usuario }}</span></td>
          <td>{{ p.cliente_safe.replace('_',' ') }}</td>
          <td>{{ p.viaje_key.replace('_',' ') }}</td>
          <td style="text-align:right;font-weight:700;color:#6c3483">${{ "%.2f"|format(p.total_aprobado) }}</td>
          <td>
            <form method="POST" action="/guardar_autorizacion" style="display:inline">
              <input type="hidden" name="usuario"     value="{{ p.usuario }}">
              <input type="hidden" name="cliente_safe" value="{{ p.cliente_safe }}">
              <input type="hidden" name="viaje_key"   value="{{ p.viaje_key }}">
              <textarea class="comment-input" name="comentario_socia" placeholder="Comentario (opcional)…"></textarea>
          </td>
          <td style="text-align:center;white-space:nowrap">
              <button type="submit" name="accion" value="autorizado" class="btn btn-success">✔ Autorizar</button>
              &nbsp;
              <button type="submit" name="accion" value="rechazado" class="btn btn-danger">✗ Rechazar</button>
            </form>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endif %}
</div>
</body></html>"""


@app.route("/autorizar_viaje")
@login_required
def autorizar_viaje():
    if session.get("rol") != "socia":
        return redirect(url_for("dashboard"))
    users = load_users()
    pendientes = []
    for usr, dat in users.items():
        if dat.get("rol") == "admin":
            continue
        if usr not in EMPLOYEES_REQUIRING_SOCIA:
            continue
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
        if not os.path.isdir(user_dir):
            continue
        for cliente_safe in sorted(os.listdir(user_dir)):
            cliente_dir = os.path.join(user_dir, cliente_safe)
            if not os.path.isdir(cliente_dir):
                continue
            for viaje_key in sorted(os.listdir(cliente_dir)):
                trip_dir = os.path.join(cliente_dir, viaje_key)
                if not os.path.isdir(trip_dir):
                    continue
                apro = _load_aprobacion(trip_dir)
                st = apro.get("status", "pendiente")
                st_socia = apro.get("status_socia", "")
                if st in ("aprobado", "aprobado_parcial") and st_socia == "pendiente":
                    # calcular total aprobado
                    detalle = apro.get("detalle", {})
                    total_ap = sum(
                        float(v.get("aprobado", 0)) for v in detalle.values()
                        if v.get("status") in ("aprobado", "parcial")
                    )
                    pendientes.append({
                        "usuario":     usr,
                        "nombre":      dat.get("nombre", usr),
                        "cliente_safe": cliente_safe,
                        "viaje_key":   viaje_key,
                        "total_aprobado": round(total_ap, 2),
                    })
    return render_template_string(AUTORIZAR_HTML, pendientes=pendientes)


@app.route("/guardar_autorizacion", methods=["POST"])
@login_required
def guardar_autorizacion():
    if session.get("rol") != "socia":
        return redirect(url_for("dashboard"))
    usuario      = request.form.get("usuario", "")
    cliente_safe = nfc(request.form.get("cliente_safe", ""))
    viaje_key    = nfc(request.form.get("viaje_key", ""))
    accion       = request.form.get("accion", "autorizado")
    comentario   = request.form.get("comentario_socia", "").strip()
    if not usuario or not cliente_safe or not viaje_key:
        return "Parámetros incompletos.", 400

    trip_dir = os.path.join(UPLOADS_DIR, safe_name(usuario),
                            safe_name(cliente_safe), safe_name(viaje_key))
    apro = _load_aprobacion(trip_dir)
    apro["status_socia"]     = accion
    apro["comentario_socia"] = comentario
    apro["autorizado_por"]   = session["nombre"]
    apro["fecha_socia"]      = datetime.now().strftime("%d/%m/%Y %H:%M")
    _save_aprobacion(trip_dir, apro)

    # Notificar al empleado
    users = load_users()
    emp_email  = users.get(usuario, {}).get("email", "")
    emp_nombre = users.get(usuario, {}).get("nombre", usuario)
    if emp_email:
        accion_txt = "autorizado ✔" if accion == "autorizado" else "rechazado ✗"
        body_emp = f"""
        <html><body>
        <h2>Resultado de autorización de viáticos</h2>
        <p>Tu viaje ha sido <strong>{accion_txt}</strong> por la socia.</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse">
          <tr><td><b>Cliente</b></td><td>{cliente_safe.replace('_',' ')}</td></tr>
          <tr><td><b>Viaje</b></td><td>{viaje_key.replace('_',' ')}</td></tr>
          <tr><td><b>Decisión</b></td><td>{accion_txt}</td></tr>
          {'<tr><td><b>Comentario</b></td><td>' + comentario + '</td></tr>' if comentario else ''}
        </table>
        </body></html>
        """
        send_email([emp_email], f"Resultado autorización viáticos — {viaje_key.replace('_',' ')}", body_emp)

    return redirect(url_for("autorizar_viaje"))


# ─── RESUMEN POR CLIENTE ───────────────────────────────────────────────────────

@app.route("/resumen_clientes")
@login_required
def resumen_clientes():
    if session.get("rol") != "admin":
        return redirect(url_for("resumen_viaticos"))

    users = load_users()
    # Acumular por cliente → viaje_key → {empleados, total_cliente, total_comprobado}
    clientes = {}  # cliente_safe → {viajes: [...], total_cliente, total_comprobado}

    for usr, dat in users.items():
        if dat.get("rol") == "admin":
            continue
        nombre = dat.get("nombre", usr)
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
        if not os.path.isdir(user_dir):
            continue
        for cliente_safe in sorted(os.listdir(user_dir)):
            cliente_dir = os.path.join(user_dir, cliente_safe)
            if not os.path.isdir(cliente_dir):
                continue
            for viaje_key in sorted(os.listdir(cliente_dir)):
                trip_dir = os.path.join(cliente_dir, viaje_key)
                if not os.path.isdir(trip_dir):
                    continue
                archivos = [f for f in os.listdir(trip_dir)
                            if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))]
                if not archivos:
                    continue
                gastos_cli, total_emp, total_cli, aprobacion = \
                    _gastos_para_viaje(usr, cliente_safe, viaje_key)
                viaje_status_cli = aprobacion.get("status", "pendiente")
                if viaje_status_cli == "pendiente":
                    total_aprobado = 0.0
                else:
                    total_aprobado = sum(
                        g.get("aprobado", 0) for g in gastos_cli
                        if g.get("status_apro") in ("aprobado", "parcial")
                    )

                if cliente_safe not in clientes:
                    clientes[cliente_safe] = {"viajes": [], "total_cliente": 0.0,
                                              "total_comprobado": 0.0, "total_aprobado": 0.0}
                clientes[cliente_safe]["viajes"].append({
                    "usuario":       usr,
                    "nombre":        nombre,
                    "viaje_key":     viaje_key,
                    "num_archivos":  len(archivos),
                    "comprobado":    total_emp,
                    "aprobado":      round(total_aprobado, 2),
                    "total_cliente": round(total_cli, 2),
                    "status_apro":   aprobacion.get("status","pendiente"),
                })
                clientes[cliente_safe]["total_cliente"]    += total_cli
                clientes[cliente_safe]["total_comprobado"] += total_emp
                clientes[cliente_safe]["total_aprobado"]   += total_aprobado

    # Calcular totales globales
    gran_total_cli  = sum(c["total_cliente"]    for c in clientes.values())
    gran_total_comp = sum(c["total_comprobado"] for c in clientes.values())
    gran_total_apro = sum(c["total_aprobado"]   for c in clientes.values())

    return render_template_string(CLIENTES_HTML,
        clientes=sorted(clientes.items()),
        gran_total_cli=gran_total_cli,
        gran_total_comp=gran_total_comp,
        gran_total_apro=gran_total_apro)


CLIENTES_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Viáticos a cobrar — V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#6c3483;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#d7bde2;text-decoration:none;font-size:13px}
.topbar a:hover{color:white}
.container{max-width:1100px;margin:24px auto;padding:0 16px}
h1{font-size:22px;color:#6c3483;margin-bottom:4px}
.subtitle{color:#777;font-size:13px;margin-bottom:20px}

.grand-bar{display:flex;gap:32px;flex-wrap:wrap;background:white;border-radius:10px;
           padding:18px 24px;margin-bottom:24px;box-shadow:0 2px 8px rgba(0,0,0,.07)}
.gi .val{font-size:24px;font-weight:700}
.gi .lbl{font-size:11px;color:#aaa;text-transform:uppercase;letter-spacing:.4px;margin-top:3px}

.client-card{background:white;border-radius:10px;
             box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px;overflow:hidden}
.client-header{background:#6c3483;color:white;padding:11px 18px;
               display:flex;justify-content:space-between;align-items:center}
.client-header h2{font-size:15px;font-weight:600}
.big-amount{font-size:22px;font-weight:700}

table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#f5eef8;color:#6c3483;padding:8px 12px;text-align:left;
         border-bottom:2px solid #d7bde2;font-weight:600}
th.num{text-align:right}
tbody tr:nth-child(even){background:#fdf4ff}
tbody tr:hover{background:#f5eef8}
td{padding:7px 12px;border-bottom:1px solid #eee}
td.num{text-align:right;font-family:monospace}
.badge-status{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700}
.bs-pend{background:#fef9e7;color:#7d6608;border:1px solid #f0c040}
.bs-apro{background:#d5f5e3;color:#155724}
.bs-parc{background:#fde8cc;color:#7d4e00}
.bs-rech{background:#fde8e8;color:#c0392b}
.totals-row td{font-weight:700;background:#f5eef8;border-top:2px solid #d7bde2}
.empty{padding:20px;color:#aaa;text-align:center;font-style:italic}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700;font-size:16px">🏢 Viáticos a cobrar por cliente — V&V</span>
  <div style="display:flex;gap:16px">
    <a href="/resumen_viaticos">👤 Resumen por empleado</a>
    <a href="/dashboard">⚙ Panel Admin</a>
    <a href="/logout">Salir</a>
  </div>
</div>

<div class="container">
  <h1>Viáticos a cobrar a clientes</h1>
  <p class="subtitle">Importe total que debe cobrarse a cada cliente por los viajes realizados en su favor.</p>

  <!-- Totales globales -->
  <div class="grand-bar">
    <div class="gi">
      <div class="val" style="color:#6c3483">${{ "%.2f"|format(gran_total_cli) }}</div>
      <div class="lbl">Total a cobrar (tarifa cliente)</div>
    </div>
    <div class="gi">
      <div class="val" style="color:#27ae60">${{ "%.2f"|format(gran_total_apro) }}</div>
      <div class="lbl">Total aprobado (reembolso empleados)</div>
    </div>
    <div class="gi">
      <div class="val" style="color:#2980b9">${{ "%.2f"|format(gran_total_comp) }}</div>
      <div class="lbl">Total comprobado (solicitado)</div>
    </div>
    <div class="gi">
      <div class="val" style="color:#e67e22">${{ "%.2f"|format(gran_total_cli - gran_total_apro) }}</div>
      <div class="lbl">Diferencia (utilidad en viáticos)</div>
    </div>
  </div>

  {% if not clientes %}
  <div class="empty">No hay viajes comprobados aún.</div>
  {% endif %}

  {% for cliente_safe, datos in clientes %}
  <div class="client-card">
    <div class="client-header">
      <h2>🏢 {{ cliente_safe.replace('_',' ') }}</h2>
      <div>
        <div class="big-amount">${{ "%.2f"|format(datos.total_cliente) }}</div>
        <div style="font-size:11px;opacity:.8">a cobrar (tarifa cliente)</div>
      </div>
    </div>
    <table>
      <thead>
        <tr>
          <th>Viaje</th>
          <th>Empleado</th>
          <th class="num">Comprobantes</th>
          <th class="num">Comprobado</th>
          <th class="num">Aprobado</th>
          <th class="num">A cobrar</th>
          <th>Estado aprobación</th>
        </tr>
      </thead>
      <tbody>
      {% for v in datos.viajes %}
      <tr>
        <td>{{ v.viaje_key.replace('_',' ') }}</td>
        <td>{{ v.nombre }}</td>
        <td class="num">{{ v.num_archivos }}</td>
        <td class="num">${{ "%.2f"|format(v.comprobado) }}</td>
        <td class="num" style="color:#27ae60">${{ "%.2f"|format(v.aprobado) }}</td>
        <td class="num" style="color:#6c3483;font-weight:700">${{ "%.2f"|format(v.total_cliente) }}</td>
        <td>
          {% if v.status_apro == 'aprobado' %}<span class="badge-status bs-apro">✔ Aprobado</span>
          {% elif v.status_apro == 'aprobado_parcial' %}<span class="badge-status bs-parc">⚡ Parcial</span>
          {% elif v.status_apro == 'rechazado' %}<span class="badge-status bs-rech">✗ Rechazado</span>
          {% else %}<span class="badge-status bs-pend">⏳ Pendiente</span>{% endif %}
        </td>
      </tr>
      {% endfor %}
      <tr class="totals-row">
        <td colspan="3" style="text-align:right;color:#555;font-size:12px">Total {{ cliente_safe.replace('_',' ') }}</td>
        <td class="num">${{ "%.2f"|format(datos.total_comprobado) }}</td>
        <td class="num" style="color:#27ae60">${{ "%.2f"|format(datos.total_aprobado) }}</td>
        <td class="num" style="color:#6c3483">${{ "%.2f"|format(datos.total_cliente) }}</td>
        <td></td>
      </tr>
      </tbody>
    </table>
  </div>
  {% endfor %}
</div>
</body></html>"""


# ─── VER COMPROBANTE ──────────────────────────────────────────────────────────

@app.route("/comprobantes_viaje")
@login_required
def comprobantes_viaje():
    """Admin/socia: lista y permite ver/descargar todos los comprobantes de un viaje."""
    if session.get("rol") not in ("admin", "socia"):
        return "Sin permiso", 403
    usuario   = request.args.get("usuario", "")
    viaje_key = nfc(request.args.get("viaje_key", ""))
    cliente   = nfc(request.args.get("cliente", ""))
    if not usuario or not viaje_key:
        return "Parámetros incompletos.", 400

    users = load_users()
    nombre_emp = users.get(usuario, {}).get("nombre", usuario)

    # Localizar directorio del viaje
    trip_dir = None
    user_dir = os.path.join(UPLOADS_DIR, safe_name(usuario))
    if os.path.isdir(user_dir):
        for cf in os.listdir(user_dir):
            cand = os.path.join(user_dir, cf, safe_name(viaje_key))
            if os.path.isdir(cand):
                trip_dir = cand
                if not cliente:
                    cliente = cf
                break
    if not trip_dir:
        return "Viaje no encontrado.", 404

    # Listar archivos comprobantes (excluir JSONs internos)
    SKIP = {"_gastos_calculados.json","_config_viaje.json",
            "_aprobacion.json","_overrides.json"}
    archivos = []
    for fn in sorted(os.listdir(trip_dir)):
        if fn in SKIP or fn.startswith("Comprobación") or fn.startswith("Anticipo"):
            continue
        fp = os.path.join(trip_dir, fn)
        if os.path.isfile(fp):
            ext = os.path.splitext(fn)[1].lower()
            size_kb = round(os.path.getsize(fp) / 1024, 1)
            archivos.append({"nombre": fn, "ext": ext, "size_kb": size_kb})

    html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Comprobantes — {viaje_key.replace('_',' ')}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}}
.topbar{{background:#1a5276;color:white;padding:12px 24px;display:flex;
         justify-content:space-between;align-items:center}}
.topbar a{{color:#aed6f1;text-decoration:none;font-size:13px}}
.container{{max-width:900px;margin:28px auto;padding:0 16px}}
h1{{font-size:20px;color:#1a5276;margin-bottom:4px}}
.subtitle{{color:#888;font-size:13px;margin-bottom:20px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:16px;margin-top:16px}}
.card{{background:white;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);
       overflow:hidden;display:flex;flex-direction:column}}
.card-thumb{{background:#ebf5fb;height:140px;display:flex;align-items:center;
             justify-content:center;font-size:48px;cursor:pointer}}
.card-thumb img{{max-width:100%;max-height:140px;object-fit:contain}}
.card-body{{padding:12px;flex:1}}
.card-name{{font-size:12px;font-weight:600;word-break:break-all;margin-bottom:4px}}
.card-meta{{font-size:11px;color:#888}}
.card-actions{{padding:8px 12px;display:flex;gap:8px;border-top:1px solid #eee}}
.btn{{padding:5px 12px;border-radius:5px;border:none;cursor:pointer;font-size:12px;
      font-weight:600;text-decoration:none;display:inline-block}}
.btn-view{{background:#2980b9;color:white}}
.btn-dl{{background:#27ae60;color:white}}
</style></head>
<body>
<div class="topbar">
  <span style="font-weight:700">📎 Comprobantes — {nombre_emp} / {viaje_key.replace('_',' ')}</span>
  <a href="/resumen_viaticos">← Volver al resumen</a>
</div>
<div class="container">
  <h1>{viaje_key.replace('_',' ')}</h1>
  <p class="subtitle">Colaborador: <strong>{nombre_emp}</strong> &nbsp;|&nbsp;
     Cliente: <strong>{cliente}</strong> &nbsp;|&nbsp;
     {len(archivos)} comprobante(s)</p>

  {'<p style="color:#888;margin-top:20px">No se encontraron comprobantes subidos para este viaje.</p>' if not archivos else ''}
  <div class="grid">
"""
    ICONS = {".pdf":"📄",".jpg":"🖼",".jpeg":"🖼",".png":"🖼",
             ".xml":"📋",".xlsx":"📊",".xls":"📊"}
    for a in archivos:
        icon = ICONS.get(a["ext"], "📎")
        view_url = f"/ver_comprobante/{usuario}/{cliente}/{viaje_key}/{a['nombre']}"
        dl_url   = f"/descargar_comprobante/{usuario}/{cliente}/{viaje_key}/{a['nombre']}"
        thumb = f'<img src="{view_url}" onerror="this.parentNode.innerHTML=\'{icon}\'">' \
                if a["ext"] in (".jpg",".jpeg",".png") else icon
        html += f"""
    <div class="card">
      <div class="card-thumb" onclick="window.open('{view_url}','_blank')">{thumb}</div>
      <div class="card-body">
        <div class="card-name">{a['nombre']}</div>
        <div class="card-meta">{a['ext'].upper()[1:]} &nbsp;·&nbsp; {a['size_kb']} KB</div>
      </div>
      <div class="card-actions">
        <a href="{view_url}" target="_blank" class="btn btn-view">👁 Ver</a>
        <a href="{dl_url}" class="btn btn-dl">⬇ Descargar</a>
      </div>
    </div>"""

    html += "\n  </div>\n</div>\n</body></html>"
    return html


@app.route("/descargar_comprobante/<usuario>/<cliente_safe>/<viaje_key>/<archivo>")
@login_required
def descargar_comprobante_archivo(usuario, cliente_safe, viaje_key, archivo):
    if session.get("rol") not in ("admin", "socia"):
        return "Sin permiso", 403
    trip_dir  = os.path.join(UPLOADS_DIR, safe_name(usuario),
                             safe_name(cliente_safe), safe_name(viaje_key))
    file_path = os.path.join(trip_dir, archivo)
    if not os.path.isfile(file_path):
        return "Archivo no encontrado.", 404
    return send_file(file_path, as_attachment=True, download_name=archivo)


@app.route("/ver_comprobante/<usuario>/<cliente_safe>/<viaje_key>/<archivo>")
@login_required
def ver_comprobante(usuario, cliente_safe, viaje_key, archivo):
    rol = session.get("rol")
    usr_session = session.get("usuario")
    # Trabajador solo puede ver sus propios archivos
    if rol == "trabajador" and usr_session != usuario:
        return "Acceso denegado.", 403
    trip_dir = os.path.join(UPLOADS_DIR, safe_name(usuario),
                            safe_name(cliente_safe), safe_name(viaje_key))
    file_path = os.path.join(trip_dir, archivo)
    if not os.path.isfile(file_path):
        return "Archivo no encontrado.", 404
    return send_file(file_path, as_attachment=False)


# ─── CONFIGURACIÓN DE EMAIL ───────────────────────────────────────────────────

EMAIL_CONFIG_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Configuración de correo — V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#adc8e6;text-decoration:none;font-size:13px}
.container{max-width:600px;margin:28px auto;padding:0 16px}
h1{font-size:20px;color:#1a3a5c;margin-bottom:4px}
.card{background:white;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);padding:24px;margin-bottom:20px}
label{display:block;font-size:12px;font-weight:600;color:#555;margin-bottom:4px;margin-top:14px}
input[type=text],input[type=password],input[type=number]{width:100%;padding:9px;border:1px solid #ccc;
  border-radius:6px;font-size:13px}
.btn{padding:9px 20px;border-radius:6px;border:none;cursor:pointer;font-size:13px;font-weight:600;
     text-decoration:none;display:inline-block;margin-top:16px;margin-right:8px}
.btn-primary{background:#1a3a5c;color:white}
.btn-secondary{background:#95a5a6;color:white}
.alert{padding:10px 14px;border-radius:6px;font-size:13px;margin-bottom:14px}
.alert-ok{background:#d1e7dd;color:#0f5132}
.alert-err{background:#f8d7da;color:#842029}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700">⚙ Configuración de correo</span>
  <a href="/dashboard">← Dashboard</a>
</div>
<div class="container">
  <h1>Configuración SMTP</h1>
  {% if msg %}
  <div class="alert {{ 'alert-ok' if ok else 'alert-err' }}">{{ msg }}</div>
  {% endif %}
  <div class="card">
    <form method="POST">
      <label>Servidor SMTP</label>
      <input type="text" name="smtp_server" value="{{ cfg.smtp_server or 'smtp.office365.com' }}" required>
      <label>Puerto SMTP</label>
      <input type="number" name="smtp_port" value="{{ cfg.smtp_port or 587 }}" required>
      <label>Usuario SMTP (correo de envío)</label>
      <input type="text" name="smtp_user" value="{{ cfg.smtp_user or '' }}" required>
      <label>Contraseña SMTP</label>
      <input type="password" name="smtp_password" placeholder="Dejar en blanco para no cambiar">
      <label>Correo remitente (from_email)</label>
      <input type="text" name="from_email" value="{{ cfg.from_email or '' }}" required>
      <label>Nombre remitente</label>
      <input type="text" name="from_name" value="{{ cfg.from_name or 'Viáticos V&V' }}">
      <br>
      <button type="submit" name="accion" value="guardar" class="btn btn-primary">💾 Guardar configuración</button>
      <button type="submit" name="accion" value="probar" class="btn btn-secondary">📧 Enviar correo de prueba</button>
    </form>
  </div>
</div>
</body></html>"""


@app.route("/config_email", methods=["GET","POST"])
@login_required
def config_email():
    if session.get("rol") != "admin":
        return redirect(url_for("dashboard"))
    cfg = load_email_config()
    msg = None
    ok = False
    if request.method == "POST":
        accion = request.form.get("accion", "guardar")
        new_cfg = {
            "smtp_server":   request.form.get("smtp_server", "").strip(),
            "smtp_port":     int(request.form.get("smtp_port", 587) or 587),
            "smtp_user":     request.form.get("smtp_user", "").strip(),
            "smtp_password": request.form.get("smtp_password", "").strip() or cfg.get("smtp_password", ""),
            "from_email":    request.form.get("from_email", "").strip(),
            "from_name":     request.form.get("from_name", "Viáticos V&V").strip(),
        }
        with open(EMAIL_CONFIG_F, "w", encoding="utf-8") as f:
            json.dump(new_cfg, f, ensure_ascii=False, indent=2)
        cfg = new_cfg
        if accion == "probar":
            users = load_users()
            admin_email = users.get(session["usuario"], {}).get("email", "")
            resultado = send_email(
                [admin_email] if admin_email else [new_cfg["from_email"]],
                "Correo de prueba — Sistema de Viáticos V&V",
                "<html><body><h2>¡Funciona!</h2><p>La configuración de correo es correcta.</p></body></html>"
            )
            msg = "Correo de prueba enviado correctamente." if resultado else "Error al enviar correo de prueba. Revisa la consola."
            ok = resultado
        else:
            msg = "Configuración guardada."
            ok = True
    return render_template_string(EMAIL_CONFIG_HTML, cfg=cfg, msg=msg, ok=ok)


# ─── PUNTO DE ENTRADA ─────────────────────────────────────────────────────────

def get_local_ip():
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

@app.route("/reportes")
@login_required
def reportes():
    if session.get("rol") != "admin":
        return redirect(url_for("resumen_viaticos"))

    users = load_users()
    # Listas para filtros
    empleados_lista = [(u, d.get("nombre", u)) for u, d in users.items()
                       if d.get("rol") != "admin"]
    # Recopilar clientes únicos
    clientes_set = set()
    for usr in [u for u, d in users.items() if d.get("rol") != "admin"]:
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
        if os.path.isdir(user_dir):
            for cs in os.listdir(user_dir):
                if os.path.isdir(os.path.join(user_dir, cs)):
                    clientes_set.add(cs)
    clientes_lista = sorted(clientes_set)

    # Leer filtros GET
    date_from_s = request.args.get("date_from", "")
    date_to_s   = request.args.get("date_to",   "")
    fil_usuario = request.args.get("usuario",   "")
    fil_cliente = request.args.get("cliente",   "")
    formato     = request.args.get("formato",   "")

    def parse_dt(s):
        try: return datetime.strptime(s, "%Y-%m-%d")
        except: return None

    dt_from = parse_dt(date_from_s)
    dt_to   = parse_dt(date_to_s)

    filas_col = []  # Por colaborador
    filas_cli = []  # Por cliente

    for usr, dat in users.items():
        if dat.get("rol") == "admin":
            continue
        if fil_usuario and usr != fil_usuario:
            continue
        nombre = dat.get("nombre", usr)
        user_dir = os.path.join(UPLOADS_DIR, safe_name(usr))
        if not os.path.isdir(user_dir):
            continue
        for cliente_safe in sorted(os.listdir(user_dir)):
            if fil_cliente and cliente_safe != fil_cliente:
                continue
            cliente_dir = os.path.join(user_dir, cliente_safe)
            if not os.path.isdir(cliente_dir):
                continue
            for viaje_key in sorted(os.listdir(cliente_dir)):
                trip_dir_r = os.path.join(cliente_dir, viaje_key)
                if not os.path.isdir(trip_dir_r):
                    continue
                archivos_r = [f for f in os.listdir(trip_dir_r)
                              if f.lower().endswith((".pdf",".jpg",".jpeg",".png"))]
                if not archivos_r:
                    continue
                # Extraer fecha desde viaje_key (últimos 8 chars = DDMMAAAA)
                fecha_viaje = None
                vk_tail = viaje_key[-8:] if len(viaje_key) >= 8 else ""
                try:
                    fecha_viaje = datetime.strptime(vk_tail, "%d%m%Y")
                except Exception:
                    pass
                # Filtrar por fecha
                if dt_from and fecha_viaje and fecha_viaje < dt_from:
                    continue
                if dt_to and fecha_viaje and fecha_viaje > dt_to:
                    continue
                fecha_str = fecha_viaje.strftime("%d/%m/%Y") if fecha_viaje else "—"
                gastos_r, total_emp_r, total_cli_r, apro_r = \
                    _gastos_para_viaje(usr, cliente_safe, viaje_key)
                viaje_status_r = apro_r.get("status", "pendiente")
                total_aprobado_r = sum(
                    g.get("aprobado", 0) for g in gastos_r
                    if g.get("status_apro") in ("aprobado", "parcial")
                ) if viaje_status_r != "pendiente" else 0.0
                saldo_r = round(total_emp_r - total_aprobado_r, 2)
                filas_col.append({
                    "empleado":   nombre,
                    "viaje":      viaje_key.replace("_", " "),
                    "fecha":      fecha_str,
                    "comprobado": round(total_emp_r, 2),
                    "aprobado":   round(total_aprobado_r, 2),
                    "saldo":      saldo_r,
                })
                filas_cli.append({
                    "cliente":   cliente_safe.replace("_", " "),
                    "viaje":     viaje_key.replace("_", " "),
                    "empleado":  nombre,
                    "a_cobrar":  round(total_cli_r, 2),
                    "aprobado":  round(total_aprobado_r, 2),
                })

    if formato == "excel":
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Por Colaborador"
        ws1.append(["Empleado", "Viaje", "Fecha", "Comprobado", "Aprobado", "Saldo"])
        for r in filas_col:
            ws1.append([r["empleado"], r["viaje"], r["fecha"],
                        r["comprobado"], r["aprobado"], r["saldo"]])
        ws2 = wb.create_sheet("Por Cliente")
        ws2.append(["Cliente", "Viaje", "Empleado", "A cobrar cliente", "Aprobado"])
        for r in filas_cli:
            ws2.append([r["cliente"], r["viaje"], r["empleado"],
                        r["a_cobrar"], r["aprobado"]])
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name="reporte_viaticos.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    REPORTES_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Reportes — V&amp;V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f6f9;color:#333;font-size:14px}
.topbar{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
        justify-content:space-between;align-items:center}
.topbar a{color:#adc8e6;text-decoration:none;font-size:13px}
.container{max-width:1100px;margin:24px auto;padding:0 16px}
h1{font-size:20px;color:#1a3a5c;margin-bottom:12px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);
      margin-bottom:20px;overflow:hidden}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;font-weight:600;font-size:13px}
.card-body{padding:16px}
.filter-row{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end}
.filter-row label{font-size:12px;font-weight:600;color:#555;display:block;margin-bottom:3px}
.filter-row input,.filter-row select{padding:7px 10px;border:1px solid #ccc;border-radius:5px;font-size:13px}
.btn{padding:8px 18px;border:none;border-radius:6px;cursor:pointer;
     font-size:13px;font-weight:600;text-decoration:none;display:inline-block}
.btn-primary{background:#1a3a5c;color:white}
.btn-success{background:#27ae60;color:white}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{background:#ebf5fb;color:#1a3a5c;padding:7px 10px;text-align:left;
         border-bottom:2px solid #d6eaf8}
td{padding:7px 10px;border-bottom:1px solid #f0f0f0}
.num{text-align:right}
tr:hover td{background:#f8fbff}
h2{font-size:15px;color:#1a5276;margin-bottom:10px}
</style>
</head>
<body>
<div class="topbar">
  <span style="font-weight:700">📊 Reportes de Viáticos</span>
  <div style="display:flex;gap:16px">
    <a href="/dashboard">← Dashboard</a>
    <a href="/logout">Salir</a>
  </div>
</div>
<div class="container">
<h1>Reportes de Viáticos</h1>
<div class="card">
  <div class="card-header">Filtros</div>
  <div class="card-body">
    <form method="GET" action="/reportes">
      <div class="filter-row">
        <div>
          <label>Fecha desde</label>
          <input type="date" name="date_from" value="{{ date_from_s }}">
        </div>
        <div>
          <label>Fecha hasta</label>
          <input type="date" name="date_to" value="{{ date_to_s }}">
        </div>
        <div>
          <label>Empleado</label>
          <select name="usuario">
            <option value="">— Todos —</option>
            {% for u, n in empleados_lista %}
            <option value="{{ u }}" {% if u == fil_usuario %}selected{% endif %}>{{ n }}</option>
            {% endfor %}
          </select>
        </div>
        <div>
          <label>Cliente</label>
          <select name="cliente">
            <option value="">— Todos —</option>
            {% for c in clientes_lista %}
            <option value="{{ c }}" {% if c == fil_cliente %}selected{% endif %}>{{ c }}</option>
            {% endfor %}
          </select>
        </div>
        <div style="display:flex;gap:8px;align-items:flex-end">
          <button type="submit" class="btn btn-primary">🔍 Generar reporte</button>
          <a href="/reportes?{{ request.query_string.decode() }}&formato=excel" class="btn btn-success">📥 Exportar Excel</a>
        </div>
      </div>
    </form>
  </div>
</div>

<div class="card">
  <div class="card-header">Por colaborador ({{ filas_col|length }} registros)</div>
  <div class="card-body" style="padding:0">
    {% if filas_col %}
    <table>
      <thead>
        <tr>
          <th>Empleado</th><th>Viaje</th><th>Fecha</th>
          <th class="num">Comprobado</th><th class="num">Aprobado</th><th class="num">Saldo</th>
        </tr>
      </thead>
      <tbody>
      {% for r in filas_col %}
      <tr>
        <td>{{ r.empleado }}</td>
        <td>{{ r.viaje }}</td>
        <td>{{ r.fecha }}</td>
        <td class="num">${{ "%.2f"|format(r.comprobado) }}</td>
        <td class="num">${{ "%.2f"|format(r.aprobado) }}</td>
        <td class="num" style="color:{% if r.saldo > 0 %}#a04000{% elif r.saldo < 0 %}#0a5c36{% else %}#555{% endif %}">
          ${{ "%.2f"|format(r.saldo) }}
        </td>
      </tr>
      {% endfor %}
      <tr style="font-weight:bold;background:#ebf5fb">
        <td colspan="3">Total</td>
        <td class="num">${{ "%.2f"|format(filas_col|sum(attribute='comprobado')) }}</td>
        <td class="num">${{ "%.2f"|format(filas_col|sum(attribute='aprobado')) }}</td>
        <td class="num">${{ "%.2f"|format(filas_col|sum(attribute='saldo')) }}</td>
      </tr>
      </tbody>
    </table>
    {% else %}
    <div style="padding:24px;text-align:center;color:#aaa">Sin registros para los filtros seleccionados.</div>
    {% endif %}
  </div>
</div>

<div class="card">
  <div class="card-header">Por cliente ({{ filas_cli|length }} registros)</div>
  <div class="card-body" style="padding:0">
    {% if filas_cli %}
    <table>
      <thead>
        <tr>
          <th>Cliente</th><th>Viaje</th><th>Empleado</th>
          <th class="num">A cobrar cliente</th><th class="num">Aprobado</th>
        </tr>
      </thead>
      <tbody>
      {% for r in filas_cli %}
      <tr>
        <td>{{ r.cliente }}</td>
        <td>{{ r.viaje }}</td>
        <td>{{ r.empleado }}</td>
        <td class="num">${{ "%.2f"|format(r.a_cobrar) }}</td>
        <td class="num">${{ "%.2f"|format(r.aprobado) }}</td>
      </tr>
      {% endfor %}
      <tr style="font-weight:bold;background:#ebf5fb">
        <td colspan="3">Total</td>
        <td class="num">${{ "%.2f"|format(filas_cli|sum(attribute='a_cobrar')) }}</td>
        <td class="num">${{ "%.2f"|format(filas_cli|sum(attribute='aprobado')) }}</td>
      </tr>
      </tbody>
    </table>
    {% else %}
    <div style="padding:24px;text-align:center;color:#aaa">Sin registros para los filtros seleccionados.</div>
    {% endif %}
  </div>
</div>
</div>
</body></html>"""
    return render_template_string(REPORTES_HTML,
        empleados_lista=empleados_lista, clientes_lista=clientes_lista,
        date_from_s=date_from_s, date_to_s=date_to_s,
        fil_usuario=fil_usuario, fil_cliente=fil_cliente,
        filas_col=filas_col, filas_cli=filas_cli,
        request=request)


@app.route("/admin/casetas", methods=["GET","POST"])
@login_required
def admin_casetas():
    if session.get("rol") != "admin":
        return redirect(url_for("dashboard"))
    casetas = load_casetas()
    msg = None
    if request.method == "POST":
        accion = request.form.get("accion","")
        if accion == "guardar":
            nueva = {}
            for key in casetas:
                val_str = request.form.get(f"c_{key}", "").strip()
                try:
                    nueva[key] = round(float(val_str), 2)
                except Exception:
                    nueva[key] = casetas[key]
            # Ciudad nueva
            new_city = request.form.get("new_city","").strip().upper()
            new_monto = request.form.get("new_monto","").strip()
            if new_city and new_monto:
                try:
                    nueva[new_city] = round(float(new_monto), 2)
                except Exception:
                    pass
            save_casetas(nueva)
            casetas = nueva
            msg = "✅ Casetas guardadas correctamente."
        elif accion == "eliminar":
            city_del = request.form.get("city_del","").strip().upper()
            if city_del in casetas:
                del casetas[city_del]
                save_casetas(casetas)
                msg = f"✅ '{city_del}' eliminado."
    casetas_sorted = dict(sorted(casetas.items()))
    return render_template_string("""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><title>Casetas — V&V</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f4f6f9;font-size:14px;color:#333}
header{background:#1a3a5c;color:white;padding:12px 24px;display:flex;
       justify-content:space-between;align-items:center;font-weight:bold}
header a{color:#adc8e6;text-decoration:none;font-size:13px;margin-left:16px}
.container{max-width:700px;margin:24px auto;padding:0 16px 40px}
.card{background:white;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:16px}
.card-header{background:#1a3a5c;color:white;padding:10px 16px;font-weight:bold;font-size:13px;border-radius:8px 8px 0 0}
.card-body{padding:18px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#ecf0f1;padding:8px 10px;text-align:left;border-bottom:2px solid #dee2e6}
td{padding:7px 10px;border-bottom:1px solid #f0f0f0}
input[type=number],input[type=text]{padding:5px 8px;border:1px solid #ccc;border-radius:4px;font-size:13px;width:100%}
.btn{padding:7px 16px;border:none;border-radius:5px;cursor:pointer;font-size:13px;font-weight:bold}
.btn-primary{background:#1a3a5c;color:white}.btn-danger{background:#c0392b;color:white}
.btn:hover{opacity:.88}
.msg{padding:10px;border-radius:6px;background:#d1e7dd;color:#0f5132;margin-bottom:14px;font-size:13px}
.note{font-size:11px;color:#777;margin-top:6px}
</style></head><body>
<header>🛣 Gestión de Casetas (CAPUFE)
  <div><a href="/dashboard">← Dashboard</a></div>
</header>
<div class="container">
{% if msg %}<div class="msg">{{ msg }}</div>{% endif %}
<div class="card">
  <div class="card-header">Costos de casetas por ciudad — ida y vuelta desde CDMX</div>
  <div class="card-body">
    <p class="note" style="margin-bottom:14px">Montos predeterminados basados en tarifas CAPUFE. Edite los valores según sea necesario. El sistema usará estos montos al calcular el anticipo cuando el modo de transporte es automóvil.</p>
    <form method="POST">
      <input type="hidden" name="accion" value="guardar">
      <table>
        <thead><tr><th>Ciudad</th><th style="width:160px">Monto ($ ida+vuelta)</th><th style="width:80px"></th></tr></thead>
        <tbody>
        {% for city, monto in casetas.items() %}
        <tr>
          <td>{{ city }}</td>
          <td><input type="number" name="c_{{ city }}" value="{{ '%.2f'|format(monto) }}" step="0.01" min="0"></td>
          <td>
            <form method="POST" style="display:inline">
              <input type="hidden" name="accion" value="eliminar">
              <input type="hidden" name="city_del" value="{{ city }}">
              <button class="btn btn-danger" style="padding:4px 10px;font-size:12px"
                      onclick="return confirm('¿Eliminar {{ city }}?')">✕</button>
            </form>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      <div style="margin-top:14px">
        <button type="submit" class="btn btn-primary">💾 Guardar cambios</button>
      </div>
    </form>
    <hr style="margin:18px 0;border-color:#eee">
    <form method="POST">
      <input type="hidden" name="accion" value="guardar">
      <div style="display:flex;gap:10px;align-items:flex-end;flex-wrap:wrap">
        <div style="flex:1;min-width:180px">
          <label style="font-size:12px;font-weight:bold;display:block;margin-bottom:4px">Nueva ciudad</label>
          <input type="text" name="new_city" placeholder="ej. HERMOSILLO" style="padding:7px">
        </div>
        <div style="width:160px">
          <label style="font-size:12px;font-weight:bold;display:block;margin-bottom:4px">Monto ($ ida+vuelta)</label>
          <input type="number" name="new_monto" step="0.01" min="0" placeholder="0.00">
        </div>
        <button type="submit" class="btn btn-primary" style="height:34px">➕ Agregar</button>
      </div>
      {% for city, monto in casetas.items() %}
      <input type="hidden" name="c_{{ city }}" value="{{ '%.2f'|format(monto) }}">
      {% endfor %}
    </form>
  </div>
</div>
</div></body></html>""", casetas=casetas_sorted, msg=msg)


def open_browser():
    import time; time.sleep(1.2)
    webbrowser.open("http://127.0.0.1:5050")

if __name__ == "__main__":
    ip = get_local_ip()
    threading.Thread(target=open_browser, daemon=True).start()
    print("="*60)
    print("  Viáticos V&V — Villar & Villar Abogados")
    print(f"  Acceso local:  http://127.0.0.1:5050")
    print(f"  Acceso en red: http://{ip}:5050  ← compartir este link")
    print("  Para cerrar:   Ctrl+C")
    print("="*60)
    print("\n  Usuarios creados:")
    for u,d in load_users().items():
        print(f"    {u:12} → {d['nombre']} ({d['rol']})")
    print()
    PORT = int(os.environ.get('PORT', 5050))
    app.run(host="0.0.0.0", port=PORT, debug=False)
